# -*- coding:utf-8 -*-

#千锋老师电话：13691511443

#【day】安装pycharm\webstrom\datagrip
#双击pycharm-professional-2019.1.3 or WebStorm-2019.1.2 or datagrip-2019.1.4
#   Do not import settings
#       ok
#   Next
#   start using webstorm
#   evaluate for free
#       evaluate
#   create new project
#   help
#       edit custom vm options..
#           create
#               -javaagent:F:\installed\PyCharm 2019.1.3\bin\jetbrains-agent.jar
#               重启webstrom
#                   help
#                       register
#                           license sever
#                               discover server
#                               activate
#【day1】注释(已授)
#
'''
注释
'''


#【day2】print()打印(字符串，整数，浮点数，运算, 变量，布尔值)(已授)
'''
print('陈紫妍很漂亮')
print('陈紫妍很漂亮' , '陈迪希很漂亮')
print('陈紫妍的年龄是' , 9)
print(123)
print(123.123)
print(18+2)
print('18+2')
'''


#【day3】input()作用：从外部获取变量的值(已授)
'''
age = input() #等待输入('值'),输入的内容可以保存在age里
print('陈紫妍的年龄是', age)
'''
'''
age = input('请输入您的年龄:')
print('陈紫妍的年龄是', age)
'''
'''
age = input('用户输入的所有内容都会是字符串')
'''
'''
#欢迎界面
print('欢迎来到大鱼吃小鱼')
name = input('请输入玩家的名字:')
age = input('请输入玩家的年龄：')
height = input('请输入玩家的身高：')
print('玩家的名字是：', name)
print('玩家的年龄是：', age)
print('玩家的身高是：', height)
'''


#【day4】标识符(已授)
#规则：
#   只能由字母、数字、下划线组成
#   开头不能是数字
#   不能是python的关键字(已经被定义了功能的字符)
#       import keyword print(keyword.kwlist)
#           ['False', 'None', 'True', 'and', 'as', 'assert', 'async', 'await', 'break', 'class', 'continue', 'def', 'del', 'elif', 'else', 'except', 'finally', 'for', 'from', 'global', 'if', 'import', 'in', 'is', 'lambda', 'nonlocal', 'not', 'or', 'pass', 'raise', 'return', 'try', 'while', 'with', 'yield']
#   区分大小写
#   见名知意思
#   驼峰原则
#        首单词小写，从第二个单词开始，首字母大写
#        cylIsGoodman
#作用：
#   给变量、函数等命名的

'''
作业判断一下是否为标识符
adb
12cdb
False
_abc
abc123
___nameAge
_abc_123
%$12ab
123abc
a_c_1
'''

#【day5】创建变量、删除变量
#变量概述
#   程序可操作的存储空间的名称 or 程序运行期间能改变的数据 or 每个变量都有特定的类型
#作用：
#   将不同类型的数据存储到内存
#定义变量
#   变量名 = 初始值
#       初始值：为了确定变量的数据类型
#   age = 10
#数据的存储
#   变量 = 数据值
#注意：在使用变量前必须定义，否则会报错

'''
age = 18
print('age = ' , age)
'''
'''
del age
print(age) #nameerror:age is not definded
'''
'''
name = 'mia'
age = 9
print(name)

name = 'zengxiaoxuan'
age = 10
print(name)
'''

'''
name = 'mia'
del name
print(name)
'''


#【day6】连续定义变量
'''
num1 = 1
num2 = num1
num3 = num2
'''
'''
num1 = num2 = num3 = 1
print(num1, num2, num3)
'''
'''
money1 = 100
money2 = 100
money3 = 100
'''
#当赋的值相同，又需要赋值给多个变量时，需要用到连续定义变量
#格式：变量名1 = 变量名2 = 变量名3 = ..... = 初始值
'''
money1 = money2 = money3 = 100
print(money1, money2, money3)
'''

#【day7】交互式定义变量
'''
num1, num2 = 1, 2
print(num1, num2)
'''
#当赋值不同时，又需要赋值给多个变量时，需要用到交互式定义变量
#格式：变量名1， 变量名2 = 初始值1， 初始值2
'''
money1, money2 = 100, 1000
print(money1, money2)
'''



#【day8】数据类型
#Python数据类型
#   number(数字)
#        整数、浮点数(小数)、复数
#    string(字符串)
#    boolean(布尔值)
#        True,False
#    None(空值)
#    list(列表)
#    tuple(元组)
#    dict(字典)
#    set(集合)

'''
age = 1
age = 1.1
age = '年龄'
age = True
age = None
age = [1,2,3]
age = (1,2,3)
age = {'one':'a', 'two':'123', 'three':123}
age = {5,6,'abc',5}
'''
'''
#作业分别打印一下不同数据类型
age = 
print(age)
'''

#【day9】tepe() 查看变量类型
#变量会随着初始值的数据类型变化而变化
'''
age = 
age_type = type(age)
print(age_type)
'''
'''
age = 
print(type(age))
'''
'''
作业
num1 = input('请输入一个数字:')
num2 = input('请再输入一个数字:')
print(num1 , type(num1))
print(num2 , type(num2))
print('num1+num2的字符串拼接' , num1+num2)
'''
'''
print(int(num1) , type(int(num1)))
print(int(num2) , type(int(num2)))
print('num1+num2的运算' , int(num1) + int(num2))
'''

#【day10】id() 查看变量在内存中的地址
#内存定义
#内存的房间号id
'''
age = '123'
print(id(age))
'''
'''
讲到函数作用域的时候在详细讲地址
num1 = 10
print(id(num1))
num1 = 20
print(id(num1))
num2 = num1
print(id(num2))
'''
#【day】内置函数
#概念
#    python内置的函数，能够直接调用
#dir
#    功能：打印类的所有属性和方法
#    返回值：这个类所有属性和方法的列表
#    示例1
#       import scrapy
#       print(dir(scrapy)) --> ['Request',...]
#    示例2
#       import random
#       print(dir(random)) --> ['choice',...]

#【day11】常量
#程序运行期间不能改变的数据
'''
age = 18 #18就是常量
'''

#【day12】数字类型
#分类：整数、浮点数、复数
'''
#整数:Python的整数可以处理任意大小的整数，当然包括负整数，在程序中的表示和数学中的写法一样
num1 = 1
print(type(num1))
'''
'''
#浮点数:由整数部分与小数部分组成，浮点数运算可能有四舍五入的误差
float1 = 1.1
float2 = 2.2
print(float1 + float2)
'''
'''
#复数:由实数部分和虚数部分组成
'''

#【day13】数字类型转换
'''
#浮点数转整数,int()只取整数部分
print(int(1.1))
print(int(1.9))
'''
'''
#整数转浮点数，float()
print(float(1))
'''
'''
#字符串转整数
print(int('123'))
'''
'''
#字符串不能是非整数格式
print('abc')
print('123abc')
'''
'''
#当+-只作为正负号时，可以转换
print('+123')
print('-123')
#print('123+123') #报错
'''
'''
#字符串转浮点数
print(float('12.12'))
'''

#【day14】数学功能
'''
#abs()返回参数的绝对值
num1 = -10
num2 = abs(num1)
print(num2)
'''
'''
num1 = -10
print(abs(num1))
'''
'''
#max()求给定参数的最大值
print(max(1,2,3,4,5,6,7,8,9))
'''
'''
num1 = 100
num2 = 99
print(max(num1,num2))
'''
'''
#min()求给定参数的最小值
print(min(1,2,3,4,,6,7,8,9))
'''
'''
#求x的y次方，例2^5
print(pow(2,5))
'''
'''
#round(x[,n])，求x四舍五入，保留到小数点后n位，[]可有可无的参数，默认为0
print(round(3.45))
print(round(3.55))
print(round(3.45 , 1))
print(round(3.55 , 1))
'''

#【day15】导入一个库
#库：封装一些功能
#import math #math是一个数学相关的库
'''
#向上取整
print(math.ceil(18.1))
print(math.ceil(18.9))
'''
'''
#向下取整
print(math.floor(18.1))
print(math.floor(18.9))
'''
'''
#返回整数部分和小数部分
print(math.modf(22.3))
'''
'''
#开方
print(math.sqrt(16))
'''

#【day16】随机数 random.choice()
#import random
'''
在序列（列表、元祖、集合）中随机抽取一个元素
print(random.choice([1,2,3,4,5]))
print(random.choice(range(5))) #range(5) == [0,1,2,3,4]
print(random.choice('cyl')) #'cyl' == ['c','y','l']
'''
'''
#作业1：产生一个0~99的随机数
import random
random_num = random.choice(range(100))
print(random_num)
'''
'''
#作业2：开发一款彩票软件，用户能输入1位数的数字（整数），随机生成一个1位数，如果这个整数等于随机数，则中奖，反之不中
import random
int_num = input('请输入1位数字：')
random_num = random.choice(range(1))
if int_num == random_num:
    print('恭喜中奖')
else:
    print('继续加油')
'''

#【day17】随机数 random.randrange([start,] stop [,step])
#从start到stop之间，按基数递增的集合中，随机取1个随机数
#start:指定范围的开始值，包含在范围内，默认为0
#stop:指定范围的结束只，不包含在范围内
#step:指定基数，默认为1
'''
import random
print(random.randrange(1,100,2)) #在此序列中任取一个：[1,3,5,7...99]
'''

#【day18】随机数 random.random()
#随机生成[0,1)之间的浮点数
'''
import random
print(random.random())
print(random.random()*5)
'''

#【day19】随机排序 random.shuffle()
#将序列的所有元素随机排序
'''
import random
list1 = [1,2,3,4,5]
print(list1)
random.shuffle(list1)
print(list1)
'''

#【day20】random.uniform(start,stop)
#随机生成一个范围为[start,stop]之间的实数（整数或小数）
'''
import random
print(random.uniform(3,9))
'''


#【day21】由变量、常量和运算组成的式子
#阅读表达式：
#   功能：
#       age = 1 age = 1 + 1 num1 , num2= 1 , 2 num3 = num1 + num2
#   值：
#       1 2 1 2 3

#【day22】算术运算符和算术运算表达式
#算术运算符：+ - * / % ** //    加 减 乘 除 取模 求幂 取整
#算术运算表达式 1+1 2*2 a/3
#   功能：进行相关数学符号的计算，不会改变变量的值
#   值：相关的数学运算结果
'''
num1,num2 = 5,3
print(num1 + num2)
print(num1 - num2)
print(num1 * num2)
print(num1 / num2)
print(num1 % num2)
print(num1 ** num2)
print(num1 // num2)
#print打印的是表达式的值
'''

#【day23】赋值运算符和赋值运算表达式
#赋值运算符 =
#赋值运算表达式
#   格式：变量 = 表达式
#       功能：计算了等号右侧的值并赋值给左侧变量
#       值：赋值结束后变量的值
'''
a = b = 1 + 2
print(a , b)
num1 = 3
num2 = num1 + 2
print(num1 , num2)
'''

#【day24】复合运算符
#复合运算符：+= -= *= /= %/ **= //=
#a += b -> a = a + b
#a -= b -> a = a - b
#a *= b -> a = a * b
#a /= b -> a = a / b
#a %= b -> a = a % b
#a **= b -> a = a ** b
#a //= b -> a = a // b

'''
a,b = 1,1
a = a * b
print(a)
a *= b
print(a)
'''
#【day25】位运算符
#a、按位与运算符 &
#二进制数进行按位与运算，相同位都为1，结果为1，否则为0
'''
print(5 & 7)
5(10) -> 101(2)
7(10) -> 111(2)
  101
& 111
------
  101
101(2) -> 5(10)
'''
#b、按位或运算符 |
#二进制数进行按位或运算，相同位相等时，结果为原数，否则为1
'''
print(9 | 8)
  1001
| 1000
-------
  1001
'''
#c、按位异或运算符 ^
#二进制数进行按位异或运算，相同位相异时，结果为1，否则为0
'''
print(9 ^ 8)
  1001
^ 1000
-------
  0001
'''
#d、按位取反运算符 ~
#每个二进制数位取反，0变1,1变0
#？以上结论错误，涉及源码和补码问题
'''
print(~8)
~0000001000(计算机是64位系统，前面1000前有60个0)
------------------------
 1111110111(计算机的补码，需要求原码：符号位不变，其他位取反，最后在加1)
------------------------
 1000001000
+         1
-------------
 1000001001(取后4位求10进制，符号1位负)
'''
#e、左移动运算符 <<
#各二进制位全部左移动若干位，由<<右侧的数字决定，高位丢弃，低位补0
'''
print(2<<2)
2<<0010
--------
   1000
'''
#f、左移动运算符 >>
#各二进制位全部右移动若干位，由>>右侧的数字决定，高位补0，低位丢弃
'''
print(13>>2)
2>>1101
--------
   0011
'''

#【day26】关系运算符和关系运算表达式
#== != > < >= <=
#关系运算表达式：
#格式：表达式1 关系运算符 表达式2
#功能：计算‘表达式1’和‘表达式2’的值
#值：如果关系成立，整个关系运算表达式的值为真True，否则为假False
'''
print(1==1)
print(1!=1)
print(1>2)
print(1<2)
print(1>=1)
print(1>=2)
print(1<=2)
'''

#【day25】if语句
#格式：
#if 表达式：
#   语句
#逻辑：当程序执行到if语句时，首先计算“表达式”的值，如果表达式的值为真，则执行if下面的语句。如果表达式的值为假，则跳过整个if语句继续向下执行。
'''
age = int(input('请输入你的年龄:'))
if age > 18:
    print('恭喜你已经是成年人了')
if age < 18:
    print('对不起你还是未成年人')
'''
'''
#假False 0 0.0 '' None
#真True 
age = 17
print(age > 18)
print(age < 18)
'''
'''
score = int(input('输入你的得分：'))
if score < 60:
    print('你的等级为D')
if score >= 60 and score < 80:
    print('你的等级为C')
if score >= 80 and score < 100:
    print('你的等级为B')
if score == 100:
    print('你的等级为A')
'''

#【day26】if-else语句
#格式：
#if 表达式：
#   语句1
#else:
#   语句2
#逻辑：当程序执行到if-else语句时，首先计算表达式的值，如果表达式的值为真，则执行语句1，执行完语句1会跳出整个if-else语句，如果表达式的值为假，则执行语句2，执行完语句2后跳出整个if-else语句
'''
#从控制台输入一个整数，判断是否为偶数
num = int(input('请输入一个整数：'))
if num % 2 == 0:
    print('输入的是【偶数】')
else:
    print('输入的是【奇数】')
'''

'''
#作业：从控制台输入一个三位数，如果是水仙花数就打印'是水仙花数'，否则打印'不是水仙花数'
#水仙花数是指一个 3 位数，它的每个位上的数字的 3次幂之和等于它本身（例如：1^3 + 5^3+ 3^3 = 153）。
num = int(input('请输入一个整数：'))
num_one = num % 10
num_ten = num // 10 % 10
num_hundred = num // 100
if num == num_one ** 3 + num_ten ** 3 + num_hundred ** 3:
    print('是水仙花数')
else:
    print('不是水仙花数')

#字符串的索引
#a = input('请输入一个三位数:')
#if int(a[0])**3 + int(a[1])**3 + int(a[2])**3 == int(a[0])*100 + int(a[1])*10 + int(a[2]):
#    print('是水仙花数')
#else:
#    print('不是水仙花数')
'''

'''
#作业：从控制台输入一个五位数，如果是回文数就打印'是回文数'，否则打印'不是回文数'
#“回文”是指正读反读都能读通的句子，它是古今中外都有的一种修辞方式和文字游戏，如“我为人人，人人为我”等。在数学中也有这样一类数字有这样的特征，成为回文数
#12321,12221,32123，即万位与个位相等，千位与十位相等
num = int(input('请输入一个五位数：'))
num_one = num % 10
num_ten = num // 10 % 10
num_thousand = num // 1000
num_ten_thousand = num // 10000 % 10
if num_one == num_ten_thousand and num_ten == num_thousand:
    print('是回文数')
else:
    print('不是回文数')

#字符串的索引
#num = input('请输入一个五位数：')
#if int(num[0]) == int(num[4]) and int(num[1]) == int(num[3]):
#    print('是回文数')
#else:
#    print('不是回文数')
'''
'''
#作业：从控制台输入两个数，输出较大的值，不准使用max()和min()
num1 = input('输入一个数：')
num2 = input('输入另一个数：')
if num1 > num2:
    print(num1)
else:
    print(num2)
'''

'''
#作业：从控制台输入三个数，输出较大的值
num1 = input('输入一个数：')
num2 = input('输入另一个数：')
num3 = input('再输入另一个数：')
if num1 > num2 > num3:
    print(num1)
if num1 > num3 > num2:
    print(num1)
if num2 > num1 > num3:
    print(num2)
if num2 > num3 > num1:
    print(num3)
if num3 > num2 > num1:
    print(num3)
if num3 > num1 > num2:
    print(num3)
num_max = max(num1,num2,num3)
print(num_max)
'''

#【day27】逻辑运算符
#逻辑与 and
#逻辑与运算表达式：表达式1 and 表达式2
#值：如果表达式1的值为True，表达式2的值为True，整个逻辑与运算表达式的值为True
#   如果表达式1的值为True，表达式2的值为False，整个逻辑与运算表达式的值为False
#   如果表达式1的值为False，表达式2的值为True，整个逻辑与运算表达式的值为False
#   如果表达式1的值为False，表达式2的值为False，整个逻辑与运算表达式的值为False
#总结:有False，则False
'''
num1 = 10
num2 = 20
if num1 and num1:
    print('*****')
'''
'''
num1 = 0
num2 = 20
if num1 and num2:
    print('*****')
'''
'''
num1 = 10
num2 = 20
if num1-10 and num2:
    print('*****')
'''
'''
#短路原则：
#表达式1 and 表达式2 and 表达式3 and ...... and 表达式n
#有一个为False，则整体都为False
#如果前面有一个表达式为False，则后面的表达式都不会运算了，因为整体为False了
'''

#逻辑或 or
#逻辑或运算表达式：表达式1 or 表达式2
#值：表达式1的值为True，表达式2的值为True，则True。
#   表达式1的值为False，表达式2的值为True，则True。
#   表达式1的值为True，表达式2的值为False，则True。
#   表达式1的值为False，表达式2的值为False，则False。
#总结：有一个True，则True。
'''
num1 = 0
num2 = 1
if num1 or num2:
    print('****')
'''
'''
#短路原则：
#表达式1 and 表达式2 and 表达式3 and ...... and 表达式n
#有一个为True，则整体都为True
#如果前面有一个表达式为True，则后面的表达式都不会运算了，因为整体为True了
'''

#逻辑非 not
#逻辑非运算表达式：not 表达式
#值：如果表达式为True，则not 表达式为Fasle
#   如果表达式为False，则not 表达式为True
'''
if not 1:
    print('******')
'''
'''
if not 0:
    print('******')
'''

#【day28】成员运算符
#in：如果指定序列中找到值，则为True，否则为Fasle
#not in：如果指定序列中没有找到值，则为True，否则为False
'''
num = 10
list1 = [1,2,3,4,5,6,7,8,9,10]
if num in list1:
    print('****')
'''
'''
num = 11
list1 =[1,2,3,4,5,6,7,8,9,10]
if num not in list1:
    print('****')
'''


#【day29】身份运算符
#is：判断两个标识符是不是引用同一个对象
#is not：判断两个标识符是不是引用不同的对象
'''
'''

#【day30】运算符优先级
'''
**
~ + - #+ -为正负号
* / % //
+ -
>> <<
&
^ |
<= < > >=
== !=
= += -= *= /= %= **= //=
is is not
in not in
or and not 
'''

'''
#作业：
#从而控制输入一个年份，判断是否是闰年(被4整出但是不被100整除，或被400整除)
yeas = int(input('请输入一个年份：'))
if (yeas % 4 == 0 and yeas % 100 != 0) or yeas % 400 == 0:
    print('此年份为闰年')
else:
    print('次年份为平年')
'''

#【day31】字符串
#什么是字符串？
#   字符串是以单引号或双引号括起来的任意文本
#'abc' "abc" 注意引号本身不是数据,字符串内容不可变
#创建字符串
'''
str1 = 'justin is good man！'
str2 = 'justin is handsome man!'
str3 = 'justin is nice man!'
'''
#字符串运算:
#字符串连接:加法
'''
str1 = 'justin is a'
str2 = ' good man'
str3 = str1 + str2
print(str3)
print(str1 , str2)
'''
#输出重复字符串：乘法
'''
str1 = 'cyl'
str2 = str1 * 3
print(str2)
'''
#访问字符串中的某一个字符:通过索引下标查找字符串
#索引：字符串的每一个元素都有一个数字下标，索引下标能立刻访问该元素
#   格式：变量名[下标]
'''
str1 = 'cyl'
str2 = str1[1]
print(str2)
'''
'''
#错误示范：以下代码会报错，因为字符串不可变
str1 = 'cyl'
str1[2] = 'c'
print(str1)
'''
#截取字符串中的一部分
'''
#截取范围:给定下标开始~给定下标的前一位
str1 = 'cyl is a good man!'
str2 = str1[4:13]
print(str2)
'''
#省略:右侧——截取范围：从给定下标开始，截取到最后
'''
str1 = 'cyl is a good man!'
str2 = str1[8:]
print(str2)
'''
#判断字符串中是否包含一些文本
'''
str1 = 'cyl is a good man!'
print('good' in str1)
print('bad' not in str1)
'''
'''
str1 = 'cyl is a good man!'
str2 = 'good'
if str1 in str2:
    print('在')
'''
#格式化输出：在字符串中用占位符站位，替换变量的值
#%d 整数占位符 %s 字符串占位符 %f 浮点数占位符
'''
name = 'cyl'
print('%s is good man')
'''
'''
age = 18
print('he is %d year')
'''
'''
num1 = int(input('请输入数字：'))
print('%.2d' % num1) #保留2位整数
'''
'''
money = 10.1
print('he has %.1f yuan' % money) #%.1f 精确到小数点后1位
'''
'''
name = 'cyl'
age = 18
money = 100.1
str1 = '%s is good man,he is %d year old, he has %.1f yuan'
print(str1 % (name, age, money))
'''
#转义字符 \
#将一些字符转换成有特殊含义的字符 \n占1个字符
#\n 换行符
'''
print('床前明月光，疑是地上霜。\n举头望明月，低头思故乡。')
'''
#\\取消转义
'''
print('\\n是换行字符')
'''
#\'转义引号
'''
print('i say :\'east or west, home is the best.\'')
'''
#\"转义双引号
'''
print('i say :\"east or west, home is the best.\"')
'''
#多行换行
#print('君不见黄河之水天上来，奔流到海不复回。\n君不见高堂明镜悲白发，朝如青丝暮成雪。\n人生得意须尽欢，莫使金樽空对月。\n天生我材必有用，千金散尽还复来。')
'''
print(#三引号
君不见黄河之水天上来，奔流到海不复回。
君不见高堂明镜悲白发，朝如青丝暮成雪。
人生得意须尽欢，莫使金樽空对月。
天生我材必有用，千金散尽还复来。#三引号)
'''
#\t 制表符 = Tab
'''
print('此处：\t（空格四下）')
'''
#r 字符串内部的转义符不转义
'''
print(r'\\\t\\')
'''
'''
print(r'D:\Youku Files\download\temp')
'''

#【day32】字符串的方法与函数
#eval(str)
#功能：将最外层的字符串去掉
#     如果字符串中只包含数字和运算，去掉字符串后，会进行表达式的计算，并将结果返回
#     如果是一个列表字符串，去掉字符串后就返回一个列表，即列表字符串转列表
'''
num1 = eval('123+1')
print(num1)
'''
'''
strList = '[{'a':'b'}]'
lis = eval(strList)
print(lis)
'''
#len(str)
#功能：返回字符串的长度(字符的个数)
'''
print(len('cyl'))
'''
#isdigit()
#功能：字符串中只包含数字，则返回True

#isalnum()
#功能：字符串中只包含数字和字母，则返回True

#isalpha()
#功能：字符串中只包含字母或汉字，则返回True

#.lower()方法
#功能：将str中的大写字母转换成小写字母
'''
str1 = 'JUSTIN IS GOOD MAN! '
print(str1.lower())
'''
#.upper()方法
#功能：将str中的小写字母转换成大写字母
'''
str1 = 'justin is good man!'
print(str1.upper())
'''
#.swapcase()方法
#功能：切换str中的字母大小写
'''
str1 = 'justin IS gOOD MAN!'
print(str1.swapcase())
'''
#capitalize()
#功能：首单词首字母大写，其他小写
'''
str1 = 'justin IS gOOD MAN!'
print(str1.capitalize())
'''
#.title()
#功能：每个单词首字母大写，其他小写
'''
str1 = 'justin IS gOOD MAN!'
print(str1.title())
'''
#.center(width[,fillchar]) char:character 字符
#功能：返回一个指定长度的居中字符串，fillchar为填充的字符,默认为空格
'''
print('justin'.center(40))
print('justin'.center(40,'*'))
'''
#.ljust(width[,fillchar])
#功能：返回一个指定长度的左对齐字符串，fillchar为填充的字符,默认为空格
'''
print('justin'.ljust(40))
print('justin'.ljust(40,'*'))
'''
#.rjust(width[,fillchar])
#功能：返回一个指定长度的右对齐字符串，fillchar为填充的字符,默认为空格
'''
print('justin'.rjust(40))
print('justin'.rjust(40,'*'))
'''
#.zfill(width)
#功能：返回一个指定长度的右对齐字符串，左端用0补齐
'''
print('justin'.zfill(40))
'''
#.count(str[,start][,end])
#功能：统计start到end范围内出现str的个数,默认从头到尾
'''
print('justin is a very very nice man!'.count('very',12,24))
'''
#.find(str[,start][,end])
#功能：从左至右检测str字符串是否包含在字符串的指定范围内，默认从头到尾检测，如果在则返回第一次str出现的开始下标，没有包含则返回-1
'''
print('justin is a very nice man!'.find('very'))
print('justin is a very nice man!'.find('very',15,24))
'''
#.rfind(str[,start][,end])
#功能：从右至左检测str字符串是否包含在字符串的指定范围内，默认从头到尾检测，如果在则返回第一次str出现的开始下标，没有包含则返回-1
'''
print('justin is a very very nice man!'.rfind('very'))
'''
#.index(str, start=0, end=len(str))
#功能:跟find()一样，区别在于字符串中不包含str时，会报异常
'''
print('justin is a very nice man!'.index('very'))
print('justin is a very nice man!'.index('good'))
'''
#.rindex(str, start=0, end=len(str))
#功能:跟rfind()一样，区别在于字符串中不包含str时，会报异常
'''
print('justin is a very very nice man!'.rindex('very'))
print('justin is a very very nice man!'.rindex('good'))
'''
#.lstrip(char= )
#功能：从左到右截掉相连的char
'''
print('      justin is a good man!'.lstrip())
print('******justin is a good man!'.lstrip('*'))
'''
#.rstrip(char= )
#功能：从右到左截掉相连的char
'''
print('justin is a good man!         '.rstrip())
print('justin is a good man!*********'.rstrip('*'))
'''
#.strip(char)
#功能：截掉两端相连的char,如果断点不是需要截取的字符串，则无法截取指定字符串
#参数char：一个字符串，字符串可以是多个字符，截取方式为每个字符独立截取，也就是说Str.strip('()')，Str的两端窃取'('和')'
'''
print('   justin   '.strip())
print('***justin***'.strip('*'))
print('([{},{}])'.strip('()'))
print('   \n([{},{}])   \n  \t '.strip(' \n()\t\r'))
'''
#.split()
#原型：.split(str=' ', num=-1)
#功能：用指定符号切割字符串并返回一个列表，其元素为字符串切割后的元素
#參數：str指定用切割，num切割次數
#     str默认用空格切割,num默认为-1,表示全部切割
'''
str1 ='justin is good man!'
list1 = str1.split(' ', 1)
print(list1)
'''
#.splitlines(keepends = False)
#功能：按照行分隔,同.split('\n')，keppends为True时，分割后的列表中保留\n
'''
str1 = '山穷水复疑无路\n柳暗花明又一村'
list1 = str1.splitlines
print(list1)
'''

'''
#控制台输入一个时间，打印该时间的下一秒
time1 = input('请输入时间（hh:mm:ss）：')
timeList = time1.split(':')
second = int(timeList[2])
minute = int(timeList[1])
hour = int(timeList[0])
second += 1
if second == 60:
    second = 0
    minute += 1
    if minute == 60:
        minute = 0
        hour += 1
        if hour == 24:
            hour = 0
print('你所输入时间的下一秒为：%.2d:%.2d:%.2d' % (hour, minute, second))
'''
#.replace('被替换','替换')
#功能：替换字符串中的元素

#.join()
#功能：用str1连接str2形成新的字符串
'''
str1 = ''
str2 = ['a', 'b', 'c']
str3 = str1.join(str2)
print(str3)
'''

#【day33】while语句
#while 表达式：
#   语句
#逻辑：当程序执行到while语句时，首先计算‘表达式’的值，
#   如果‘表达式’的值为假，则结束整个while语句，向下执行。
#   如果‘表达式’的值为真，则执行‘语句’，执行完‘语句’，再计算‘表达式’的值，
#       如果表达式的值为假，结束整个while语句，
#       如果表达式的值还为真，则执行‘语句’，
#           如此循环反复，知道表达式的值为假，结束整个循环
'''
num = 0
while num < 2:
    print('num的值为%d' % num)
    num += 1 #num = num + 1
'''
'''
#求和1+2+3+......+100
num = 1
sum1 = 0
while num <= 100:
    sum1 += num
    num += 1
print(sum1)
'''
'''
#将字符串中的每个字符一个一个打印出来
str1 = 'Justin is good man！'
num = 0
while num < len(str1):
    print(str1[num])
    num += 1
'''

#考核
'''
#打印出所有三位数的水仙花数
num = 100
while num < 1000:
    if num == (num // 100 ) ** 3 + (num // 10 % 10) ** 3 + (num % 10) ** 3:
        print(num)
    num += 1
'''
'''
#告诉我五位数中有多少个回文数
count = 0
num = 10000
while num < 100000:
    if (num // 10000 == num % 10) and (num // 1000 % 10 == num // 100 % 10):
        count += 1
    num += 1
print(count)
'''
'''
#从控制台输入一个数，判断是否是质数
num1 = int(input('请输入一个数：'))
for i in range(2,num1):
    if num1 == 2:
        print('%d是质数' % num1)
    if num1 % i == 0:
        print('%d不是质数' % num1)
        break
    if i == num1 - 1:
        print('%d是质数' % num1)
'''
'''
#从控制台输入一个数，分解质因数
#方法一：
num1 = int(input('请输入一个数：'))
renum1 = num1
num2 = [2, 3, 5, 7]
for i in num2:
    num1 = renum1
    while num1 % i == 0:
        print(i)
        num1 /= i
        
#方法二：
num = int(input('请输入数字：'))
i = 2
while num != 1:
    if num % i == 0:
        print(i)
        num //= i
    else:
        i += 1
'''


'''
#从控制台中输入一个字符串，返回这个字符串中有多少个单词
#方法一：
str1 = input('请输入英文句子：')
list1 = str1.strip().split(' ')
for i in list1:
    if len(i) == 0:
        list1.remove(i)
print(len(list1))
#方法二：
str1 = input('请输入英语句子，我能查看你的句子有几个单词：')
str2 = str1.strip()
index1 = 0
count1 = 0
while index1 < len(str1):
    while str2[index1] != ' ':
        index1 += 1
        if index1 == len(str2):
            break
    count1 += 1
    if index1 == len(str1):
        break
    while str2[index1] == ' ':
        index1 += 1
print(count1)

'''


'''
#从控制台中输入一个字符串，打印出这个字符串中的所有数字字符的和
str1 = input('请输入字符串：')
sum = 0
for i in range(len(str1)):
    if str1[i] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
        sum += int(str1[i])
print('字符串中数字的和为%d' % sum)
'''

#【day34】ASKII码值
#print(ord('a')) #ord()求ASCII码值
#print(chr(65)) #ASCII码值求对应符号

#作业：输入字符串，字母大小写切换
'''
str = input('请输入字符串：')
res = ''
for i in str:
    if i >= 'a' and i <= 'z':
        ch = chr(ord(i) - 32)
        res += ch
    elif i >= 'A' and i <= 'Z':
        ch = chr(ord(i) + 32)
        res += ch
    else:
        res += i
print(res)
'''
'''
#作业:验证码——随机生成六为数的验证码，包含数字，大写字母，小写字母
import random
str_ = ''
for i in range(6):
    ty = random.randrange(3)
    if ty == 0:
        #随机生成一个大写字母
        ch = chr(random.randrange(ord('A'), ord('Z') + 1))
        str_ += ch
    elif ty == 1:
        #随机生成一个小写字母
        ch = chr(random.randrange(ord('a'), ord('z') + 1))
        str_ += ch
    elif ty == 2:
        #随机生成一个数字
        ch = chr(random.randrange(ord('0'), ord('9') + 1))
        str_ += ch
print(str_)
'''


#【day35】字符串的比较
#从第一个字符开始比较，谁的ASKII码大谁就大，如果相等则比较下一个字符的ASKII码值的大小，谁码大谁大
'''
print(ord('a'), ord('b'), ord('c'), ord('d'))
print('ab' < 'd')
print('ab'<'abc')#长度不够用\0比较
'''



#【day36】布尔值与空值
'''
boolean1 = True #真
boolean2 = False #假
print(boolean1, boolean2)
'''
#空值：是Python里一个特殊的值，用None表示，None不能理解为0，因为0是有意义的，而None是一个特殊值。
'''
n = None
print(n)
'''

#【day37】列表list
#列表是有序的集合
#创建列表，格式：列表名 = [列表项1， 列表项2， 列表项3，......, 列表项n]
#作用：存储数据
'''
#创建一个空列表
list1 = []
print(list1)
'''
'''
#创建一个带元素的列表,元素可以是数字类型、字符串类型、布尔值、空值
list1 = [1, 2, 3]
list2 = [1, 2, '3', 'abc', True, None]
print(list1, list2)
'''

#【day38】访问列表数据
'''
# 取值，格式：列表名[下标]
list1 = [1, 2, 3, 4, 5]
num = list1[2]
print(num)
'''
'''
# 替换
list1[2] = 30
print(list1)
#【day】装饰器
'''
'''
#注意列表不能越界(下标超出了可表示的范围)
list1 = [1, 2, 3, 4, 5]
list1[5] = 100
'''
'''
#存储5个人的年龄，求他们的平均年龄
list1 = [45,23,56,8,5]
index = 0
sum = 0
while index < len(list1):
    sum += list1[index]
    index += 1
mean = sum / index
print(mean, sum, index)
'''

#【day39】列表的操作
#列表合并
'''
list1 = [1, 2, 3]
list2 = [2, 3, 4]
list3 = list1 + list2
print(list3)
'''
#列表重复
'''
list1 = [1, 2, 3]
list2 = list1 * 3
print(list2)
'''
#判断元素是否在列表中
'''
list1 = [1, 2, 3]
print(3 in list1)
print(4 in list1)
'''
#列表截取
'''
list1 = [1, 2, 3, 4, 5, 6, 7, 8, 9]
list2 = list1[0]
list3 = list1[3:6]
list4 = list1[4:]
list5 = list1[:5]
print(list4)
'''
#二维列表:列表元素为列表（列表嵌套）
'''
list1 = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
print(list1[2][2]) #第一次索引访问第一层列表，第二次索引访问嵌套列表
'''

#【day40】列表的方法
#末尾追加新元素
'''
list1 = [1, 2, 3]
list1.append(6)
list1.append([1, 2, 3])
print(list1)
'''
#末尾追加列表中的多个值
'''
list1 = [1, 2, 3]
list1.extend([4, 5, 6])
print(list1)
'''
#在指定下标处插入指定元素，原数据向后顺移
'''
list1 = [1, 2, 3]
list1.insert(1, [100, 100])
print(list1)
'''
#移除指定下标处的元素，默认最后一个元素,并返回删除的数据
#.pop(x=list[-1])
'''
list1 = [1, 2, 3]
list1.pop()
print(list1)
'''
'''
list1 = [1, 2, 3]
list1.pop(1)
print(list1)
'''
'''
list1 = [1, 2, 3]
num = list1.pop(1)
print(list1, num)
'''
#移除列表中的首次出现的指定元素
'''
list1 = [1, 2, 3, 2]
list1.remove(2)
print(list1)
'''
#清除列表中所有元素
'''
list1 = [1, 2, 3]
list1.clear()
print(list1)
'''
#从指定范围的列表中找出首个指定值的下标并返回，默认从头到尾
#.index(value [,start][,end])
'''
list1 = [1, 2, 3, 2, 4]
num1 = list1.index(2)
num2 = list1.index(2, 2, 4)
print(num1, num2)
'''
#返回列表中元素的个数
'''
list1 = [1, 2, 3]
print(len(list1))
'''
#返回列表中最大值
'''
list1 = [1, 2, 3]
print(max(list1))
'''
#返回列表中最小值
'''
list1 = [1, 2, 3]
print(min(list1))
'''
#返回统计指定元素的出现次数
'''
list1 = [1, 2, 3, 4, 5, 1, 2, 3, 4, 5]
count_num = list1.count(5)
print(count_num)
'''
'''
#练习
#从以下列表中，删除所有的5

list1 = [5, 6, 7, 4, 5, 6, 7, 5, 4, 5, 6, 5, 6]
count_num = list1.count(5)
num = 0
while num < count_num:
    list1.remove(5)
    num += 1
print(list1)

list1 = [5, 6, 7, 4, 5, 6, 7, 5, 4, 5, 6, 5, 6]
count_num = list1.count(5)
for i in range(count_num):
    list1.remove(5)
print(list1)
'''
'''
#练习倒序
list1 = [1, 2, 3, 4, 5]
num = 0
while num < len(list1):
    num1 = list1.pop()
    list1.insert(num, num1)
    num += 1
print(list1)
'''
#颠倒列表元素
'''
list1 = [1, 2, 3, 4, 5]
list1.reverse()
print(list1)
'''
#排序,默认升序
'''
list1 = [2, 4, 1, 3, 5]
list1.sort()
print(list1)
'''
#拷贝
#浅拷贝，引用拷贝
'''
list1 = [1, 2, 3, 4, 5] #栈区list1保存id（2861510451720），通过id指向堆区保存的12345
list2 = list1 #list2栈区保存的是list1的id，通过此id，list2也指向堆区的12345
list1[0] = 2 #修改堆区的12345为22345
print(list1)
print(list2)
print(id(list1))
print(id(list2))
'''
#深拷贝，内存拷贝
'''
list1 = [1, 2, 3, 4, 5] #栈区保存id（2625976099336），堆区保存12345
list2 = list1.copy() #栈区新开辟空间id（2625976099400），堆区保存12345
list1[0] = 2 #修改list1堆区的内容，不影响list2堆区的内容
print(list1)
print(list2)
print(id(list1))
print(id(list2))
'''
#列表的转换
#元组转列表
'''
tuple1 = (1,2,3)
print(type(tuple1))
list1 = list(tuple1)
print(list1)
'''
'''
#练习，在控制台输入5个数并找出这5个数中的第二大的值
list1 = []
num = 0
while num < 5:
    value1 = int(input('请输入数字:'))
    list1.append(value1)
    num += 1
list1.sort()
num_max = max(list1)
count_max = list1.count(num_max)
num1 = 0
while num1 < count_max:
    list1.pop()
    num1 += 1
num_second = list1[-1]
print(num_second)
'''

#条件控制语句
'''
#if 语句
#if_else 语句
'''
#if_elif_else 语句
#逻辑：当程序执行到if_elif_else语句时,首先计算‘表达式1’的值，
#       如果‘表达式1’的值为真，则执行‘语句1’。执行完‘语句1’，则跳过整个if_elif_else语句
#       如果‘表达式1’的值为假，则计算‘表达式2’的值
#           如果‘表达式2’的值为真，则执行‘语句2’。执行完‘语句2’，则跳过整个if_elif_else语句
#           如果‘表达式2’的值为假，则计算‘表达式3’的值
#               如果‘表达式3’的值为真，则执行‘语句3’。执行完‘语句3’，则跳过整个if_elif_else语句
#               如果‘表达式3’的值为假，则计算‘表达式4’的值。。。。。
#                   如此下去，知道某个表达式的值为真，则停止
#                       如果elif的表达式没有一个真，且有else，则执行语句e
#格式：
'''
if 表达式1:
    语句1
elif 表达式2:
    语句2
elif 表达式3:
    语句3
......
elif 表达式n:
    语句n
else:      #可有可无
    语句e
'''
'''
#控制台输入一个孩子的年龄，系统判断该孩子处于什么阶段
#方法一：能实现需求，但是程序执行效率低，因为每个if程序都会去判断一次
age = int(input('请输入年龄：'))
if age < 3:
    print('婴儿')
if (age >= 3) and (age < 7):
    print('儿童')
if (age >= 7) and (age < 18):
    print('青年')

#方法二:
age = int(input('请输入年龄：'))
if age < 3:
    print('婴儿')
elif (age >= 3) and (age < 7):
    print('儿童')
elif (age >= 7) and (age < 18):
    print('青年')
else:
    print('成年')
    
#方法三:
#elif还有一个隐藏功能，就是能执行到下一个表达式，则上一个表达式一定不满足，因此可以简化表达式的判断条件，从而提高程序效率
#每个el都是对它上面所有表达式的否定
age = int(input('请输入年龄：'))
if age < 3:
    print('婴儿')
elif age < 7:
    print('儿童')
elif age < 18:
    print('青年')
else:
    print('成年')
'''

#【day41】死循环
#表达式永远为真的循环
'''
while 1:
    print('justin is good man!')
'''

#【day42】while_else语句
#逻辑：在条件语句（表达式1）为False时执行else中的‘语句2’
#格式：
'''
#while 表达式1：
#   语句1
#else：
#   语句2
'''
'''
num1 = 0
while num1 <= 3:
    print('justin is good man!')
    num1 += 1
else:
    print('justin is nice man!')
print('justin is very very good man!')
'''

#【day42】for语句
#逻辑：按顺序取‘集合’中的每个元素赋值给‘变量’，在去执行语句，如此循环往复，直到取完‘集合’中的元素截止
#格式：
'''
for 变量名 in 集合：
    语句
'''
'''
for i in [1, 2, 3, 4, 5]:
    print(i)
'''

#【day43】列表生成器
#range([start=0,] end[, step=1])函数
#功能：生成数列
'''
list1 = range(10)
print(list1)
print(type(list1))
'''
'''
#range的参数可以控制循环的次数
for i in range(10):
    print(i)
'''
'''
#range(2, 10, 2) == [2, 4, 6, 8]
for i in range(2, 10, 2):
    print(i)
'''
'''
#求1+2+3+......+100的和
sum1 = 0
for i in range(1, 101):
    sum1 += i
print(sum1)
'''

#【day44】枚举器
#enumerate(list)
#作用：拿出一个列表的下标以及元素——[0, 1, 2],[1, 2, 3]
'''
print(type(enumerate([1, 2, 3])))
for index, i in enumerate([1, 2, 3]): #index, i = 下标, 元素
    print(index, i)
'''

#【day45】break语句
#作用：为了跳出while 和 for 循环
'''
#在for循环中，不修改range(10),当i==5时，跳出循环
for i in range(10):
    print(i)
    if i == 5:
        break

num1 = 0
while num1 < 10:
    print(num1)
    if num1 == 5:
        break
    num1 += 1
'''
'''
#注意：只能跳出距离它最近的那一层循环
for i in range(2):
    print(i)
    for k in range(3):
        if k == 1:
            break
        print(k)
'''
'''
#注意：循环语句可以有else语句，break导致循环截止，不会执行else
num1 = 0
while num1 < 5:
    print(num1)
    num1 += 1
    if num1 == 3:
        break
else:
    print('over')
'''

#【day46】continue语句
#作用：跳过当前循环中的剩余语句，然后继续下一次循环
#注意：只跳过距离最近的循环
'''
for i in range(10):
    print(i)
    if i == 3: #当i==3时，不打印*和&
        continue
    print('*')
    print('&')
'''
'''
num = 0
while num < 10:
    print(num)
    if num == 3:
        num += 1
        continue
    print('*')
    print('&')
    num += 1
'''

'''
作业：
#打印99乘法表
num1 = 1
str1 = ''
list1 = [1, 2, 3, 4, 5, 6, 7, 8, 9]
for i in list1:
    for k in range(1, i+1):
        product1 = k * i
        str_1 = '%d * %d = %d,' % (k, i, product1)
        str1 += str_1
        print(str1)
'''
'''
#输入两个数，求这两个数的最大公约数
num1 = int(input('输入数字1：'))
num2 = int(input('输入数字2：'))
list1 = []
list2 = []
if num1 > 1 and num2 > 1:
    for i in range(1, num1):
        if num1 % i == 0:
            list1.append(i)
    for i in range(1, num2):
        if num2 % i == 0:
            list2.append(i)
    relist = list1 + list2
    set_list = list(set(relist))
    max_num = max(set_list)
    print(max_num)
else:
    print('请输入大于1的数')
'''
'''
#输入一个字符串，将字符串中的大写字母转小写，小写字母转大写
str1 = input('请输入字符串：')
str2 = str1.swapcase()
print(str2)
'''
'''
#随机生成一个6位数的验证码（大写、小写、数字）
import random
list1 = []
str1 = ''
str2 = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
for i in range(6):
    str3 = random.choice(str2)
    str1 += str3
print(str1)
'''

#插入知识点
#   强制终止程序，强制停止程序，强制退出
#       exit()
#       当遍历多个列表元素时，只想打印第一个观察值，可以在pring()下面加上exit(),遍历打印了第一条数据后，就会被强制停止
#   列表推导式
#       lis = ['&x2343', '&x2343只保留文字']
#       lis = [re.sub('&x2343', '', li) for li in lis]
#       a = [li for li in lis if len(li)>0]

#【day47】turtle模块绘图
#简单的绘图工具，提供一个小海龟，可以把它理解成一个机器人，只能听得懂有限的命令
#导入 turtle库
#   运动命令
#       .forward(d) 向前移动d长度
#       .backward(d) 向后移动d长度
#       .right(d) 向右转动d度
#       .left(d) 向左转动d度
#       .goto(x,y) 移动到坐标为(x,y)的位置
#       .speed(speed) 笔画绘制的速度[0,10]
#   笔画控制命令
#       .up() 笔画抬起，在移动的时候不会绘图
#       .down() 笔画落下，在移动的时候回绘图
#       .setheading(d) 改变海龟的朝向
#       .pensize(d) 改变画笔的大小
#       .pencolor(colotstr) 改变画笔的颜色
#       .reset() 恢复所有设置，清空窗口，重置turtle状态
#       .clear() 清空窗口，不会重置turtle状态
#       .circle(r[,steps = 180]) 绘制圆形，r为半径，steps为边数
#       .begin_fill() 开始填充
#       .fillcolor(colorstr)
#       .end_fill()
#   其他命令
#       .done() 持续显示画布
#       .undo() 撤销上一次动作
#       .hideturtle() 隐藏海龟
#       .showturtle() 显示海龟
#       .screensize(x, y)
#绘图窗口的原点(0,0)在屏幕的正中间，默认海龟的方向为向右
'''
import turtle
turtle.speed(10)
turtle.forward(100)
turtle.right(90)
turtle.forward(100)
turtle.left(90)
turtle.forward(100)
turtle.left(90)
turtle.forward(100)
turtle.right(90)
turtle.forward(100)
turtle.goto(150, 50)
turtle.goto(0, 0)
turtle.up()
turtle.goto(150, 0)
turtle.right(90)
turtle.down()
turtle.pensize(5)
turtle.pencolor('red')
turtle.forward(100)
turtle.reset()
#turtle.clear()
turtle.circle(50)
turtle.begin_fill()
turtle.fillcolor('blue')
turtle.circle(50, steps=5)
turtle.end_fill()
turtle.done()
'''
'''
作业：正方形、矩形、正方体、五角星、奥运五环、围棋棋盘、国际象棋棋盘、
'''

#【day47】元组
#本质：一种有序集合
#特点：
#   与列表非常相似
#   且初始化就不能修改
#   使用小括号
#创建元组
'''
tuple1 = ()
print(type(tuple1))
'''
#创建带有元素的元组
'''
tuple1 = (1, 'abc', True)
print(tuple1)
'''
#定义一个元素的元组,必须加逗号
'''
tuple1 = (1,)
print(tuple1)
'''
#元祖元素的访问
#格式:元素名[下标]
#下标从0开始,同样不能越界（下标超出范围）
'''
tuple1 = (1, 2, 3, 4, 5)
print(tuple1[0])
'''
'''
#下标可以是负数，从-1开始，-1是最后一个元素，-2是倒数第二个元素，依次类推
tuple1 = (1, 2, 3, 4, 5)
print(tuple1[-1])
'''
'''
#元组元素不可变，是不可变其在堆区的元素值，但此堆区有保存指向其他列表的地址，列表中的值可被修改，但是不能修改这个列表地址
tuple1 = (1, 2, 3, 4, [5, 6, 7])
#tuple1[0] = 100 #会报错
#tuple1[4] = [6, 7, 8] #会报错
tuple1[4][0] = 6
print(tuple1)
'''
#删除元组
'''
tuple1 = (12, 13)
del tuple1
print(tuple1)
'''
#元组的操作
#元组的合并
'''
tuple1 = (1, 2, 3)
tuple2 = (4, 5, 6)
tuple3 = tuple1 + tuple2
print(tuple3)
'''
#元组的重复
'''
tuple1 = (1, 2, 3)
tuple2 = tuple1 * 3
print(tuple2)
'''
#判断元素是否在元组中
'''
tuple1 = (1, 2, 3)
print(1 in tuple1)
'''
#元组的截取
#格式：元组名[start：end]
#从开始下标开始截取，结束到结束下标之前
'''
tuple1 = (1, 2, 3, 4, 5)
tuple2 = tuple1[2:5]
tuple3 = tuple1[:5]
tuple4 = tuple1[2:]
print(tuple2)
'''
#二维元组:元素为一维元组的元组
'''
tuple1 = (1, 2, (1, 2, 3))
print(tuple1[2][1])
'''
#元组的函数与方法
#len() 返回元组中元素的个数
'''
tuple1 = (1, 2, 3)
print(len(tuple1))
'''
#max() 返回元组中元素的最大值
'''
tuple1 = (1, 2, 3)
print(max(tuple1))
'''
#min() 返回元组中元素的最小值
'''
tuple1 = (1, 2, 3)
print(min(tuple1))
'''
#tuple() 将列表转换成元组
'''
list1 = [1, 2, 3]
tuple1 = tuple(list1)
print(tuple1)
'''
#元组的遍历
'''
for i in (1,2,3,4,5):
    print(i)
'''

#【day48】字典
#本质：由键和值组成的无序的集合,可以通过键来访问元素
#创建字典
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
print(dict1)
print(type(dict1))
'''
#创建空字典，再赋值
'''
dict1 = {}
dict1['name'] = 'justin'
dict1['gender'] = 'male'
dict1['phone'] = '18086829907'
print(dict1)
'''
#用**kwargs可变参数传入关键字创建字典
'''
dict1 = dict(one=1, two=2, three=3)
print(dict1)
'''
#使用可迭代对象创建字典
'''
#列表压缩迭代
dict1 = dict(zip(['one', 'two', 'three', 'three'],[1, 2, 3, 4])) #如果键相同，取最后面的值
print(dict1)
list1 = list(dict1)
print(list1)

#元组对迭代
dict2 = dict([('one',1), ('two',2), ('three',3), ('three',4)]) #如果键相同，取最后面的值
print(dict2)
'''

#
#字典元素的访问
#格式：变量名[键]
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
name = dict1['name']
print(name)
'''
#注意：同一个字典中键不能重复且不可变，值不做任何限制
'''
dict1 = {'name':'justin', 'name':'cyl'} #会报错
dict2 = {['name']:'justin', 'name':'cyl'} #会报错
'''
#字典元素的删除
#格式:del 变量名[键]
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
del dict1['name']
print(dict1)
'''
#字典的遍历
#遍历得到的i是key
'''
dict1 = {'a':'1','b':'2','c':'3'}
for i in dict1:
    print(i)
'''
#枚举遍历（遍历字典）
#得到kv
'''
for k, v in dic.items():
    print(k + ':' + v)
说明：items是字典的一个方法，能将字典格式{'a':1,'b':2}，转变成dict_items([('a', 1), ('b', 2)])对象
    可以遍历这个对象取到响应的元组，k取元组的[0],v取元组的[1]
'''

'''
dict1 = {'a':'1','b':'2','c':'3'}
for v, k in enumerate(dict1):
    print(k, v)
'''
#字典的函数与方法
#str() 字典转换成字符串
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
print(str(dict1))
'''
#.clear() 字典元素的清除
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
dict1.clear()
print(dict1)
'''
#字典的浅拷贝
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
dict2 = dict1
dict1['name'] = 'cyl'
print(dict1)
print(dict2)
'''
#.copy() 字典的深拷贝
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
dict2 = dict1.copy()
dict1['name'] = 'cyl'
print(dict1)
print(dict2)
'''
#.get(k[, d]) 获得对应键的值，如果没有该键，则返回d
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
value1 = dict1.get('name')
value2 = dict1.get('height','未记录身高')
print(value1, value2)
'''
#.items()
#功能：使用字典的元素创建一个以(key,value)为一组的元组对象
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
for k,v in dict1.items():
    print(k, v)
'''

#.keys()
#功能：使用字典中的键值创建一个列表对象
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
key1 = dict1.keys()
print(key1, type(key1))
for i in key1:
    print(i)
'''
#.value()
#功能：返回字典中的值
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
value1 = dict1.values()
print(value1)
for i in value1:
    print(i)
'''
#.popitem()
#功能：删除字典中最后个元素，并返回该元素的元组对
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
dict2 = dict1.popitem()
print(dict2)
dict3 = dict1.popitem()
print(dict3)
'''
#.pop(键)
#功能：删除指定键的值并返回其值
'''
#间接修改字典的键
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
dict1['Name'] = dict1.pop('name')
print(dict1)
'''

#.setdefault(k[, d])
#功能:返回通过key查找字典中对应的value，如果字典中有key，则返回对应value。反之返回d，并且在原字典中添加该键值对
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
value1 = dict1.setdefault('name')
print(value1)
value2 = dict1.setdefault('email','417217170@qq.com')
print(value2)
print(dict1)
'''
#.update(dict)更新（插入）键值对
'''
dict1 = {'name':'justin', 'gender':'male', 'phone':'18086829907'}
dict1.update({'email':'417217170@qq.com'})
print(dict1)
'''
'''
#解析歌词：输入到时间(秒)，打印对应歌词
'''
musicLrc = '''
[01:01.80][01:01.80]有些事啊 本该笑的 而我却哭了
[01:08.70][01:01.80]有些人啊 以为长久 中途却散了
[01:16.20]摊开手掌 才发现 紧握的是 空的
[01:23.15]岁月除了 皱纹白发　又留下了什么
[01:29.20]
[01:30.69]有些爱呀 明知有毒 还是去喝了
[01:37.50]有些恨呀 瓜葛交错 不觉已忘了
[01:45.50][01:01.80][01:01.80]闭上眼睛 脑海里 平静到无 风波
[01:52.16]故事终究 开花结果 归于尘土了
[01:58.63]
[02:03.79]时光太瘦 指缝太宽 漏去的不可逆转
[02:11.04]伤口太深 疤痕太浅 过往被风轻云淡
[02:18.71][01:01.80]越长大 越明白 身后的 那几年
[02:25.11]原来是命运 设局的 一层一层铺垫
[02:32.32]
[02:32.78]人生很长 缘分很短 错过的不可逆转
[02:39.88][01:01.80][01:01.80]欲望很近 梦想很远 有取舍方能实现
[02:47.69]越长大 越怀念 回不去 的从前
'''
'''
import time
rlcDict = {}
musicLrcList = musicLrc.splitlines()
for lrcLine in musicLrcList:
    lrcLineList = lrcLine.split(']')
    for i in range(len(lrcLineList) - 1):
        timeStr = lrcLineList[i][1:]
        timeList = timeStr.split(':')
        time1 = float(timeList[0]) * 60 + float(timeList[1])
        rlcDict[time1] = lrcLineList[-1]
allTimeList = []
for i in rlcDict.keys():
    allTimeList.append(i)
allTimeList.sort()
second = float(input('输入秒数：'))
for i in allTimeList:
    if second < allTimeList[0]:
        print('时间太小，至少要输入大于%f' % allTimeList[0])
        break
    elif second < i:
        result_index = allTimeList.index(i) - 1
        result_time = allTimeList[result_index]
        print(rlcDict[result_time])
        break
for t in rlcDict:
    time.sleep(t)
    print(rlcDict[t])
'''

#【day49】set 集合
#类似dict，是一组key的集合，不存储value
#本质：无序、无重复元素的集合、无下标
#创建set需要一个liest或tuple或者dict作为输入集合
#重复元素在set中会自动过滤
'''
set1 = set([1, 2, 3, 4, 5, 5, 4, 3, 2, 1])
print(set1)
'''
'''
set1 = set((1, 2, 3, 4, 5, 5, 4, 3, 2, 1))
print(set1)
'''
'''
set1 = set({'a':'1', 'b':'2'})
print(set1)
'''
#添加
'''
set1 = set([1, 2, 3, 4, 5])
set1.add(6)
print(set1)
'''
'''
set1 = set([1, 2, 3, 4, 5])
set1.add([1, 2, 3]) #报错：set的元素不能是列表，因为列表是可变的
print(set1)
'''
'''
set1 = set([1, 2, 3, 4, 5])
set1.add((1, 2, 3))
print(set1)
'''
'''
set1 = set([1, 2, 3, 4, 5])
set1.add({'a':'1', 'b':'2'}) #报错：set元素不能是字典，因为存在value
print(set1)
'''
#.update() 打碎插入
#可插入列表、字典、元组、字符串
'''
#打碎插入字典
set1 = set([1, 2, 3, 4, 5])
set1.update({'a':'1', 'b':'2'}) #报错：set元素不能是字典，因为存在value
print(set1)
'''
'''
#打碎插入字符串
set1 = set([1, 2, 3, 4, 5])
set1.update('abc') #报错：set元素不能是字典，因为存在value
print(set1)
'''
#.remove()
#删除指定元素
'''
set1 = set([1, 2, 3, 4, 5])
set1.remove(3)
print(set1)
'''
#set的遍历
'''
set1 = set([1, 2, 3, 4, 5])
for i in set1:
    print(i)
'''
#set的枚举遍历
'''
set1 = set([1, 2, 3, 4, 5])
for index, data in enumerate(set1):
    print(index, data)
'''
#交集、并集
'''
set1 = set([1, 2, 3])
set2 = set([2, 3, 4])
#求交集
a1 = set1 & set2
print(a1, type(a1))
#求并集
a2 = set1 | set2
print(a2)
'''
#类型转换
#list1 -> set
'''
list1 = [1, 2, 3, 4, 5, 5, 6]
set1 = set(list1)
'''
#tuple -> set
'''
tuple1 = (1, 2, 3, 4, 5)
set1 = set(tuple1)
'''
#set -> list
'''
set1 = {1, 2, 3}
list1 = list(set1)
'''
#set -> tuple
'''
set1 = {1, 2, 3}
tuple1 = tuple(set1)
'''
#使用set去重list
'''
list1 = [1, 2, 3, 3, 2, 1]
list2 = list(set(list1))
print(list2)
'''

#【day50】迭代器
#可迭代对象：可以直接作用于for循环的对象，统称为可迭代对象(Iterable)。可以用
# .isinstance() 可判断一个对象是否是可迭代对象
#可以直接作用于for循环的数据类型一般分为两类：
#   1、集合数据类型，如list、dict、set、string。
#   2、是generator,包括生成器和带yield的generator。

#判断对象是否为Iterable对象
'''
from collections import Iterable
print(isinstance([], Iterable))
print(isinstance((), Iterable))
print(isinstance({}, Iterable))
print(isinstance('', Iterable))
print(isinstance((x for x in range(10)), Iterable))
'''

#迭代器：不但可以直接作用于for循环，还可以被next()函数不断的调用并返回下一个值，直到抛出一个stopIteration错误，表示无法继续返回下一个值
#可以被next()函数调用并不断返回下一个值的对象称为迭代器(Iterator对象)，可以使用isinstance()函数判断一个对象是否是iterator对象
'''
from collections import Iterator
print(isinstance([], Iterator))
print(isinstance((), Iterator))
print(isinstance({}, Iterator))
print(isinstance('', Iterator))
print(isinstance((x for x in range(10)), Iterator)) #只有[x for x in range(10)]是迭代器
'''
#迭代器的使用
'''
num = (x for x in range(5))
print(next(num))
print(next(num))
print(next(num))
print(next(num))
print(next(num))
'''
'''
num = (x for x in [212, 2131, 5, 15, 1231])
print(next(num))
print(next(num))
print(next(num))
print(next(num))
print(next(num))
'''
#iter()函数
#功能：list、tuple、dict、string转成Iterator对象
'''
from collections import Iterator
print(isinstance(iter([]), Iterator))
print(isinstance(iter(()), Iterator))
print(isinstance(iter({}), Iterator))
print(isinstance(iter(''), Iterator))
'''
'''
iter1 = iter([1, 2, 3, 4, 5])
print(next(iter1))
print(next(iter1))
'''
'''
#inter的运用：输入end才结束input程序
endstr = 'end'
str = ''
for line in iter(input, endstr):
    str += line + '\n'
print(str)
'''

#【day51】函数
#定义：在一个完整的项目中，某些功能在反复使用，那么我们将这个功能封装起来，当我们是用这个功能的时候，直接调用函数即可。
#本质：函数就是对功能的封装
#优点：简化代码结构，增加了代码的复用度；如果想修改某些功能或者调试某些BUG，只需要修改对应的函数即可

#定义函数，格式：
#def 函数名(参数1, 参数2, 参数3,......参数n):
#   语句
#   return 表达式

#函数的构成：
#def:函数代码块以def关键字开始
#函数名：遵循标识符规则
#参数列表(参数1, 参数2, 参数3,......参数n):任何传入函数的参数和变量必须放在圆括号之间，用逗号分隔。函数从函数的调用者那里获取信息
#()：是参数列表的开始和结束
#冒号：函数内容（封装的功能）以冒号开始，并且缩进
#语句：函数封装的功能
#return：结束函数并返回信息给函数的调用者
#表达式：即为要返回给函数调用者的信息
#例子：我给钱让小明去买水，买了水并给我，参数为钱，return的是水，封装的是去买水的动作
#注意：最后的return可以不写，相当于return None

#最简单的函数——无参数，无返回
'''
def myPrint():
    print('justin is good man!')
'''
#函数的调用
#格式：函数名(参数列表)
#函数名：是要使用的功能的函数函数名字
#参数列表：函数的调用者给函数传递的信息
#本质：实参给形参赋值的过程
#注意：没有实参，括号不能省略
'''
myPrint()
'''
'''
练习：编写一个函数，给函数一个字符串和一个年龄，在函数内部打印出来
def myPrint(name, age): #参数列表(形参)：形式参数,定义函数时小括号里面的变量，本质是变量
    print('我的名字是：%s，我的年龄是：%d' % (name, age))
myPrint('justin', 30) #参数列表(实参):实际参数,调用函数是给函数传递的数据，本质是值
#函数调用的本质是，实参给形参赋值的过程
'''
#参数必须按顺序传递，个数目前要对应
'''
def personalData(name, weight, height): 
    print('运动员：%s, 体重：%d, 身高：%d' % (name, weight, height))
personalData('justin', 70) 
'''
#编写函数，实现功能：给函数两个数，返回这两个数的和
'''
def mySum(num1, num2):
    sum1 = num1 + num2
    return sum1 #将结果返回给函数的调用者
    print('****') #return结束函数，return后面的代码不执行
num = mySum(5 , 6)
print(num)
'''

#【day52】传递参数
#值传递：传递的值是不可变类型——string、tuple、number
'''
#传递的是不可变的值，到函数中去赋值后，变量的值不变
def func1(num):
    print(num)
    print(id(num))
    num = 10
    print(id(num))
temp = 20
print(id(temp))
func1(temp)
print(temp)
'''
#引用传递：传递的值是可变类型——list、dict、set
'''
#传递的是可变的值，到函数中去赋值后，变量的值可变
def func1(list0):
    list0[0] = 100
list1 = [1, 2, 3, 4, 5]
func1(list1)
print(list1)
'''

#【day53】关键字参数
#允许函数调用时参数的顺序与定义时不一致
'''
def personalData(name, weight, height):
    print('运动员：%s, 体重：%d, 身高：%d' % (name, weight, height))
personalData(weight=70, height=170, name='justin')
'''

#【day54】默认参数
#概念：调用函数时，如果没有给函数传递参数，则使用默认参数
'''
def personalData(name='justin', weight=70, height=170):
    print('运动员：%s, 体重：%d, 身高：%d' % (name, weight, height))
personalData()
personalData('mia')
personalData(height=170, name='mia')
'''
'''
#如果要使用默认参数，最好将默认参数放到最后
#因为在传递参数时，依然遵循对齐原则
def personalData(name='justin', weight, height = 170):
    print('运动员：%s, 体重：%d, 身高：%d' % (name, weight, height))
personalData(70)
'''
#【day55】*args 不定长参数
#概念：能处理比定义时更多的参数
#格式：*变量名
#加了*号的变量存放所有未命名的变量参数，如果在函数调用时没有指定参数，它就是一个空元组，
'''
def invitation(name, *args):
    print('今天我邀请了：\n%s' % name)
    print(args, type(args))
    for i in args:
        print(i)
invitation('mia','dice','coco')
'''
'''
#使用不定长参数，求随意传入数字的和
def mySum(*args):
    sum1 = 0
    for i in args:
        sum1 += i
    return sum1
print(mySum(1,2,3,4,5))
'''
#【day56】键值对不定长参数字典
#格式：**变量名
#和*变量所代表的意义类似
#使用kwargs必须使用关键字参数，kwargs保存的是字典格式的数据
#理论上函数的参数时无限的，但实际上最好不要超过6、7个
'''
def func1(**kwargs):
    print(kwargs, type(kwargs))
func1(x=1, y=2, z=3)
'''
'''
#*和**连用，实参能接受任何形式
def func1(*args, **kwargs):
    print(args, kwargs)
func1(1, 2, 3, x=1, y=2)
'''

#【day57】匿名函数
#概念：不使用def这样的语句定义函数，使用lambda来创建匿名函数
#特点：lambda只是一个表达式，函数体比def简单
#     lambda的主题是一个表达式，而不是代码块，仅仅只能在lambda表达式中封装简单的逻辑
#     lambda函数有自己的命名空间，且不能访问自有参数列表之外的或全局命名空间的参数
#     虽然lambda是一个表达式且看起来只能写一行，与c和c++内联函数不同
#格式：lambda 参数1，参数2...参数n : 表达式
'''
sum1 = lambda num1, num2 : num1 + num2
print(sum1(1, 3))
'''
'''
#作业：将以前的题目用函数封装
'''

#【day58】装饰器
#是一个闭包，把一个函数当做参数返回一个替代版的函数，本质上就是一个返回函数的函数
#当发现别人的函数缺少一些功能而又不能修改原函数，则用装饰器加以修改
'''
#简单的装饰器：
def func1():
    print('justin is good man!')
def outer():
    print('*******')
    func1()
outer()
'''
'''
def func1():
    print('justin is good man!')
def outer(func):
    def inner():
        print('*******')
        func()
    return inner
f = outer(func1)
f() #函数是func1的加强版本
'''
'''
#魔法装饰器@
def outer(func1):
    def inner(age):
        if age < 0:
            age = 0
        func1(age)
    return inner
@outer #使用@符号讲装饰器运用到函数
def say(age):
    print('cyl is %d years old' % age)
say(-1)
'''
#通用装饰器
'''
def outer(func):
    def inner(*args, **kwargs):
        #添加修改的功能
        print('*******')
        func(*args, **kwargs)
    return inner
@outer
def say(name, age):
    print('my name is %s, i am %d years old!' % (name, age))
say('justin', 30)
'''

#【day59】偏函数
#修改原函数的默认值得到的函数为偏函数
'''
print(int('1010', base=2))
def int2(str1, base = 2):
    return int(str1, base)
print(int2('1011'))
'''
#生成偏函数
'''
import functools
int3 = functools.partial(int, base=2) #把指定函数的参数固定住，形成一个新的函数
print(int3('1011'))
'''

#【day60】变量的作用域
#变量可以使用的范围
#程序的变量并不是在所有位置都能使用的，访问的权限决定于变量的位置——变量在哪个位置定义的
#作用域：
#   块级作用域
#       缩进内定义的变量
#   局部作用域
#       函数内定义的变量，在函数的外部是不能访问到局部变量的
#   全局作用域
#       在脚本中，直接顶格定义的变量
#
'''
num = int(input('请输入一个整数：'))
if num % 2 == 0:
    info = '输出的数是偶数'
print(info)
'''
#局部变量：局部作用域不能被外部函数调用
'''
def func1():
    name = 'justin'
func1()
print(name)
'''
#全局变量：当前脚本中任意位置，都能访问到全局变量
'''
name = 'justin'
def func1():
    print(name)
func1()
'''
'''
def func1():
    print(name)
name = 'justin'
func1()
'''
'''
def func1():
    print(name)
func1()
name = 'justin'
'''
#调用变量遵循就近层原则，在外部调用则在外部找变量，在内部调用则在内部
'''
name = 'justin'
def func1():
    name = 'mia'
    #print(name)
func1()
print(name)
'''
'''
name = 'justin'
def func1():
    name = 'mia'
    def inner():
        name = 'dici'
        print(name)
    inner()
func1()
print(name)
'''
#作用域链：程序在运行的时候，去查找变量的规则，
#        先从内层查找，没有则往上一层查找
#        作用域链，在程序加载的时候，作用域就已经形成了，作用域链也生成了。
'''
name = 'justin'
def func1():
    print(name)
def func2():
    name = 'green'
    func1()
func2()
'''
#global
#功能：将局部变量转化成全局变量
#格式：global 变量名
'''
name = 'justin'
def func1():
    global name
    name = 'mia'
func1()
print(name)
'''
#nonlocal
#功能：将局部内部变量升上一层
#格式： nonlocal 变量名
'''
name = 'justin'
def func1():
    name = 'mia'
    def inner():
        nonlocal name
        name = '陈紫妍'
    inner()
    print(name)
func1()
print(name)
'''
#闭包
'''
#函数名即变量（函数也是一个数据）
def f():
    print('f()函数被调用了！')
a = f
b = a
b()
'''
'''
#输出什么？
def f():
    name = 'justin'
    def inner():
        print(name)
    return inner
a = f()
a()
'''
'''
#输出什么？
def f():
    temp = []
    for i in range(10):
        def inner():
            print(i)
        temp.append(inner) #[inner(i),inner(i)...inner(i)] ，for循环执行完后temp中保存了10个inner函数，参数为变量i而不是值
    return temp #此时i==9，当在调用10个inner(i)时，i的值都是9
funcs = f()
funcs[0]()
for f in funcs:
    f()
'''

#【day61】异常处理
#当程序遇到问题，不让程序结束，而越过错误向下执行，

#try-except-else
#格式：
'''
#try:
#   语句t
#except 错误码1 as e：
#   语句1
#except 错误码2 as e:
#   语句2
#......
#except 错误码n as e:
#   语句n
#else:
#   语句e
#注意：else可有可无
'''
#作用:用来检测try语句模块中的错误，从而让except语句捕获错误信息并处理，没有捕获到的错误则else
#逻辑：当程序执行到try-except-else语句时
#     1、如果当try'语句t'执行出现错误,会匹配第一个错误码，如果匹配上，就执行对应的语句
#     2、如果当try'语句t'执行出现错误，没有匹配的异常，错误将会被提交到上一层的try语句。或者到程序的最上层
#     3、如果当try'语句t'执行没有错误，执行else下的'语句e'(你得有)
'''
try:
    #print(num)
    #print(3/0)
    print(3/1)
except ZeroDivisionError as e:
    print('除数为零了')
except NameError as e:
    print('没有定义变量')
else:
    print('代码没有问题')
print('*************')
'''
#使用except而不使用任何的错误类型
'''
try:
    print(4/0)
except:
    print('程序出现了异常') #此时记录一个错误日志
'''
#使用except带有多种异常
'''
try:
    print(num)
except (ZeroDivisionError, NameError):
    print('出现了NameError或ZeroDivisionError')
'''
#特殊
'''
#1、异常本质是一个类，所有的异常类都继承自BaseException，它是异常基类，所以如果用BaseException捕获异常的时候，能捕获所有的异常错误
try:
    print(3/0)
except BaseException as e:
    print('异常1')
except ZeroDivisionError as e:
    print('异常2')
'''
'''
#2、跨越多层调用,main调用func2，func2调用func1，func1出现了错误，这是只要main捕获到了就能处理
def func1(num):
    print(1/num)
def func2(num):
    func1(num)
def main():
    func2(0)
try:
    main()
except ZeroDivisionError as e:
    print('ZeroDivisionError')
'''

#try-except-finally
#格式：
'''
#try:
#   语句t
#except 错误码1 as e：
#   语句1
#except 错误码2 as e:
#   语句2
#......
#except 错误码n as e:
#   语句n
#finally:
#   语句f
'''
#作用：语句t无论有无错误，无论是否捕获到错误，都将执行语句f
'''
try:
    print(1/0)
except ZeroDivisionError as e:
    print('ZDE')
finally:
    print('必须执行它')
'''

#【day62】断言
'''
#当断言正确时，打印后面的字符串
def func(num, div):
    assert (div != 0), 'div不能为0' 
    return num/div
print(func(10, 2))
print(func(10, 0))
'''

#【day63】文件读写
#读文件
#过程：
#   打开文件
#   读文件
#   关闭文件
'''
'''
#open(path, flag[, encoding][, errors])
#path:要打开文件的路径
#flag：打开的方式：只需记住r读w写，二进制加个b
#           r    以只读的方式，文件的描述符放在文件的开头
#           rb   以二进制的方式打开一个文件用于只读，文件的描述符放在文件的开头
#           r+   打开一个文件用于读写，文件的描述符放在文件的开头
#           w    打开一个文件只用于写入，如果该文件已经存在会覆盖，如果不存在则会创建新文件
#           wb   打开一个文件只用于写入二进制，如果该文件已经存在会覆盖，如果不存在则会创建新文件
#           w+   打开一个文件用于读写
#           a    打开一个文件用于追加，如果文件存在，文件描述符放到文件末尾
#           a+   打开一个文件用于追加
#encoding:编码方式（utf-8）
#errors:错误处理

#打开文件
'''
path = r'D:\qian_feng_education\first_project\test.txt'
#f = open(path, 'r', encoding='utf-8', errors='ignore')
f = open(path, 'r')
'''
#读文件内容
'''
#1 .read()读取文件全部内容
str1 = f.read()
print(str1)
'''
'''
#2 .read(10)读取指定文件字符数
str1 = f.read(10)
print(str1)
str2 = f.read(10)
print(str2)
'''
'''
#3 .readline()读取整行，包括'\n'字符
str1 = f.readline()
print(str1)
str2 = f.readline()
print(str2)
'''
'''
#4 .readline(10)读取指定文件字符数
str1 = f.readline(10)
str2 = f.readline(10)
print(str1)
print(str2)
'''
'''
#5 .readlines() 读取所有行并返回列表
list1 = f.readlines()
print(list1)
'''
'''
#6 .readlines(2) 若给定额数字大于零，返回实际size字节的行数
list1 = f.readlines(2)
print(list1)
list2 = f.readlines(22)
print(list2)
'''
#.seek()修改描述符的位置,参数数字代表字符数
'''
f.seek(0)
'''
#关闭文件
'''
f.close()
'''
#一个完整的读文件过程
'''
path = r'D:\qian_feng_education\first_project\test.txt'
try:
    f1 = open(path, 'r', encoding='utf-8', errors='ignore')
    print(f1.read())
except NameError as e:
    print('f1未打开')
finally:
    if f1:
        f1.close()
'''
#with 不管文件是否打开，程序都会关闭文件
'''
path = r'D:\qian_feng_education\first_project\test.txt'
with open(path, 'r') as f2:
    print(f2.read())
'''
#写文件
'''
path = r'D:\qian_feng_education\first_project\test.txt'
f = open(path, 'w')
#1、将信息写入缓冲区，
f.write('my daughter is beautiful girl!\n')
#2、刷新缓冲区,将信息立刻写入文件，而不是被动的等待文件关闭时刷新缓冲区写入信息
f.flush()
#没有.flush()的情况时，执行等待(不关闭文件),信息不能写入，当有.flush()时，信息立即写入
#while True:
    #pass
#3、当缓冲区满了，会自动将信息写入文件
f.close() #自动刷新缓冲区
'''
#追加写入
'''
path = r'D:\qian_feng_education\first_project\test.txt'
f = open(path, 'a', encoding='utf-8', errors='ignore')
f.write('my daughter name is mia')
f.close()
'''
'''
#简单版追加
path = r'D:\qian_feng_education\first_project\test.txt'
with open(path, 'a', encoding='utf-8', errors='ignore') as f:
    f.write(' and mia is very good!')
'''
#二进制文件读写
#在读写是一定要记得编码和解码
'''
path = r'D:\qian_feng_education\first_project\binary.txt'
with open(path, 'wb') as f:
    str1 = 'my daughter name is mia and she is very good!'
    f.write(str1.encode('utf-8')) #给字符串编码
with open(path, 'rb') as f:
    data = f.read()
    print(data,type(data))
    newData = data.decode('utf-8') #给字符串解码
'''
#【day64】list-tuple-dict-set的文件操作
#列表保存、元组保存、字典保存、集合保存
#数据持久模块——可以在多个程序间使用的数据，相当于把数据存到磁盘
'''
import pickle
myList = [1, 2, 3, 4, 5, 'abc']
#myTuple = (1, 2, 3, 4, 5, 'abc')
#myDict = {'a':'1','b':'2','c':'3'}
#mySet = {1, 2, 3, 4, 5, 'abc'}
path = r'D:\qian_feng_education\first_project\LTDS.txt'
with open(path, 'wb') as f:
    pickle.dump(myList, f) #dump装载
with open(path, 'rb') as f:
    tempList = pickle.load(f) #laod加载
    print(tempList)
'''

#【day65】os模块
#包含了普遍操作系统的功能
'''
import os
'''
#显示当前使用系统的类型 nt：windows posix：Linux、Unix Max OS X
'''
print(os.name)
'''
#打印操作系统详细的信息(windows不支持)
'''
#print(os.uname())
'''
#获取操作系统的所有的环境变量
'''
print(os.environ)
'''
#获取指定环境变量的地址
'''
print(os.environ.get('APPDATA'))
'''
#修改环境变量的地址（网上可查）
'''
'''
#获得当前目录
'''
print(os.curdir, type(os.curdir))
'''
#获取当前工作目录,即当前Python脚本所在的目录
'''
print(os.getcwd())
'''
#以列表的形式返回指定目录下的所有的文件(包括文件夹，文件夹也是文件的一种)
'''
print(os.listdir(r'D:\qian_feng_education\first_project'))
'''
#创建目录（创建文件夹）
#os.mkdir(path) .代表当前路径，.\代表当前路径下的路径，..代表上一层路径
'''
os.mkdir(r'.\justinTest')
'''
#os.makedirs(path) 多层级创建目录（当父级不存在时，自动创建父级）
'''
path = r'D:\qian_feng_education\creatTextDir\01'
os.makedirs(path)
'''
#os.redir(path)指定路径中删除目录，默认在当前目录下
'''
os.rmdir(r'.\justinTest')
'''

#os.stat(path)获取文件的属性
'''
print(os.stat('test.txt'))
'''
#os.rename(path1,path2)重命名
'''
os.rename('teat.py','test.py')
'''
#os.remove(path)删除普通文件
'''
os.remove('123.txt')
'''
#运行shell命令
'''
#可在任务管理器中查看执行文件名
os.system('notepad') #笔记本
os.system('taskkill /f /im notepad.exe') #关闭指定程序
os.system('write') #写字板
os.system('mspaint') #画板
os.system('msconfig') #系统设置
os.system('shutdown -s -t 500') #定时关机
os.system('shutdown -a') #取消定时关机
'''
'''
import os.path
'''
#查看当前或相对路径的绝对路径
'''
print(os.path.abspath('.'))
print(os.path.abspath(r'.\123.txt'))
'''
#路径的拼接
'''
path1 = r'D:\qian_feng_education\first_project'
path2 = 'knowledgePoints.py'
path3 = os.path.join(path1, path2) #注意参数2里，开始不要有\，之后可以有\
print(path3)
'''
'''
#注意参数2里，开始不要有\，之后可以有\
path1 = r'D:\qian_feng_education\first_project'
path2 = r'venv\123.txt'
path3 = os.path.join(path1, path2)
print(path3)
'''
#拆分最后一个文件或文件名，返回路径和文件名的元组
#作用：取绝对路径中的最后一个文件名或文件
#os.path.split(path)
'''
path = r'D:\qian_feng_education\first_project\qianfeng.py'
tempPath = os.path.split(path)
print(tempPath[0])
'''
#拆分获取扩展名并返回元组
'''
path = r'D:\qian_feng_education\first_project\qianfeng.py'
print(os.path.splitext(path))
'''
#获得文件名,即获取路径的最后一个节点，即获取最后一个\后的字符串
'''
path = r'D:\qian_feng_education\first_project\qianfeng.py'
print(os.path.basename(path)) -->qianfeng.py
'''
#获得文件的目录名
'''
path = r'D:\qian_feng_education\first_project\qianfeng.py'
print(os.path.dirname(path))
'''
#os.path.isdir(path)判断是否为目录
'''
path1 = r'D:\qian_feng_education\first_project'
path2 = r'D:\qian_feng_education\first_project\qianfeng.py'
print(os.path.isdir(path1))
print(os.path.isdir(path2))
'''
#判断指定目录下文件是否存在
'''
path = r'D:\qian_feng_education\first_project\qianfeng.py'
print(os.path.isfile(path)) 
'''
#判断目录文件是否存在
'''
path = r'D:\qian_feng_education\first_project'
print(os.path.exists(path))
'''
#获得文件大小(字节)
'''
path = r'D:\qian_feng_education\first_project\qianfeng.py'
print(os.path.getsize(path))
'''

#【day】shutil文件夹和文件操作

#简介：复制文件内容，粘贴到另外文件中
#copyfileobj(fsrc, fdst, length=16*1024)
#   参数
#       fsrc： 源文件
#       fdst： 复制至fdst文件
#       length： 缓冲区大小，即fsrc每次读取的长度
#   功能：将fsrc文件内容复制至fdst文件，length为fsrc每次读取的长度，用做缓冲区大小
#   优点：可设置每次读取的长度来控制缓存
#   缺点：必须先创建fdst文件
'''
import shutil
f1 = open("file.txt","r")
f2 = open("file_copy.txt","a+")
shutil.copyfileobj(f1,f2,length=1024)
'''

#简介：复制粘贴文件
#copyfile(src, dst)： 将src文件内容复制至dst文件
#   参数
#       src： 源文件路径
#       dst： 复制至dst文件，若dst文件不存在，将会生成一个dst文件；若存在将会被覆盖
#       follow_symlinks：设置为True时，若src为软连接，则当成文件复制；如果设置为False，复制软连接。默认为True。Python3新增参数
#   优点：当目标文件不存在时自动创建
#   缺点：不能设置读取长度，不能控制缓存
'''
import shutil
shutil.copyfile("file.txt","file_copy.txt")
'''

#简介：复制文件到粘贴生成另一文件或在指定文件夹中粘贴生成同名的另一文件
#copy(src, dst)： 将文件src复制至dst。dst可以是个目录，会在该目录下创建与src同名的文件，若该目录下存在同名文件，将会报错提示已经存在同名文件。权限会被一并复制。本质是先后调用了copyfile与copymode而已
#   参数
#       src：源文件路径
#       dst：复制至dst文件夹或文件
#       follow_symlinks：设置为False时，src, dst皆为软连接，可以复制软连接权限，如果设置为True，则当成普通文件复制权限。默认为True。Python3新增参数
#   优点：文件
'''
import shutil, os
shutil.copy("file.txt","file_copy.txt") #复制文件到粘贴生成另一文件
# 或者
shutil.copy("file.txt",os.path.join(os.getcwd(),"copy")) #在指定文件夹中粘贴生成同名的另一文件
'''

#简介：复制文件夹，粘贴生成文件夹及文件夹中所有内容
#copytree(src, dst, symlinks=False, ignore=None)： 拷贝文档树，将src文件夹里的所有内容拷贝至dst文件夹
#   参数
#       src：源文件夹
#       dst：复制至dst文件夹，该文件夹会自动创建，需保证此文件夹不存在，否则将报错
#       symlinks：是否复制软连接，True复制软连接，False不复制，软连接会被当成文件复制过来，默认False
#       ignore：忽略模式，可传入ignore_patterns()
#           ignore_patterns(*patterns)： 忽略模式，用于配合copytree()方法，传递文件将会被忽略，不会被拷贝
#               功能：指定文件名，被指定的文件不会被拷贝到另一个文件夹中
#               参数patterns：文件名称，元组
#       copy_function：拷贝文件的方式，可以传入一个可执行的处理函数，默认为copy2，Python3新增参数
#       ignore_dangling_symlinks：sysmlinks设置为False时，拷贝指向文件已删除的软连接时，将会报错，如果想消除这个异常，可以设置此值为True。默认为False,Python3新增参数
'''
import shutil,os
folder1 = os.path.join(os.getcwd(),"aaa")
# bbb与ccc文件夹都可以不存在,会自动创建
folder2 = os.path.join(os.getcwd(),"bbb","ccc")
# 将"abc.txt","bcd.txt"忽略，不复制
shutil.copytree(folder1,folder2,ignore=shutil.ignore_patterns("abc.txt","bcd.txt")
'''

#简介：删除文件夹及其中的文件和文件夹
#rmtree(path, ignore_errors=False, onerror=None)： 移除文档树，将文件夹目录删除
#   参数
#       ignore_errors：是否忽略错误，默认False
#       onerror：定义错误处理函数，需传递一个可执行的处理函数，该处理函数接收三个参数：函数、路径和excinfo
'''
import shutil,os
folder1 = os.path.join(os.getcwd(),"aaa")
shutil.rmtree(folder1)
'''

#简介：移动文件或文件夹
#   文件夹1到文件夹2
#        if 文件夹2存在：
#            文件夹1剪切到文件夹2下
#        if 文件夹2不存在
#            文件夹1更名为文件夹2
#   文件1到文件夹2
#        if 文件夹2存在：
#            文件1剪切到文件夹2下
#        if 文件夹2不存在
#            文件1更名为文件夹2的名字
#   文件1到文件2
#       文件1更名为文件2
#move(src, dst)： 将src移动至dst目录下。若dst目录不存在，则效果等同于src改名为dst。若dst目录存在，将会把src文件夹的所有内容移动至该目录下面
#   参数
#       src：源文件夹或文件
#       dst：移动至dst文件夹，或将文件改名为dst文件。如果src为文件夹，而dst为文件将会报错
#       copy_function：拷贝文件的方式，可以传入一个可执行的处理函数。默认为copy2，Python3新增参数
'''
import shutil,os
# 示例一，将src文件夹移动至dst文件夹下面，如果bbb文件夹不存在，则变成了重命名操作
folder1 = os.path.join(os.getcwd(),"aaa")
folder2 = os.path.join(os.getcwd(),"bbb")
shutil.move(folder1, folder2)
# 示例二，将src文件移动至dst文件夹下面，如果bbb文件夹不存在，则变成了重命名操作
file1 = os.path.join(os.getcwd(),"aaa.txt")
folder2 = os.path.join(os.getcwd(),"bbb")
shutil.move(file1, folder2)
# 示例三，将src文件重命名为dst文件(dst文件存在，将会覆盖)
file1 = os.path.join(os.getcwd(),"aaa.txt")
file2 = os.path.join(os.getcwd(),"bbb.txt")
shutil.move(file1, file2)
'''



#【day65】窗体控制
'''
import win32con
import win32gui
'''
#查出qq窗体编号
#win32gui.FindWindow(className, captionName)
#使用spy工具查询窗体的className和captionName
'''
QQwin = win32gui.FindWindow('TXGuiFoundation', 'QQ')
'''
#显示窗体
'''
win32gui.ShowWindow(qqwin, win32con.SW_SHOW)
'''
#隐藏窗体
'''
win32gui.ShowWindow(QQwin, win32con.SW_HIDE)
'''
'''
#一启动qq就不停的显示窗体和隐藏窗体
import time
import win32gui
import win32con
while True:
    QQwin = win32gui.FindWindow('TXGuiFoundation', 'QQ')
    win32gui.ShowWindow(QQwin, win32con.SW_HIDE)
    time.sleep(1)
    win32gui.ShowWindow(QQwin, win32con.SW_SHOW)
    time.sleep(1)
'''
#控制窗体的位置和大小
#win32gui.SetWindowPos(*args, **kwargs)
#参数1：控制的窗体对象
#参数2：大致方位
#   win32con.HWND_TOPMOST 靠上方
#参数3：位置的x
#参数4：位置的y
#参数5：窗口的长度
#参数6：窗口的宽度
#参数7：一直显示
#   win32con.SWP_SHOWWINDOW
'''
import win32gui
import win32con
QQwin = win32gui.FindWindow('TXGuiFoundation', 'QQ')
win32gui.SetWindowPos(QQwin, win32con.HWND_TOPMOST, 100, 100, 300, 300, win32con.SWP_SHOWWINDOW)
'''
'''
#控制qq窗体，随机出现在屏幕的任意位置
'''
'''
import win32gui
import win32con
import random
while True:
    x = random.randrange(900)
    y = random.randrange(600)
    QQwin = win32gui.FindWindow('TXGuiFoundation', 'QQ')
    win32gui.SetWindowPos(QQwin, win32con.HWND_TOPMOST, x, y, 300, 300, win32con.SWP_SHOWWINDOW)
'''

#【day66】win32com.client系统客户端
#语音合成

#让电脑给你读文章
'''
import win32com.client
dehua = win32com.client.Dispatch('SAPI.SPVOICE')
dehua.Speak("She's my master's wife.")
'''

#语音模块(语音识别)
#语音模块加载报错处理方法
#C:\Users\surface\Anaconda3\Lib\site-packages\pythonwin\Pythonwin.exe
#Tools->commakepy->utility->Microsoft speech object library [5.4]
'''
import win32com.client
from win32com.client import constants
import os
import win32com.client
import pythoncom

mia = win32com.client.Dispatch('SAPI.SPVOICE')  # 生成speaker对象
mia.Speak('你好，我的名字叫米亚，很高兴认识你')

class SpeechRecognition:
    def __init__(self, wordsToAdd):
        self.speaker = win32com.client.Dispatch('SAPI.SpVoice')
        self.listener = win32com.client.Dispatch('SAPI.SpSharedRecognizer')
        self.context = self.listener.CreateRecoContext()
        self.grammar = self.context.CreateGrammar()
        self.grammar.DictationSetState(0)
        self.wordsRule = self.grammar.Rules.Add('wordsRule', constants.SRATopLevel + constants.SRADynamic, 0)
        self.wordsRule.Clear()
        [self.wordsRule.InitialState.AddWordTransition(None, word) for word in wordsToAdd]
        self.grammar.Rules.Commit()
        self.grammar.CmdSetRuleState('wordsRule', 1)
        self.grammar.Rules.Commit()
        self.eventHandler = ContextEvents(self.context)
        self.say('Started successfully')
    def say(self, phrase):
        self.speaker.Speak(phrase)

class ContextEvents(win32com.client.getevents('SAPI.SpSharedRecoContext')):
    def OnRecognition(self, StreamNumber, StreamPosition, RecognitionType, Result):
        newResult = win32com.client.Dispatch(Result) #生成语音识别对象
        voiceCommand = newResult.PhraseInfo.GetText()  # 生成语音转文字变量
        print('我说:', voiceCommand)
        for k in dictKeyWord():
            if voiceCommand == k:
                v = dictKeyWord()[k]
                mia.Speak(v)
                print(v)

def dictKeyWord():
    dict1 = {'陈贤民是谁': '他是世界上最好的人',
             '你的爸爸是谁': '我爸爸叫陈艺龙，他是一个超厉害的人',
             '能给我讲一个故事吗': '我讲故事可是要收钱的哦',
             '彭淑贤是谁': '她是我爸爸的老婆'}
    return dict1

if __name__ == '__main__':
    wordsToAdd = list(dictKeyWord())
    speechReco = SpeechRecognition(wordsToAdd)
    while True:
        pythoncom.PumpWaitingMessages()
'''

#【day67】进程模块
'''
import win32process
import win32con
import win32gui
import win32api
import ctypes
'''
#通过程序来修改内存值
#思路：
#通过spy找到游戏窗体进程id，再通过进程id找到对应的内存地址，最后修改内存值
'''
#修改植物大战僵尸的阳光数
PROCESS_ALL_ACCESS = (0X000F0000|0X00100000|0XFFF) #最高全向变量，最高权限是通过二进制数进行|是位运算的到的值
win = win32gui.FindWindow('MainWindow', '植物大战僵尸中文版') #找窗体
hid, pid = win32process.GetWindowThreadProcessId(win) #根据窗体找到进程号（任务管理器中的pid号）
p = win32api.OpenProcess(PROCESS_ALL_ACCESS, False, pid) #以最高权限打开pid进程
md = ctypes.windll.LoadLibrary(r'c:\windows\system32\kernel32.dll') #md——mydata加载内核模块，内核模块存放于c:\windows\system32\kernel32.dll
data = ctypes.c_long() #定义一个数据类型为c语言中的长整型数据类型的变量，用于存放指定内存中的值
md.ReadProcessMemory(int(p), 485053864, ctypes.byref(data), 4, None) #用内核模块读取p进程的485053864内存，将值保存到data中，保存4个字节，None表示内存信息错误我们不处理。内存地址通过memsearch软件查到
newData = ctypes.c_long(10000) #设置新值
md.WriteProcessMemory(int(p), 485053864, ctypes.byref(newData), 4, None)#用内核模块将新值写入p进程485053864内存，写入4个字节，None表示内存信息错误我们不处理
'''

#【day68】递归
#递归调用:一个函数调用了自身，称递归调用
'''
def func1():
    print('&&&&')
def func2():
    func1()
func1()
'''
#递归函数：一个会调用自身的函数称为递归函数
'''
def func1():
    func1()
'''
#凡是循环能干的事，递归都能干，但是不是一般人能写出来

#写递归的方方式：
#   1、写出临界条件
#   2、找这一次和上一次的关系
#   3、假设当前函数已经能用，调用自身计算上一次的结果，再求出本次的结果

''' 
#输入一个数（大于等于1），求1+2+3+...n的和
'''
'''
方式一：函数写
def func1(n):
    sum = 0
    for i in range(1, n + 1):
        sum += i
    return sum
result = func1(5)
print(result)
'''
'''
#方式二：写递归
1+2+3+4+5=?
sum1(1) + 0 = sum2(1)
sum1(1) + 2 = sum1(2)
sum1(2) + 3 = sum1(3)
sum1(3) + 4 = sum1(4)
sum1(4) + 5 = sum1(5)
5 + sum1(4)
5 + 4 + sum1(3)
5 + 4 + 3 + sum1(2)
5 + 4 + 3 + 2 + sum1(1)
5 + 4 + 3 + 2 + 1 = 15
def sum1(n):
    if n == 1:
        return 1
    else:
        return n + sum1(n-1)
result = sum1(5)
print('result=', result)
'''
'''
递归求解5!*
def func(n):
    if n == 1:
        return 1
    else:
        return n * func(n - 1)
result = func(5)
print(result)
'''

#【day69】栈数据结构与队列数据结构
#栈：可以存数据，
#栈存数据的特点：数据先进后出，数据后进先出
#用列表模拟栈结构
'''
stack = []
#压栈(向栈里存数据)
stack.append('a')
print(stack)
stack.append('b')
print(stack)
stack.append('c')
print(stack)
#出栈(在栈取数据)
result1 = stack.pop()
print(result1, stack)
result2 = stack.pop()
print(result2, stack)
result3 = stack.pop()
print(result3, stack)
'''
#队列：可以存储数据
#队列存储数据的特点：数据先进先出
'''
#创建队列方法一
import collections
#在数据集合模块中封装了队列数据结构
#创建一个队列
queue = collections.deque()
print(queue, type(queue))
#进队——存数据
queue.append('A')
print(queue)
queue.append('B')
print(queue)
queue.append('C')
print(queue)
#出队——取数据
result1 = queue.popleft()
print(result1, queue)
result2 = queue.popleft()
print(result2, queue)
result3 = queue.popleft()
print(result3, queue)
'''

#创建队列方法二
'''
import queue
q = queue.Queue()
q.put() #进队——添加元素
q.get() #出队——取出元素
'''
#栈、队列、递归，可以去循环的处理一些事情

#【day70】递归遍历目录
#创建多层级文件夹
'''
import os
path = r'D:\qian_feng_education\creatDirTest\{}'
for i in range(5):
    path1 = path.format(i)
    os.makedirs(path1)
    #创建三级目录
    path2 = os.path.join(path1, str(i))
    os.makedirs(path2)
    #创建二级目录下的普通文件
    fileName1 = '{}.txt'.format('123')
    path3 = os.path.join(path1, fileName1)
    path3.format(i)
    with open(path3, 'w') as f1:
        f1.write('i love you, my wife!')
    #创建三级目录下的普通文件
    fileName2 = '{}.txt'.format('456')
    path4 = os.path.join(path2, fileName2)
    with open(path4, 'w') as f2:
        f2.write('i love you, my daughter')
'''
#递归遍历目录
'''
import os
def getAllDirRecursion(path1, sp = ''):
    #得到指定目录下的所有文件名
    sp += '  '
    fileList = os.listdir(path1)
    #处理每一个文件
    for fileName in fileList:
        fileAbsPath = os.path.join(path1, fileName)
        if os.path.isdir(fileAbsPath): #os.path.isdir(绝对路径)， 路径的拼接：os.path.join(paht, fileName)
            print(sp + '目录：', fileName)
            #递归调用
            getAllDir(fileAbsPath, sp)
        else:
            print(sp + '普通文件：', fileName)
path = r'D:\qian_feng_education\first_project\temp'
getAllDirRecursion(path)
'''

#【day71】深度遍历——模拟栈结构遍历目录(后期可用于爬虫，遍历网页)
'''
import os
def getAllDirDeep(path):
    #列表模拟栈
    stack = []
    #压栈
    stack.append(path)
    #处理栈：当栈为空的时候，结束循环
    while len(stack) != 0:
        #路径出栈
        dirPath = stack.pop()
        #获取指定目录下所有文件名
        filesList = os.listdir(dirPath)
        #循环处理所有文件
        for fileName in filesList:
            #合成得文件的绝对路径
            fileAbsPath = os.path.join(dirPath, fileName)
            #是文件夹的路径则压栈
            if os.path.isdir(fileAbsPath):
                print('目录:' + fileName)
                stack.append(fileAbsPath)
            #否则打印
            else:
                print('普通文件:' + fileName)

path = r'D:\qian_feng_education\first_project\temp'
getAllDirDeep(path)
'''

#【day72】广度遍历——队列结构遍历目录(后期可用于爬虫，遍历网页)
'''
import os
import collections
def getAllDirQueue(path):
    queue = collections.deque()
    #进队
    queue.append(path)
    while len(queue) != 0:
        #出队数据
        dirPath = queue.popleft()
        #找出所有的文件
        filesList = os.listdir(dirPath)
        for fileName in filesList:
            #合成绝对路径
            fileAbsPath = os.path.join(dirPath, fileName)
            #判断是否为目录，是目录进队，不是目录打印
            if os.path.isdir(fileAbsPath):
                queue.append(fileAbsPath)
                print('目录：', fileName)
            else:
                print('普通文件：', fileName)
path = r'D:\qian_feng_education\first_project\temp'
getAllDirQueue(path)
'''

#【day72】time模块
#UTC(世界协调时间):格林尼治天文时间，世界标准时间
#在中国来说是UTC+8
#DST(夏令时)：节约能源而人为规定的时间制度，在夏季时间调快一个小时

#时间的表示形式
'''
1、时间戳：以整型或浮点型表示时间的一个以秒为单位的时间间隔
    这个时间间隔的基础值是从1970年1月1号零点开始算起，即1970年1月1日00:00:00开始距离现在有多少秒，就是我们当前时间的多少秒
2、元组：一种Python的数据结构表示，这个元组有9个整型内容
    year month day hours minutes seconds weekday Juliaday flag(1（夏令时）或-1（根据当前日期自己判断）或0（正常格式）)
3、格式化字符串：
    %a 本地(locale)简化星期名称
    %A 本地完整星期名称
    %b 本地简化月份名称
    %B 本地完整月份名称
    %c 本地相应的日期和时间表示
    %d 一个月中的第几天（01-31）
    %H 一天中的第几个小时（24小时制，00-23）
    %I 第几个小时（12小时制，01-12）
    %j 一年中的第几天（001-366）
    %m 月份（01-12）
    %M 分钟数（00-59）
    %p 本地am或者pm的相应符
    %S 秒（01-61）
    %U 一年中的星期数。（00-53星期天是一个星期的开始。）第一个星期天之前的所有天数都放在第0周
    %w 一个星期中的第几天（0-6，0是星期天）
    %W 和%U基本相同，不同的是%W以星期一为一星期的开始
    %x 本地相应日期
    %X 本地相应时间
    %y 去掉世纪的年份（00-99）
    %Y 完整的年份
    %Z 时区的名字（如果不存在为空字符）
    %% ‘%’字符
'''

#时间戳、本地时间元组、字符串之间的转化
'''
import time
#以浮点数形式，返回当前时间的时间戳
c = time.time()
print(c, type(c))
#将时间戳转化为UTC时间元组
t = time.gmtime(c)
print(t, type(t))
#将时间戳转化为本地时间元组
b = time.localtime(c)
print(b, type(b))
#将本地时间元组转化为时间戳
m = time.mktime(b)
print(m, type(m))
#将本地时间元组转化为字符串
s = time.asctime(b)
print(s)
#将时间戳转化为字符串
p = time.ctime(c)
print(p)
#将本地时间元组转化为给定格式时间的字符串，参数2为时间元组，如果没有参数2，则默认转化当前时间
b = time.localtime(time.time())
q = time.strftime('%Y-%m-%d %X', b) #%X = %H:%M:%S
print(q, type(q))
#将时间字符串转为时间元组
w = time.strptime(q, '%Y-%m-%d %X') #参数1：时间字符串，参数2：用这样的格式去读q
print(w)
'''
#time的方法
#延迟一个时间，整型或者浮点型
'''
time.sleep(1) #参数1的单位为秒
'''
#返回当前程序的cpu时间，cpu运行这段代码用了多少时间
'''
y1 = time.perf_counter()
print(y1)
time.sleep(1)
y2 = time.perf_counter()
print(y2)
time.sleep(1)
y3 = time.perf_counter()
print(y3)
'''
#性能测试
'''
import time
sum1 = 0
for i in range(100000000):
    sum1 += i
t = time.perf_counter()
print(t)
'''

#【day73】datetime模块
#datetime比time高级了不少，可以理解为datetime基于time进行了封装，提供了更为实用的函数，datetime模块的接口更直观，更容易调用
#模块中的类：
#   datetime 同时有时间和日期
#   timedelta 主要用于计算时间的跨度（时间差）
#   tzinfo 时区相关
#   time 只关注时间
#   date 只关注时区
'''
import datetime
'''
#datetime.datetime的方法
'''
#获取当前时间
d1 = datetime.datetime.now()
print(d1, type(d1))
#获取指定时间
d2 = datetime.datetime(1999, 10, 1, 10, 28, 25, 123123)
d2 = datetime.datetime(year=2017, month=7, day=17)
print(d2)
#将datetime对象时间转为格式化字符串
d3 = d1.strftime('%Y-%m-%d %X')
print(d3, type(d3))
#将格式化字符串转化为datetime对象时间
d4 = datetime.datetime.strptime(d3, '%Y-%m-%d %X')#转换的格式要与字符串格式一致，即d3的格式为'%Y-%m-%d %X'，参数2的格式为'%Y-%m-%d %X'
print(d4)
#datetime对象时间可以进行运算,得到一个timedelta时间差对象
d5 = d1 - d2
print(d5, type(d5))
#提取间隔的天数
print(d5.days)
#提取间隔天数除外的秒数
print(d5.seconds)
'''

#【day74】日历模块
'''
import calendar
'''
#返回指定年的日历
'''
m = calendar.calendar(2019)
print(m)
'''
#返回指定某年某月的日历
'''
m = calendar.month(2019, 6)
print(m, type(m))
'''
#闰年返回True,平年返回False
'''
m = calendar.isleap(2000)
print(m)
'''
#返回指定年月的weekday的第一天和这个月所有天数
'''
m = calendar.monthrange(2019, 3)
print(m)
'''
#返回某个月以每一周为元素的列表
'''
print(calendar.monthcalendar(2019, 6))
'''

'''
#作业：
1、在开房信息中输入一个人的名字显示全部的开房信息
import os
name = input('输入你的名字：')
path = r'D:\迅雷下载\数据\2000W开房数据'
filesName = os.listdir(path)
for fileName in filesName:
    fileAbsPath = os.path.join(path, fileName)
    with open(fileAbsPath, 'r', encoding='utf-8') as f:
        list1 = f.readlines()
        for i in list1:
            if name in i:
                print(i)

2、在多层文件夹及文件中提取邮箱数据，并分类整理邮箱，将相同邮箱放入新建的文件夹中
import os

def rwfile(fileAbsPath, path2):
    with open(fileAbsPath, 'r', encoding='utf-8') as f:
        contentsList = f.readlines()
    sum1 = 0
    for content in contentsList:
        emailAddName = content.split('@')[1].split('.')[0]
        print(emailAddName)
        dirNamePath = os.path.join(path2, emailAddName)
        if not os.path.isdir(dirNamePath):
            os.mkdir(dirNamePath)
        emailFile = os.path.join(dirNamePath, emailAddName) + '.txt'
        print(emailFile)
        with open(emailFile, 'a', encoding='utf-8') as f:
                f.write(content)
                sum1 += 1
                print('成功写入数据第{}条'.format(sum1))

# import collections
def getAllTxt(path1, path2):
    stack = []  # queue = collections.deque()
    stack.append(path1)
    while len(stack) != 0:
        filePath = stack.pop()  # queue.popleft()
        filesNamelist = os.listdir(filePath)
        for fileName in filesNamelist:
            fileAbsPath = os.path.join(filePath, fileName)
            if os.path.isdir(fileAbsPath):
                stack.append(fileAbsPath)
                print('目录:', fileName)
            elif 'txt' in fileAbsPath:
                print('txt文件:', fileName)
                rwfile(fileAbsPath, path2)

path1 = r'D:\qian_feng_education\testdata'
path2 = r'D:\qian_feng_education\first_project\emailType'
getAllTxt(path1, path2)
'''

#【day74】模块概述
'''
概述：目前代码比较少，写在一个文件中还体现不出什么缺点，但随着代码量越来越多，代码就越来越难以维护。
为了解决难以维护的问题，我们把很多相似功能的函数分组，分别放到不同的文件中去，这样每个文件中的内容相对较少，
而且对于每个文件的大致功能可用文件名来体现。很多编程语言都是这么来组织代码结构。一个.py文件就是一个模块

优点：
1、提高代码的可维护性
2、提高了代码的复用度，当一个模块完毕，它可以被多个地方引用
3、引用其他的模块（内置模块、三方模块、自定义模块）
4、避免函数名和变量名的冲突
'''

#【day75】引入标准模块

#引入内置模块
#import语句
#格式：import module1[, module2][, module3][,......modulen]
'''
import os, time, random
'''
'''
import sys
'''
#sys.argv属性能接受黑屏终端传入的信息
#在黑屏终端中输入：
#c:\python 模块的绝对路径 justing 30 male
#sys.argv = ['模块的绝对路径', 'justing', '30', 'male']
'''
print(sys.argv)
for i in sys.argv:
    print(i)
name = sys.argv[1]
age = sys.argv[2]
gender = sys.argv[3]
print(name, age, gender)
'''
#黑屏终端执行，自动查找所需模块的路径的列表
'''
print(sys.path)
'''

#【day76】使用自定义模块(将自定义模块放在执行文件同级)
#创建自定义模块
#   一个.py文件，就是一个自定义模块
#引入自定义模块，直接引用其文件名即可，不用加.py
#注意：一个模块只能被引入一次，不管你执行了多少次import语句
'''
import myModule
'''
#引入模块后，就能使用模块中的内容了
#格式:模块名.函数名/变量名
'''
myModule.sayHi()
print(myModule.yy)
'''
#from...import...语句
#作用：从模块中导入一个指定的部分到当前命名空间,未引用的方法不能使用
#格式：from moduleName import name1[, name2][, name3]...
'''
from myModule import sayHi
sayHi()
'''
'''
程序内容的函数可以将模块中的同名函数覆盖，注意定义名称不要重复
def sayHi():
    print('hi，hi')
'''
#from......import *
#作用：把一个模块中的所有方法都导入当前命名空间
'''
from myModule import *
'''

#【day77】__name__属性：
#模块就是一个可执行的.py文件，一个模块被另一个程序引入。我们不想让模块中的某些代码执行，我们可以用__name__属性来使程序仅调用模块中的一部分
#每一个模块都有一个__name__属性，当其值等于'__main__'时，表明该模块自身在执行。否则被引入其他文件
#当前文件如果为程序的入口文件(当前执行文件)，则__nam__属性的值为__main__
'''
import myModule
'''
'''
当__name__属性作为模块变量导入时，__name__的值为模块名
'''


#【day78】包
#思考：如果不同的人编写的模块同名了怎么办？
#解决：为了解决模块命名的冲突，引入了按目录来组织模块的方法，这些目录称为包
#特点：引入了包以后，只要顶层的包不与其他人让生冲突，那么模块都不会与别人的模块发生冲突
#注意：一个目录中只有包含了'__init__.py'文件才能被认作是一个包
#__init__.py主要是为了避免一些滥竽充数的名字，基本上目前这个文件中什么也不用写
'''
import package1.myModule
import package2.myModule
import package1.my_module
package1.myModule.sayHi()
package2.myModule.sayHi()
package1.my_module.sayHi()
'''

#【day79】第三方模块
#安装第三方模块
#pip -V
'''
mac:无需安装，自带
Linux:无需安装，自带
windows:安装Python时，勾选pip和add python.exe to Path
'''
#要安装一个第三方模块，需要知道模块的名字
#Pillow 非常强大的处理图像的工具库
#黑屏终端输入 pip install Pillow
#windows如果报错,需要升级pip：输入pip install -upgrade pip
#引入第三方模块
'''
from PIL import Image
#打开图片
im = Image.open('timg.jpg')
#查看图片信息
print(im.format, im.size, im.mode)
#设置图片大小
im.thumbnail((50, 50))
#保存成新图片
im.save('temp.jpg', 'JPEG')
'''

#【day80】面向过程、面向对象
#软件编程的实质：软件编程就是将我们的思维转变成计算机能够识别语言的一个过程
#什么是面向过程：
#   自上而下顺序执行，逐步求精；
#   其实程序是按照功能分为若干个基本模块，这些模块形成一个树状结构
#   各模块之间的关系尽可能简单，在功能上相对独立
#   每一模块内部均是由顺序、选择和循环三种基本结构组成
#   其模块化实现的具体方法是使用子程序
#   程序流程在写程序时就已决定

#什么是面向对象：
#   把数据及对数据的操作方法放在一起，作为一个相互依赖的整体——对象
#   对同类对象抽象出其共性，形成类
#   类中的大多数数据，只能用本类的方法进行处理
#   类通过一个简单的外部接口与外界发生关系，对象与对象之间通过消息进行通信
#   程序流程由用户在使用中决定

#深度理解面对对象：
#面向对象是相对面向过程而言
#面向对象和面向过程都是一种思想
#面向过程
#   强调的是功能行为
#   关注的是解决问题需要哪些步骤
#面向对象
#   将功能封装对象，强调具备了功能的对象
#   关注的是解决问题需要哪些对象
#面向对象基于面向过程的

'''
把大象放进冰箱
面对过程
    打开冰箱、把大象放进冰箱、关上冰箱
面对对象
    寻找一个能把大象放进冰箱的人，知道他能打开冰箱门，知道他能把大象装进冰箱，知道他能关上冰箱门，让他去执行就行了，至于他怎么打开冰箱，怎么装大象，怎么关门，我们不关心
'''
#面向对象的特点
#   是一种符合人们思考习惯的思想
#   可以将复杂的事情简单化
#   将程序员从执行者转化成了指挥者
#完成需求时：
#   先要去找具体所需的功能的对象使用
#   如果该对象不存在，那么创建一个具有所需功能的对象

#现实生活中我们是如何运用面对对象
#   打电话
#   媳妇：属性（身高、体重、身材、性别），功能（揉揉肩、拖鞋）
#   老公：属性（颜值、身高、体重、财富），功能（拎包、刷卡、逛街）
#   车：属性（颜色，轮子个数），功能（跑）

#类与对象的关系
#   使用计算机语言就是不断的在描述现实生活中的事物
#   Python中描述事物通过类的形式体现，类是具体事物抽象，概念上的一个定义
#   对象即是该类事物实实在在存在的个体

#类的定义
#生活中描述事物无非就是描述事物的名称、属性、行为：如：人的身高、体重等属性，有说话，打架等行为
#Python中用类来描述事物也是如此
#   属性：对应类中的成员变量
#   行为：对应类中的成员方法
#定义类其实在定义类中的成员（成员变量和成员方法）
#拥有相同（或者类似）属性和行为的对象都可以

#类的设计
#只关心3样东西
#   事物名称（类名）：人（person）
#   属性：身高（height）、年龄（age）
#   行为（功能）：跑（run）、打架（fight）

#【day81】创建类
#设计类
'''
类名：首字母大写、其他遵循驼峰原则，见名知意
属性：遵循驼峰原则，见名知意
行为（功能/方法）：遵循驼峰原则，见名知意

类名：wife
属性（身高、体重、身材、性别）
功能（揉揉肩、拖鞋）
'''
#创建类
'''
类：一种数据类型，本身并不占内存空间，跟以前所学的number，string，boolean等类似。
用类创建实例化对象（变量），对象占内存空间
'''
#格式：
'''
class 类名(父类列表):
    属性
    行为
'''
#object:基类、超类，所有类的父类，一般没有合适的父类就写object
'''
class Person(object):
    #定义属性(定义变量)
    name = 'mia'
    age = 9
    height = 150
    weight = 40
    #定义方法(定义函数)
    #注意：方法的参数必须以self当第一个参数
    #self代表类的实例（某个对象）
    def run(self):
        print('run')
    def eat(self, food):
        print('eat', food)
    def openDoor(self):
        print('我已经打开了冰箱门')
    def fillElephant(self):
        print('我已经把大象装进冰箱了')
    def closeDoor(self):
        print('我已经关上了冰箱门')
'''
#实例化对象
'''
格式：对象名(变量名) = 类名(参数列表)
注意：没有参数，小括号也不能省略
'''
'''
#有同一个类实例化出来的两个对象，是完全不同的两个个体，他们的变量名在栈区的id地址不同，在堆区保存的属性和方法的内存地址也不同
per1 = Person()
print(per1, type(per1), id(per1))
per2 = Person()
print(per2, type(per2), id(per2))
'''
#【day80】使用类实例化对象
#实例化对象要占内存空间，即有内存空间地址
'''
#格式：对象名.属性名
per = Person()
print(per)
'''
#访问属性
'''
print(per.name, per.age, per.height, per.weight)
'''
#属性赋值：
'''
格式：对象名.属性名 = 新值
per.name = 'justin'
per.age = 30
per.height = 170
per.weight = 70
print(per.name, per.age, per.height, per.weight)
'''
#访问方法
'''
格式：对象名.方法名(参数列表)
per.openDoor()
per.fillElephant()
per.closeDoor()
per.eat('apple')
'''
#问题：目前来看Person创建的所有对象属性都是一样的，不符合现实思维逻辑
'''
per1 = Person()
per2 = Person()
print(per1.name)
print(per2.name)
#用构造函数解决以上问题
'''
#【day81】对象的初始状态(构造函数)
#构造函数：__init__()
#意义：在使用类创建对象的时候，创建不同属性的对象。构造函数会自动调用
#注意：如果在定义类时未写出构造函数，默认会自动添加一个空的构造函数,即
'''
def __init__(self):
    pass
'''
'''
class Person(object):
    def __init__(self):
        print('实例化对象是，自动执行此函数')
per = Person()
'''
'''
class Person(object):
    #定义属性
    def __init__(self, name, age, height, weight):
        #print(name, age, height, weight)
        self.name = name #self相当于是正在创建的对象，即per
        self.age = age
        self.height = height
        self.weight = weight
per1 = Person('justin', 30, 170, 70)
print(per1.name, per1.age, per1.height, per1.weight)
per2 = Person('mia', 9, 150, 40)
print(per2.name, per2.age, per2.height, per2.weight)
'''

#【day82】self
#self代表类的实例，而非类
#哪个对象调用方法，那么该方法中的self就代表哪个对象
'''
class Person(object):
    def __init__(self, name, age, height, weight):
        self.name = name #self相当于是正在创建的对象，即per
        print(self.name) #self.name = per.name
        self.age = age #给当前实例对象赋值，可以将self看成一个变量，实例化不同对象的时候，self代表不同对象
        self.height = height
        self.weight = weight
    def openDoor(self):
        print(self) #per实例对象调用openDoor方法，此时的self就是per对象
    #self不是关键字，换成其他的标识符也是可以的，但是帅的人都是用self
    def closeDoor(a):
        print(a) ##per实例对象调用openDoor方法，此时的a就是per对象
per = Person('justin', 30, 170, 70)
per.openDoor()
per.closeDoor()
'''
#self.__calss__ 代表类名以及使用
'''
class Person(object):
    def __init__(self, name):
        self.name = name
    def openDoor(self):
        print(self.__class__)
        per1 = self.__class__('mia') #self.__class__ = Person,so 执行了Person('mia')
        print(per1)
per = Person('justin')
per.openDoor()
'''

#【day83】析构函数
#析构函数：__del__()
#释放对象时自动调用,释放对象即删除变量
#程序结束时会Python垃圾回收机制能自动释放对象，释放对象析构函数自动执行
'''
class Person(object):
    def __init__(self, name, age):
        self.name = name
        self.age = age
    def run(self):
        print('run')
    def eat(self, food):
        print('eat', food)
    def __del__(self):
        print('这里是析构函数')
per = Person('justin', 30)
'''
#程序不结束，手动释放对象时，析构函数自动执行
'''
class Person(object):
    def __init__(self, name, age):
        self.name = name
        self.age = age
    def run(self):
        print('run')
    def eat(self, food):
        print('eat', food)
    def __del__(self):
        print('这里是析构函数')
per = Person('justin', 30)
del per #在程序结束，或手动释放对象时，析构函数自动执行
while True:
    pass
'''
#在函数里定义的对象，会因函数调用结束而释放对象，析构函数自动执行，这样可以用来减少内存空间的浪费
'''
class Person(object):
    def __init__(self, name, age):
        self.name = name
        self.age = age
    def run(self):
        print('run')
    def eat(self, food):
        print('eat', food)
    def __del__(self):
        print('这里是析构函数')

def function():
    per = Person('justin', 30)
function()

while True:
    pass
'''

#【day84】重写__repr__与__str__函数
#重写：将函数重新定义写一遍
#__str__(self):在调用print打印对象时自动调用，是给用户用的，是一个描述对象的方法
#__repr__(self):是给机器用的，在Python解释器里面直接敲对象名再回车后调用的方法，即在黑屏终端的Python交互模式时，输入实例化对象回车后，调用__repr__():函数
#注意：在没有str时，且有repr，Python内部会执行str = repr，以后需要描述一个对象时，只需要使用str即可
'''
#优点：当一个对象的属性值很多，并且都需要打印，重写了__str__方法后，简化了代码
class Person(object):
    def __init__(self, name, age):
        self.name = name
        self.age = age
    #def __str__(self): #重写打印对象时返回的字符串
        #return '%s-%d' % (self.name, self.age)
per = Person('justin', 30)
print(per.name, per.age) #如果我们的属性有10个以上，将属性全部打印出来不美观
#需求：通过打印实例化对象名，就能打印实例化对象的属性
print(per)
'''


'''
作业：人开枪射击子弹，开枪少一发，子弹为零，再开枪不发射子弹
'''
'''
方法一：
import time
class Gun(object):
    def __init__(self, name, bulletsNumber, speed):
        self.gunName = name
        self.bulletsNumber = bulletsNumber
        self.speed = speed
    def shoot(self):
        while self.bulletsNumber > 0:
            time.sleep(self.speed)
            print('bang')
            self.bulletsNumber -= 1
        print('piang')
        print('%s的子弹打完' % self.gunName)
#gun1 = Gun('来复枪',7 ,1)
#gun1.shoot()
#gun2 = Gun('ak',50 ,0.1)
#gun2.shoot()
'''
'''
方法二：
人
类名：Person
属性：gun
行为：fire,fillBullet

弹夹
类名：Cartridge
属性：BulletCount
行为：

枪
类名：Gun
属性：cartridge
行为：shoot
from Person import Person
from Gun import Gun
from Cartridge import Cartridge
cartridge = Cartridge(5)
gun = Gun(cartridge) #实例对象被当做参数传入一个类中时，该类内部可以调用其属性和方法
per = Person(gun)
per.fire()
per.fire()
per.fire()
per.fire()
per.fire()
per.fire()
per.fillBullte(3)
per.fire()
per.fire()
per.fire()
per.fire()

'''

#【day85】访问限制
'''
class Person(object):
    def __init__(self, name, age, money, property):
        self._name = name
        self.__age__ = age #注意仅只有前面有两个下划线才能算是私有属性
        self.money = money
        self.__property = property #__property = _Person__property
    def run(self):
        print('run')
    def eat(self, food):
        print('eat', food)
    def query(self):
        print(self.__property)
    #通过自定义的方法，实现对私有属性的赋值与取值
    #私有属性一般都会给它写set方法和get方法
    def setProperty(self, property):
        #数据过滤
        if property < 0:
            self.__property = 0
        self.__property = property
    def getProperty(self):
        return self.__property
per = Person('justin', 30, 10000000, 1000000)
print(per.money) #不安全
'''
#如果要让内部属性不被外部直接访问,就在属性前面加两个下划线__,在Python中如果在属性前加两个下划线，这个属性就变成了私有属性
'''
print(per.__property)
'''
#私有属性只能在内部进行使用
'''
per.query()
per.setProperty(10)
print(per.getProperty())

#不能直接访问per.__property是因为Python解释器把per.__property变成了_Person__property,仍然可以用_Person__property访问私有属性，但是强烈建议不要这样干，因为不同版本的解释器可能解释出来的变量名不一定是这个。
per._Person__property = 1
print(per.getProperty())
#注意：在Python中__xxx__属于特殊变量，可以直接访问
print(per.__age__)
#注意：在Python中_xxx可以直接访问，但是按照约定的规则，当我们看到这样的变量时，意思是：“我虽然可以被访问，但是不要访问我”
print(per._name)
'''

#【day86】继承
#继承：有两个类，A类和B类，当我们说A类继承自B类时，那么A类就拥有了B类中的所有属性和方法
#object类是所有类的父类，还可以称为基类或超类，__init__属性就是继承自object
#继承者称为子类，被继承这称为父类
#继承的优点：1、简化代码，减少冗余
#          2、提高了代码的健壮性 #修改一个父类就能修改增加所有子类的属性和方法
#          3、提高了代码的安全性 #父类的属性不暴露，不能直接修改子类的属性
#          4、是多态的前提
#继承的缺点：1、耦合与内聚是描述类与类之间关系的。耦合性越低内聚性越高代码越好，但是继承正好相反（耦合——联系，内聚——功能）

#【day87】单继承的实现
'''
from student import Student
student = Student('mia', 18, 1000, 10000)
print(student.name, student.age)
#子类独立的属性
print(student.stuId)
#私有属性不能直接被子类继承，
print(student.__money)
student.stufunc()
#但可以通过继承共有方法访问私有属性
print(student.getMoney())
'''

'''
#作业：盖房子，用面对对象的思维
工人
类名：Worder
属性：砖、水泥
行为：和水泥、砌砖

材料
类名：MaterialScience
属性：砖、水泥
行为：
from materialScience import MaterialScience
from worker import Worker
materialScience = MaterialScience(500, 10)
worker = Worker(materialScience)
worker.mixCement()
worker.bricklaying()
worker.mixCement()
worker.bricklaying()
'''

#【day88】多继承的实现
#Children类继承了Father类的money以及Mother类的faceValue，还有两个父类的方法
'''
from children import Children
children = Children(10000000, 100)
print(children.money, children.faceValue)
children.play()
children.eat()
children.function()
'''

#【day89】多态
#一种事物的多种形态
#最终目标：人可以喂任何一种动物
'''
from cat import Cat
from mouse import Mouse
tom = Cat('tom')
jerry = Mouse('jerry')
#tom.eat()
#jerry.eat()
#思考：我在添加100中动物，也都有name属性和eat方法
#用继承解决以上问题，定义一个有name属性和eat方法的animal类，让其他动物都继承animal的属性和方法
#需求：定义一个男孩，可以喂猫和老鼠吃东西
from boy import Boy
boy = Boy()
#boy.feedCat(tom)
#boy.feedMouse(jerry)
#思考：如果人要喂100中动物，需要些100个feed方法吗
#tom和jerry都是继承自animal，这就是多态。即我喂猫喂老鼠喂大象，喂得累，我就只为动物，因为他们都是动物的多种形态。这就是多态
boy.feedAnimal(tom)
boy.feedAnimal(jerry)
'''

#【day90】对象属性与类属性
#类属性：用类名来调用
'''
class Person(object):
    name = 'justin'  # 类属性
    def __init__(self, name):
        self.name = name #对象属性
print(Person.name) #类属性用类名来调用
per = Person('mia') 
print(per.name) #调用对象属性
'''

#获取类属性的值
'''
a = Person()
b = getattr(a, 'name', None)
print(b) --> 如果Person对象中有name类属性，则答应类属性的值，如果没有则打印默认值，即None
'''


#当没有对象属性时，类属性默认为对象属性
#对象属性的优先级高于类属性
'''
class Person(object):
    name = 'justin'  # 类属性
    def __init__(self, name):
        pass
print(Person.name) #类属性用类名来调用
per = Person('mia')
print(per.name) #调用对象属性
per.name = 'mia' #添加对象属性
print(Person.name) #不会修改类属性
'''

#动态添加对象属性,且只对当前实例对象生效，对于类创建的其他实例对象不生效
'''
class Person(object):
    def __init__(self, name):
        pass
per = Person('justin')
per.age = 18
print(per.age)
per2 = Person('mia')
print(per2.age) #per2对象不包含age属性
'''

#删除对象中的name属性，在调用会使用到同名的类属性
'''
class Person(object):
    name = 'justin'
    def __init__(self, name):
        self.name = name
per = Person('mia')
print(per.name)
del per.name
print(per.name)
'''
#注意：以后千万不要讲类属性和对象属性重名，因为对象属性会屏蔽掉类属性，当删除对象属性又能使用类属性。

#【day91】动态给实例添加属性和方法并使用
#创建空类
'''
class Person(object):
    pass
per = Person()
#动态添加属性，这体现了动态语言的特点——灵活
per.name = 'justin'
print(per.name)
'''
#动态添加方法
'''
def say(self):
    print('my name is %s' % self.name)
per.speak = say
per.speak(per)
per.speak() #我们在调用实例方法时，从来没有传参传自己的
'''
# 想要不传参怎么办呢？
'''
from types import MethodType
def say(self):
    print('my name is %s' % self.name)
per.speak = MethodType(say, per) #偏函数：将per（self）即实例化对象，传入say方法。即MethodType(say, per) = say(per)
per.speak()
'''
#思考：如果我们想要限制动态添加实例的属性怎么办？
#比如：只允许给对象添加name, age, height, weight属性
#解决：定义类的时候，定义一个特殊的属性__slots__，可以限制动态添加的属性
'''
class Person(object):
    __slots__ = ('name', 'age')
per = Person()
per.name = 'justin'
per.age = 18
print(per.name, per.age)
per.weight = 80 #未在__slots__属性中定义此属性，因此报错
print(per.weight)
'''

#【day92】@property, @属性名.setter
#这两个装饰器可以你对受限制访问的属性使用点语法访问
'''
class Person(object):
    def __init__(self, age):
        #属性直接对外暴露
        self.age = age
per = Person(18)
print(per.age) #属性直接对外暴露
#缺点：1、不安全 2、数据没有过滤
per.age = -10
print(per.age)
'''
#以上问题可以用过限制访问解决，即self.__age,需要自己写set和get方法，这样才能访问
'''
class Person(object):
    def __init__(self, age):
        self.__age = age

    def setAge(self, age):
        if age < 0:
            age = 0
        else:
            self.__age = age

    def getAge(self):
        return self.__age
per = Person(15)
print(per.getAge())
per.setAge(20)
print(per.getAge())
'''

#使用限制访问，在访问私有属性时比较麻烦，我们还是想要使用实例.变量的方式去赋值私有属性和取值私有属性
'''
class Person(object):
    def __init__(self, age, name):
        self.__age = age
        self.__name = name

    #受限制的变量去掉双下划线
    @property #property装饰器
    def age(self):
        return self.__age
    @age.setter #去掉下划线.setter
    def age(self, age): #方法名相同，下面的函数会吧上面的函数覆盖掉
        if age < 0:
            self.__age = age
        else:
            self.__age = age

    #受限制的变量去掉双下划线
    @property #property装饰器
    def name(self):
        return self.__name
    @name.setter #去掉下划线.setter
    def name(self, name): #方法名相同，下面的函数会吧上面的函数覆盖掉
        self.__name = name

per = Person(15, 'justin') #经过装饰后，又可以用我们熟悉的点来调用属性了
per.age = 100 #相当于是以前使用的setAge()
print(per.age) #相当于是以前使用的getAge()
print(per.name) #相当于是以前使用的getName()
'''

#【day93】运算符重载
'''
print(1 + 2)
print('1' + '2')
'''
#不同的类型用加法会有不同的解释
'''
class Person(object):
    def __init__(self, num):
        self.num = num
per1 = Person(1)
per2 = Person(2)
print(per1 + per2) #两个对象相加，+法解释不了
'''
#使用运算符重载，重新解释两个对象相加
'''
class Person(object):
    def __init__(self, num):
        self.num = num

    #重载加法运算符
    def __add__(self, other): #self代表per1，other代表per2
        return Person(self.num + other.num) #self.num = per1.num, other.num = per2.num , self.num+self.num=3。到这一步就可以打印per1+per2了，但是打印的是对象地址，需要再重置一下打印结果

    def __str__(self):
        return 'num = ' + str(self.num)

per1 = Person(1)
per2 = Person(2)
print(per1 + per2) # per1 + per2 == per1.__add__(per2)
'''

#【day94】发短信以及发邮件
# 接口类型：互亿无线触发短信接口，支持发送验证码短信、订单通知短信等。
# 账户注册：请通过该地址开通账户http://sms.ihuyi.com/register.html
# 注意事项：
# （1）调试期间，请用默认的模板进行测试，默认模板详见接口文档；
# （2）请使用APIID（查看APIID请登录用户中心->验证码短信->产品总览->APIID）及 APIkey来调用接口；
# （3）该代码仅供接入互亿无线短信接口参考使用，客户可根据实际需要自行编写；

# !/usr/local/bin/python
# -*- coding:utf-8 -*-
'''
import http.client
import urllib

host = "106.ihuyi.com"
sms_send_uri = "/webservice/sms.php?method=Submit"
account = "C10780844"
password = "d5cb4462b9db850e1e3dc0427f2dc838"

def send_sms(text, mobile):
    params = urllib.parse.urlencode(
        {'account': account, 'password': password, 'content': text, 'mobile': mobile, 'format': 'json'})
    headers = {"Content-type": "application/x-www-form-urlencoded", "Accept": "text/plain"}
    conn = http.client.HTTPConnection(host, port=80, timeout=30)
    conn.request("POST", sms_send_uri, params, headers)
    response = conn.getresponse()
    response_str = response.read()
    conn.close()
    return response_str

if __name__ == '__main__':
    mobile = "18086829907"
    text = "您的验证码是：121254。请不要把验证码泄露给其他人。"
    #print(send_sms(text, mobile))
'''

#【day95】发邮件
#需要登录163邮箱，在设置中打开SMTP的客户端授权码界面，授权第三方访问邮件
'''
import smtplib #发邮件的库
from email.mime.text import MIMEText #邮件文本
#SMTP服务器
SMTPServer = 'smtp.163.com'
#发邮件的地址
sender = 'chenyilong112233@163.com'
#发送者邮箱的密码(授权码)
passwd = '135cylpsx'

#设置发送的内容
message = '你的账号被脱裤'
#字符串转邮件文本
msg = MIMEText(message)
#标题
msg['Subject'] = '来自你老公的问候'
#发送者
msg['From'] = sender

#创建SMTP服务器，即链接服务器,相当于打开163的登陆界面
mailServer = smtplib.SMTP(SMTPServer, 25) #邮件的端口号都是25，以后设置自己的端口号要超过1024，因为0-1024都是被系统占用的
#登陆邮箱
mailServer.login(sender, passwd)
#发送邮件
mailServer.sendmail(sender, ['chenyilong112233@163.com', '417217170@qq.com'], msg.as_string()) #参数1：发送者;参数2：收件者列表;参数3：一个函数，将列表中的字符串转换成邮件形式的字符串
#退出邮箱
mailServer.quit()
#注意：报错554，往往是内容不合法把邮件判定为垃圾邮件、收件者的邮箱不存在
'''

'''
#作业：1、封装手机类，调用发短信的方法，传参电话号码，以及文本
class MobilePhone(object):
    def __init__(self, number, text):
        self.number = number
        self.text = text

    def sendMassage(self):
        import http.client
        import urllib
        host = "106.ihuyi.com"
        sms_send_uri = "/webservice/sms.php?method=Submit"
        account = "C10780844"
        password = "d5cb4462b9db850e1e3dc0427f2dc838"
        params = urllib.parse.urlencode({'account': account, 'password': password, 'content': self.text, 'mobile': self.number, 'format': 'json'})
        headers = {"Content-type": "application/x-www-form-urlencoded", "Accept": "text/plain"}
        conn = http.client.HTTPConnection(host, port=80, timeout=30)
        conn.request("POST", sms_send_uri, params, headers)
        response = conn.getresponse()
        response_str = response.read()
        conn.close()
        print(response_str)

mobilePhon = MobilePhone('18086829907', '您的验证码是：417217。请不要把验证码泄露给其他人。')
mobilePhon.sendMassage()

#     2、封装一个邮箱的类，调用一个发邮件的方法，传参发件者、收件者、标题、内容
class Email(object):
    def __init__(self, sender, password, text, title, *addresseeList):
        self.sender = sender #chenyilong112233@163.com
        self.password = password #135cylpsx
        self.text = text #邮件内容
        self.title = title #邮箱标题
        self.addresseeList = list(addresseeList) #收件人

    def sendemail(self):
        import smtplib #发邮件的库
        from email.mime.text import MIMEText #邮件文本
        #SMTP服务器
        SMTPServer = 'smtp.163.com'
        #发件者的地址
        sender = self.sender
        #发送者邮箱的密码(授权码)
        passwd = self.password
        #设置发送的内容
        message = self.text
        #字符串转邮件文本
        msg = MIMEText(message)
        #标题
        msg['Subject'] = self.title
        #发送者
        msg['From'] = self.sender
        #创建SMTP服务器，即链接服务器,相当于打开163的登陆界面
        mailServer = smtplib.SMTP(SMTPServer, 25) #邮件的端口号都是25，以后设置自己的端口号要超过1024，因为0-1024都是被系统占用的
        #登陆邮箱
        mailServer.login(sender, passwd)
        #发送邮件
        mailServer.sendmail(self.sender, self.addresseeList, msg.as_string()) #参数1：发送者;参数2：收件者列表;参数3：一个函数，将列表中的字符串转换成邮件形式的字符串
        #退出邮箱
        mailServer.quit()

if __name__ == '__main__':
    email = Email('chenyilong112233@163.com', '135cylpsx', '幸福人生', '爱你的人', '417217170@qq.com', 'chenyilong112233@163.com')
    email.sendemail()
'''

#银行提款机系统
'''
人
类名：Person
属性：姓名 身份证 电话号 卡
行为：开户 查询 取款 存款 转账 改密 锁卡 解锁 补卡 销户

卡
类名：Card
属性：卡号 密码 余额
行为：

银行
类名：Ban
属性：用户列表 提款机
行为：

提款机
类名：
属性：
行为：开户 查询 取款 存款 转账 改密 锁卡 解锁 补卡 销户 退出

界面
类名：View
属性：
行为：管理员界面 管理员登陆 系统功能界面
'''
#见银行系统实战


#【day96】tkinter
#GUI第三方库
'''
import tkinter
#创建主窗口
win = tkinter.Tk()
#设置窗口标题
win.title('我的第一个窗口')
#设置窗口大小和位置
win.geometry('400x400+200+0') #400x400窗口大小 （200，0）初始化位置
win.mainloop()#进入消息循环
'''

#Label:标签控件
#功能：可以显示文本
'''
import tkinter
win = tkinter.Tk()
win.title('我的第一个窗口')
win.geometry('400x400+200+0')

#win:父窗体
#text:显示的文本内容
#bg:背景色
#fg:字体颜色
#font:字体的风格、大小
#width:宽
#height:高
#wraplength:指定多宽之后换行
#justify:设置换行后的对齐方式
#anchor:位置 e东 s南 w西 n北 center中 es东南 en东北 sw西南 nw西北
label = tkinter.Label(win, text='我的名字叫陈艺龙', bg='pink', font=('黑体', 10),
                      width=50, height=10, wraplength=10, justify='left', anchor='center')
label.pack() #显示出来
win.mainloop()
'''

#Button控件
'''
import tkinter
win = tkinter.Tk()
win.title('我的第一个窗口')
win.geometry('400x400+200+0')

#command参数=函数名
def function():
    print('陈艺龙很帅')

button1 = tkinter.Button(win, text='按钮', command=function, width=10, height=2)
button1.pack()

#command参数=匿名函数
button2 = tkinter.Button(win, text='按钮', command=lambda: print('陈艺龙很酷'))
button2.pack()

#command参数=对象方法
button3 = tkinter.Button(win, text='退出', command=win.quit)
button3.pack()

win.mainloop()
'''

#Entry输入控件
#功能：用于显示简单的文本内容
'''
import tkinter
win = tkinter.Tk()
win.title('我的主窗体')
win.geometry('400x400+200+0')
entry1 = tkinter.Entry(win)
entry1.pack()
entry2 = tkinter.Entry(win, show='*') #show：匿密显示，用其中的字符串显示,'%'即用%号显示
entry2.pack()
#绑定变量
e = tkinter.Variable()
entry3 = tkinter.Entry(win, textvariable=e)
entry3.pack()
#e就代表这个输入框对象
#设置值
e.set('陈艺龙') #默认文本内容
#取值
print(e.get())
print(entry3.get())
win.mainloop()
'''

'''
#作业点击按钮，输出输入框中的内容
#方法一：
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
e = tkinter.Variable()
entry = tkinter.Entry(win, textvariable=e)
entry.pack()
button = tkinter.Button(win, text='打印', width=10, height=2, command=lambda:print(entry.get()))
button.pack()
win.mainloop()

#方法二：
import tkinter
win = tkinter.Tk()
win.title('主窗体')
win.geometry('800x800+200+200')
entry = tkinter.Entry(win)
entry.pack()
def showInfo():
    print(entry.get())
button = tkinter.Button(win, text='打印', command=showInfo)
button.pack()
win.mainloop()
'''

#Text文本控件
#功能：在界面窗口显示多行文本
'''
import tkinter
win = tkinter.Tk()
win.title('主窗体')
win.geometry('800x800+200+200')
text = tkinter.Text(win, width=50, height=8) #注：height显示的行数
text.pack()
#str_ = If there is anyone out there who still doubts that America is a place where all things are possible; who still wonders if the dream of our founders is alive in our time; who still questions the power of our democracy, tonight is your answer
text.insert(tkinter.INSERT, str_) #在文本中插入
#text.insert(tkinter.END, str_) #在文本最后插入

win.mainloop()
'''

#Text带滚动条的文本控件
'''
import tkinter
win = tkinter.Tk()
win.title('主窗体')
#win.geometry('800x800+200+200')
#创建滚动条和文本
scroll = tkinter.Scrollbar()
text = tkinter.Text(win, width=50, height=4) #注：height显示的行数
#显示滚动条和文本
scroll.pack(side=tkinter.RIGHT, fill=tkinter.Y) #side:放到窗体的哪一侧, fill：填充满在y轴方向
text.pack(side=tkinter.LEFT, fill=tkinter.Y)
#关联滚动条和文本
scroll.config(command=text.yview) #滚动条动，控制文本动
text.config(yscrollcommand=scroll.set) #文本动，控制滚动条动

#str_ = If there is anyone out there who still doubts that America is a place where all things are possible; who still wonders if the dream of our founders is alive in our time; who still questions the power of our democracy, tonight is your answer
text.insert(tkinter.INSERT, str_)
win.mainloop()
'''

#CheckButton多选框控件
'''
import tkinter
win = tkinter.Tk()
win.title('主窗体')
win.geometry('800x800+200+200')

def updata():
    message = ''
    if hobby1.get() == True:
        message += 'money\n'
    if hobby2.get() == True:
        message += 'power\n'
    if hobby3.get() == True:
        message += 'people\n'
    text.delete(0.0, tkinter.END) #参数1：从开头清除，0行0列；参数2,：清空到最后
    text.insert(tkinter.INSERT, message)

hobby1 = tkinter.BooleanVar() #生成一个布尔类型的变量
check1 = tkinter.Checkbutton(win, text='money', variable=hobby1, command=updata)
check1.pack()

hobby2 = tkinter.BooleanVar() #生成一个布尔类型的变量
check2 = tkinter.Checkbutton(win, text='power', variable=hobby2, command=updata)
check2.pack()

hobby3 = tkinter.BooleanVar() #生成一个布尔类型的变量
check3 = tkinter.Checkbutton(win, text='people', variable=hobby3, command=updata)
check3.pack()

text = tkinter.Text(win, width=10, height=5)
text.pack()

win.mainloop()
'''

#Radiobutton单选框控件
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')

def updata():
    print(variable.get())

#注意：一组单选框要绑定同一个变量
variable = tkinter.IntVar()

radio1 = tkinter.Radiobutton(win, text='one', value=1, variable=variable, command=updata) #text:单选框显示的内容,value:单选框所代表的数据
radio1.pack()

radio2 = tkinter.Radiobutton(win, text='two', value=2, variable=variable, command=updata)
radio2.pack()

win.mainloop()
'''


#Listbox列表框控件1
#功能：可以包含一个或多个文本框，可单选可多选
#作用：可以在listbox控件的小窗口显示字符串
'''
import tkinter
win = tkinter.Tk()
win.title('主菜单')
win.geometry('800x800+200+200')

lb = tkinter.Listbox(win, selectmode=tkinter.BROWSE) #BROWSE,可以通过鼠标来移动Listbox中的选中位置（不是移动item）
lb.pack()
for item in ['justin', 'mia', 'dixie', 'jessica', 'psx', 'cxm']:
    #按顺序添加，即每个元素添加到tkinter的尾部，如果参数1为：tkinter.ACTIVE,则往前面添加
    lb.insert(tkinter.END, item)
#在头部添加
lb.insert(tkinter.ACTIVE, 'cyl')
#将列表当成一个元素添加
lb.insert(tkinter.END, ['czy','cdx'])

#删除 参数1：开始的索引，参数2：结束的索引，如果不指定参数2，只删除第一个索引处的内容
lb.delete(0, lb.size()) #删除多个值,列表清屏
#lb.delete(7) #删除单个值

#选中
#lb.select_set(2, 3) #选择start，end范围内的内容
#lb.select_set(0) #选中索引处的内容

#取消选中
#lb.select_clear(2, 3)
#lb.select_clear(0)

#获取到列表中的元素个数
#print(lb.size())

#从列表取值
#print(lb.get(2, 3))
#print(lb.get(1))

#返回当前选中项的索引，不是item元素
#print(lb.curselection())

#判断一个选项是否被选中
#print(lb.selection_includes(0))
#print(lb.selection_includes(1))

win.mainloop()
'''

#Listbox列表框控件2
'''
import tkinter
win = tkinter.Tk()
win.title('123')
win.geometry('800x800+200+200')
lbv = tkinter.StringVar()
#single 与 borwse 相似，但它不支持鼠标按下后移动选中位置
lb = tkinter.Listbox(win, selectmode=tkinter.SINGLE, listvariable=lbv) #与BROWSE相似 的为SINGLE，但不支持鼠标移动选中位置。
for i in ['justin', 'mia', 'dixie', 'jessica', 'psx', 'cxm']:
    lb.insert(tkinter.ACTIVE, i)
lb.pack()
#重置列表中的值
#lbv.set((1, 2))
#绑定事件
def myPrint(event):
    print(lb.get(lb.curselection()))
lb.bind('<Double-Button-1>', myPrint) #双击-按钮-左键
win.mainloop()
'''

#Listbox列表控件3
'''
import tkinter
win = tkinter.Tk()
win.title('123')
#win.geometry('800x800+200+200')
#EXTENDED,可以使listbox支持shift和control
lb = tkinter.Listbox(win, selectmode=tkinter.EXTENDED) #EXPANDED使Listbox支持Shift和Control
for i in ['justin', 'mia', 'dixie', 'jessica', 'psx', 'cxm',
          'justin1', 'mia1', 'dixie1', 'jessica1', 'psx1', 'cxm1',
          'justin2', 'mia2', 'dixie2', 'jessica2', 'psx2', 'cxm2']:
    lb.insert(tkinter.END, i)
#添加滚动条
sc = tkinter.Scrollbar(win)
sc.pack(side=tkinter.RIGHT, fill=tkinter.Y)
lb.pack(side=tkinter.LEFT, fill=tkinter.BOTH)
lb.configure(yscrollcommand=sc.set)
sc['command'] = lb.yview #给属性赋值

win.mainloop()
'''

#Listbox列表控件4
'''
import tkinter
win = tkinter.Tk()
win.title('123')
#win.geometry('800x800+200+200')
lb = tkinter.Listbox(win, selectmode=tkinter.MULTIPLE) #MULTIPLE,支持多选
for i in ['justin', 'mia', 'dixie', 'jessica', 'psx', 'cxm',
          'justin1', 'mia1', 'dixie1', 'jessica1', 'psx1', 'cxm1',
          'justin2', 'mia2', 'dixie2', 'jessica2', 'psx2', 'cxm2']:
    lb.insert(tkinter.END, i)
#添加滚动条
sc = tkinter.Scrollbar(win)
sc.pack(side=tkinter.RIGHT, fill=tkinter.Y)
lb.pack(side=tkinter.LEFT, fill=tkinter.BOTH)
lb.configure(yscrollcommand=sc.set)
sc['command'] = lb.yview #给属性赋值
win.mainloop()
'''

#Scale范围控件
#功能：供用户通过拖拽指示器改变变量的值，可以水平，可以竖直
'''
import tkinter
win = tkinter.Tk()
win.title('123')
win.geometry('800x800+200+200')
#HORIZONTAL水平
#VERTICAL垂直
#length 水平宽度 竖直高度
#tickinterval 选择值为该值的倍数

scale = tkinter.Scale(win, from_=0, to=100, orient=tkinter.HORIZONTAL, tickinterval=10, length=200)
scale.pack()
#设置值初始值
scale.set(20)
#取值
def showNum():
    print(scale.get())
tkinter.Button(win, text='按钮', command=showNum).pack()
win.mainloop()
'''


#Spinbox数值范围控件
#功能：控制数据增加减少
'''
import tkinter
win = tkinter.Tk()
win.title('123')
win.geometry('800x800+200+200')
#from_ 最小值
#to 最大值
#increment 步长 默认为1
#values 指定变化的数值 最好不要与from_to和incrememt同时使用
#command 只要数值变化就能执行对应的方法
#绑定变量
v = tkinter.StringVar()
#绑定事件
def showNum():
    print(v.get())
sp1 = tkinter.Spinbox(win, from_=0, to=100, increment=5, textvariable=v, command=showNum)
sp1.pack()
#赋值
v.set(20)
#取值
print(v.get())
sp2 = tkinter.Spinbox(win, values=(0, 2, 4, 6, 8))
sp2.pack()
win.mainloop()
'''

#Menu顶层菜单
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
def function():
    print('justin is good!')
#菜单条
menubar = tkinter.Menu(win)
#配置菜单栏
win.config(menu=menubar)
#创建菜单选项
menu1 = tkinter.Menu(menubar, tearoff=False)
#给菜单选项添加内容
for item in ['Python', 'C', 'C++', 'OC', 'Swift', 'C#', 'shell', 'Java', 'JS', 'PHP', '汇编', 'NodeJS', 'quite']:
    if item == 'quite':
        menu1.add_separator() #给菜单栏添加分割线
        menu1.add_command(label=item, command=win.quit) #退出
    else:
        menu1.add_command(label=item, command=function) #不用循环，手动添加即可实现调用不同的方法
#向菜单条上添加菜单
menubar.add_cascade(label='语言', menu=menu1)

menu2 = tkinter.Menu(menubar, tearoff=True) #tearoff撕裂，可以独立显示的菜单栏
menu2.add_command(label='red')
menu2.add_command(label='blue')
menubar.add_cascade(label='color', menu=menu2)

win.mainloop()

'''

#Menu鼠标右键菜单
'''
import tkinter
win = tkinter.Tk()
win.title('123')
win.geometry('800x800+200+200')

#菜单条
menubar = tkinter.Menu(win)

#菜单
menu = tkinter.Menu(menubar, tearoff=False)
menu.add_command(label='one')
menu.add_command(label='two')

#菜单条绑定菜单
menubar.add_cascade(label='Number', menu=menu)


#绑定事件
def showMenu(event):
    menubar.post(event.x_root, event.y_root) #post功能显示menubar，event.x_root和event.y_root鼠标点击位置的坐标
win.bind('<Button-3>', showMenu)

win.mainloop()
'''

#Combobox下拉控件
'''
import tkinter
from tkinter import ttk
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
#绑定变量
cv = tkinter.StringVar()
com = ttk.Combobox(win, textvariable=cv)
com.pack()
#设置下拉属性
com['value'] = ('四川', '重庆', '北京', '哈尔滨')
#设置下拉框默认显示的value属性
#com.current(0)
#绑定事件
def function(event):
    print(com.get()) #选其一来取值
    print(cv.get())
com.bind('<<ComboboxSelected>>', function)
win.mainloop()
'''

#Frame框架控件
#功能：在屏幕上显示一个矩形区域，作为容器控件
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')

frm = tkinter.Frame(win)
frm.pack()

#left
frm_1 = tkinter.Frame(frm)
frm_1.pack(side=tkinter.LEFT)
tkinter.Label(frm_1, text='左上', bg='red').pack(side=tkinter.TOP)
tkinter.Label(frm_1, text='左下', bg='pink').pack(side=tkinter.BOTTOM)


#right
frm_2 = tkinter.Frame(frm)
tkinter.Label(frm_2, text='右上', bg='blue').pack(side=tkinter.TOP)
tkinter.Label(frm_2, text='右下', bg='yellow').pack(side=tkinter.TOP)
frm_2.pack(side=tkinter.RIGHT)

win.mainloop()
'''

#显示数据：表格数据
'''
import tkinter
from tkinter import ttk
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
#创建表格
tree = ttk.Treeview(win)
#定义列
tree['columns'] = ('name', 'age', 'height', 'weight')
#设置列
tree.column('name', width=100)
tree.column('age', width=100)
tree.column('height', width=100)
tree.column('weight', width=100)
#设置表头
tree.heading('name', text='姓名')
tree.heading('age', text='年龄')
tree.heading('height', text='身高')
tree.heading('weight', text='体重')
#添加数据
tree.insert('', 0, text='line1', values=('mia', '9', '140', '50'))
tree.insert('', 1, text='line2', values=('justin', '30', '170', '80')) #参数2：下标
tree.pack()
win.mainloop()
'''

#显示数据：树状数据
'''
import tkinter
from tkinter import ttk
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
#创建表格
tree = ttk.Treeview(win)
tree.pack()
#添加一级树枝
tree1 = tree.insert('', 0, '中国', text='China', values=('1'), open=True) #参数1：父级窗体，参数2：插入位置，参数4：显示内容,参数5：值，参数6：是否展开
tree2 = tree.insert('', 1, '美国', text='America', values=('2'))
tree3= tree.insert('', 2, '日本', text='Japen', values=('3'))
#添加二级树枝
tree1_1 = tree.insert(tree1, 0, '黑龙江', text='HeiLongJiang', values='1_1')
tree1_2 = tree.insert(tree1, 1, '四川', text='SiChuan', values='1_2')
tree1_3 = tree.insert(tree1, 2, '北京', text='BeiJing', values='1_3')
tree2_1 = tree.insert(tree2, 0, '纽约', text='NewYork', values='2_1')
tree2_2 = tree.insert(tree2, 1, '芝加哥', text='Chicago', values='2_2')
tree2_3 = tree.insert(tree2, 2, '波士顿', text='Boston', values='2_3')
tree3_1 = tree.insert(tree3, 0, '札幌市', text='札幌市', values='3_1')
tree3_2 = tree.insert(tree3, 1, '函馆市', text='函馆市', values='3_2')
tree3_3 = tree.insert(tree3, 2, '小樽市', text='小樽市', values='3_3')
#添加三级树枝
tree1_1_1 = tree.insert(tree1_1, 0, '哈尔滨市', text='哈尔滨市', values='1_1_1')
tree1_1_2 = tree.insert(tree1_1, 0, '齐齐哈尔市', text='齐齐哈尔市', values='1_1_2')
tree1_1_3 = tree.insert(tree1_1, 0, '牡丹江市', text='牡丹江市', values='1_1_3')
win.mainloop()
'''
'''
#作业：用到递归遍历取出传入path的所有文件夹名及文件名，并且将他们用树状结构显示出来
#创建一个树状结构类，继承tkinter.Frame类，实例化对象 = 该类（传入父窗体和一个文件路径）
#树状窗体
#类名 TreeWindows
#属性 self.tree = ttk.Treeview(父窗体)对象
#方法 loadTree插入树
'''

#绝对布局
'''
import tkinter
from tkinter import ttk
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label1 = tkinter.Label(win, text='1', bg='blue')
label2 = tkinter.Label(win, text='2', bg='red')
label3 = tkinter.Label(win, text='3', bg='yellow')
#绝对布局:窗口的变化对位置没有影响
label1.place(x=10, y=10)
label2.place(x=50, y=50)
label3.place(x=100, y=100)
win.mainloop()
'''

#相对布局
'''
import tkinter
from tkinter import ttk
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label1 = tkinter.Label(win, text='justin', bg='blue')
label2 = tkinter.Label(win, text='mia', bg='red')
label3 = tkinter.Label(win, text='dixie', bg='yellow')
label4 = tkinter.Label(win, text='coco', bg='green')
#相对布局:窗口的变化对位置有影响
label1.pack(fill=tkinter.Y, side=tkinter.LEFT)
label2.pack(fill=tkinter.X, side=tkinter.TOP)
label3.pack(fill=tkinter.Y, side=tkinter.RIGHT)
label4.pack(fill=tkinter.X, side=tkinter.BOTTOM)
win.mainloop()
'''

#表格布局
'''
import tkinter
from tkinter import ttk
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label1 = tkinter.Label(win, text='justin', bg='blue')
label2 = tkinter.Label(win, text='mia', bg='red')
label3 = tkinter.Label(win, text='dixie', bg='yellow')
label4 = tkinter.Label(win, text='shuxian', bg='yellow')
#表格布局
label1.grid(row=0, column=0)
label2.grid(row=0, column=1)
label3.grid(row=1, column=0)
label4.grid(row=1, column=1)
win.mainloop()
'''

#鼠标事件
#<Button-i> 鼠标单击事件 Button-1,Button-2,Button-3表明左键，中间，右键，点击时tkinter会自动抓取鼠标位置
#<Double-Button-i> 鼠标双击事件
#<Triple-Button-i> 鼠标三击事件
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
#event对象，属性：char、keycode、num、widget、x和y、x_root和y_root
def function(event):
    print(event.x, event.y) #控件的x轴y轴位置
    #print(event.x_root, event.y_root) #主屏幕的x轴y轴位置
    print(event.widget) #触发控件的对象。widget.selection
button = tkinter.Button(win, text='按钮')
button.bind('<Button-1>', function) #bind 参数1：绑定事件，参数2：触发函数,注bind可以给所有控件绑定事件
button.pack()
label = tkinter.Label(win, text='justin')
label.bind('<Button-2>', function)
label.pack()
win.mainloop()
'''

#鼠标移动事件
#<Bi-Motion> 当鼠标左(中、右)键被按住在小构建且移动鼠标是事件触发
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label = tkinter.Label(win, text='justin')
label.pack()
def function(event):
    print(event.x, event.y)
label.bind('<B1-Motion>', function)
win.mainloop()
'''


#鼠标释放事件
#<ButtonRelease-i> 当释放鼠标左(中、右)键时事件发生
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label = tkinter.Label(win, text='justin', bg='red')
label.pack()
def function(event):
    print(event.x, event.y)
label.bind('<ButtonRelease-1>', function)
win.mainloop()
'''

#进入事件
#<Enter> 当鼠标光标进入控件时事件触发
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label = tkinter.Label(win, text='justin', bg='red')
label.pack()
def function(event):
    print(event.x, event.y)
label.bind('<Enter>', function)
win.mainloop()
'''

#离开事件
#<Leave> 当鼠标光标离开控件时事件触发
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label = tkinter.Label(win, text='justin', bg='red')
label.pack()
def function(event):
    print(event.x, event.y)
label.bind('<Leave>', function)
win.mainloop()
'''

#响应所有键盘按键的事件
#<Key> 当单击一个键时事件发生
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label = tkinter.Label(win, text='justin', bg='red')
label.focus_set() #设置焦点，使用key事件必须设置焦点，可以理解为键盘的键都作用在控件的焦点上
label.pack()
def function(event):
    print(event.char) #event.char 代表从键盘输入的和按键事件相关的字符
    print(event.keycode) #event.keycode 代表从键盘输入的和按键事件相关的键的ASKII码
label.bind('<Key>', function)
#只有给win绑定Key事件时，不需要给win设置焦点
#def function(event):
#    print(event.char)
#    print(event.keycode)
#win.bind('<Key>', function)
win.mainloop()
'''

#响应特殊按键事件
#<Shift_L> <Shift_R>
#<F1> <F2>..
#<Return> 回车键 当单击'Enter'键时事件触发
#<BackSpace> 退格键
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label = tkinter.Label(win, text='justin', bg='red')
label.focus_set() #设置焦点，使用key事件必须设置焦点，可以理解为键盘的键都作用在控件的焦点上
label.pack()
def function(event):
    print(event.char) #event.char 代表从键盘输入的和按键事件相关的字符
    print(event.keycode) #event.keycode 代表从键盘输入的和按键事件相关的键的ASKII码
label.bind('<Return>', function)
win.mainloop()
'''

#指定按键事件
#'a' 点击a键时触发
'''
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label = tkinter.Label(win, text='justin', bg='red')
label.focus_set() #设置焦点，使用key事件必须设置焦点，可以理解为键盘的键都作用在控件的焦点上
label.pack()
def function(event):
    print(event.char) #event.char 代表从键盘输入的和按键事件相关的字符
    print(event.keycode) #event.keycode 代表从键盘输入的和按键事件相关的键的ASKII码
label.bind('a', function)
win.mainloop()
'''

#组合按键事件
'''
#<Control-Alt-n>
import tkinter
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
label = tkinter.Label(win, text='justin', bg='red')
label.focus_set() #设置焦点，使用key事件必须设置焦点，可以理解为键盘的键都作用在控件的焦点上
label.pack()
def function(event):
    print(event.char) #event.char 代表从键盘输入的和按键事件相关的字符
    print(event.keycode) #event.keycode 代表从键盘输入的和按键事件相关的键的ASKII码
label.bind('<Shift-Up>', function)
win.mainloop()
'''

#关闭窗体事件
'''
import tkinter
import tkinter.messagebox
root=tkinter.Tk()
root.geometry("200x200")
def func1():
    if tkinter.messagebox.askyesno("关闭窗口", "确认关闭窗口吗"):
        root.destroy()
root.protocol("WM_DELETE_WINDOW", func1)
root.mainloop()
'''

#树结构选择事件，取树结构的text和values
'''
import tkinter
from tkinter import ttk
win = tkinter.Tk()
win.title('主窗口')
win.geometry('800x800+200+200')
#创建表格
tree = ttk.Treeview(win)
tree.pack()
#添加一级树枝
tree1 = tree.insert('', 0, '中国', text='China', values=('1'), open=True) #参数1：父级窗体，参数2：插入位置，参数4：显示内容,参数5：值，参数6：是否展开
#绑定事件
def treeViewSelect(event):
    v = event.widget.selection()
    for sv in v:
        file = tree.item(sv)['text']
        print('点击树结构后，取到树结构text为：', file)
        value = tree.item(sv)['values'][0]
        print('点击树结构后，取到树结构value为：', value)
tree.bind('<<TreeviewSelect>>', treeViewSelect)
win.mainloop()
'''

#【day97】读csv文件
'''
import csv
def readCsv(path):
    infoList=[]
    with open(path, 'r', encoding='utf-8') as f:
        allFileInfo = csv.reader(f)
        for row in allFileInfo:
            infoList.append(row)
    return infoList
path = r'D:\迅雷下载\数据\2000W开房数据\最后5000.csv'
info = readCsv(path)
'''

#【day98】写csv文件
'''
import csv
def writeCsv(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        writer = csv.writer(f)
        for rowData in data:
            writer.writerow(rowData)

path = r'D:\qian_feng_education\first_project\test.csv'
writeCsv(path, [['name','age','gender'],['陈艺龙','30','男']])
'''

#【day99】读取pdf中的文本内容,将内容写到txt
'''
import sys
import importlib
importlib.reload(sys)
import os

from pdfminer.pdfparser import PDFParser, PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LTTextBoxHorizontal, LAParams
from pdfminer.pdfinterp import PDFTextExtractionNotAllowed

def readPDF(path, callback=None, toPath=''):
    # 以二进制形式打开pdf文件
    with open(path, 'rb') as f:
        # 创建一个pdf文档分析器
        parser = PDFParser(f)
        # 创建一个pdf文档
        pdfFile = PDFDocument()
        # 将新建pdf与分析器进行链接
        parser.set_document(pdfFile)
        pdfFile.set_parser(parser)
        # 提供初始化密码
        pdfFile.initialize()
        # 检测文档是否提供Text转换
        if not pdfFile.is_extractable:  # 不能提供转换功能返回False，not之后为True，不在看此pdf
            raise PDFTextExtractionNotAllowed  # 结束
        else:
            # 解析数据
            # 创建数据管理器
            manager = PDFResourceManager()
            # 创建PDF设备对象
            laparams = LAParams()
            device = PDFPageAggregator(manager, laparams=laparams)
            # 创建解释器对象
            interpreter = PDFPageInterpreter(manager, device)
            # 开始循环处理，每次处理一页
            for page in pdfFile.get_pages():  # 获取pdf的页数
                # 解释器处理一页内容
                interpreter.process_page(page)
                # 创建图层对象
                layout = device.get_result()
                for x in layout:
                    # isinstance判断x是否为LTTextBoxHorizontal类型
                    if (isinstance(x, LTTextBoxHorizontal)):
                        if toPath == '':
                        #处理数据,因为每次处理数据的方式不一样，又不想修改类中的代码，因此传入一个callback函数，在main中自己写，再传入类中
                            str_ = x.get_text()
                            if callback != None:
                                callback(str_)
                            else:
                                print(str_)
                        else:
                            print('将输入存入文件')
                            with open(toPath, 'a', encoding='utf-8') as f:
                                # 读pdf的内容
                                str_ = x.get_text()
                                f.write(str_ + '\n')

#传toPath
path = r'D:\qian_feng_education\first_project\500d-cvresume.pdf'
toPath = r'D:\qian_feng_education\first_project\500d-cvresume.txt'
readPDF(path, toPath)

#不穿toPath
path = os.path.join(os.getcwd(), '500d-cvresume.pdf')
rwFile = RwFile()
def fuction(str_):
    print(str_ + '!')
rwFile.readPDF(path, fuction)
'''

#【day100】word自动化办公
#读取doc与docx文件并返回列表
'''
import win32com
import win32com.client
def readWrod(path):
    word = []
    #实例化系统中word的功能，即打开word软件，可以处理doc和docx两种文件
    mw = win32com.client.Dispatch('Word.Application')
    #调用实例对象的Open方法打开指定word文件
    doc = mw.Documents.Open(path)
    for paragraph in doc.Paragraphs:
        line = paragraph.Range.Text
        word.append(line)
        print(line)
    #关闭文件
    doc.Close()
    #退出word
    mw.Quit()
path = r'D:\qian_feng_education\first_project\test.docx'
readWrod(path)
'''

#读取doc和docx并另存为其他格式的文件
'''
import win32com
import win32com.client
def readWrodToOther(path, toPath):
    mw = win32com.client.Dispatch('Word.Application')
    doc = mw.Documents.Open(path)
    #将读取到的数据保存到(另一个文件中, txt文件)
    doc.SaveAs(toPath, 2)
    doc.Close()
    mw.Quit()
path = r'D:\qian_feng_education\first_project\test.docx'
toPath = r'D:\qian_feng_education\first_project\test.txt'
readWrodToOther(path, toPath)
'''

#写word
'''
import win32com
import win32com.client
import os

def makeWorkFile(path, name):
    word = win32com.client.Dispatch('Word.Application')#打开word软件
    #让文档可见
    word.Visible = True
    #创建文档
    doc = word.Documents.Add()
    #写内容
    #从头开始写
    r = doc.Range(0, 0) #光标定位到最开始
    r.InsertAfter('亲爱的%s\n' % name)
    r.InsertAfter('    我想你。。。\n')
    #存储文件
    doc.SaveAs(path)
    #关闭文件
    doc.Close()
    #退出软件
    word.Quit()

names = ['justin', 'hailey', 'mia', 'jessie']
for name in names:
    path = os.path.join(os.getcwd(), name) + '.docx'
    print(path)
    makeWorkFile(path, name)
'''


#【day101】excel自动化办公

#读xlsx并返回dict
'''
from openpyxl.reader.excel import load_workbook
def readXlsxFile(path):
    #打开文件
    file = load_workbook(filename=path)
    #所有sheet的名称
    sheets = file.get_sheet_names()

    #拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    print(sheet.max_row) #最大行数
    print(sheet.max_column) #最大列数
    print(sheet.title) #sheet名

    #读取一张表的数据
    allList = []
    for lineNum in range(1, sheet.max_row + 1):
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            #拿单元格数据
            value = sheet.cell(row=lineNum, column=columnNum).value
            if value != None: #看数据不需要None, 但如果写数据则需要None,因为单元格才对齐
                lineList.append(value)
        allList.append(lineList)

path = r''
readXlsxFile(path)
'''

#读xls和xlsx文件
'''
from collections import OrderedDict #导入有序字典
from pyexcel_xls import get_data
def readXlsAndXlsxFile(path):
    dict_ = OrderedDict() #有序字典
    xdata = get_data(path)
    for sheet in xdata:
        dict_[sheet] = xdata[sheet]
    return dict_
path = r''
dict_ = readXlsAndXlsxFile(path)
'''

#写xls文件
'''
pip install openpyxl
pip install xlrd
pip install future
pip install xlwt-future
pip install pyexcel-io
pip install ordereddict
pip install pyexcel
pip install pyexcel-xls

import xlwt
valueErorr
解决：修改xlwt源码
修改前
if flags & SRE_FLAG_LOCALE:
    raise ValueError("cannot use LOCALE flag with a str pattern")
修改后
if flags & SRE_FLAG_LOCALE:
    pass #stone20161228 raise ValueError("cannot use LOCALE flag with a str pattern")
'''
'''
import os
from collections import OrderedDict #导入有序字典
from pyexcel_xls import save_data
def makeExcelFile(path, data):
    dic = OrderedDict()
    for sheetName, sheetValue in data.items():
        dict_ = {}
        dict_[sheetName] = sheetValue
        dic.update(dict_) #.update() 似 .append()
    save_data(path, dic)
    return dic

data = {'sheet1':[['1','2','3'],['4','5','6'],['7','8','9']], 'sheet2':[['a','b','c'],['d','e','f'],['g','h','i']]}
path = os.getcwd() + '.xls'
makeExcelFile(path, data)
'''

'''
#作业：变量读取csv，追加到一个csv中去
import os
from myClass.rwfile import RwFile
toPath = r'D:\qian_feng_education\first_project\tesxData\csvInfo.csv'
rwFile = RwFile()
for num in range(1, 2):
    path = r'D:\qian_feng_education\first_project\tesxData\{}.csv'.format(num)
    csvInfo = rwFile.readCsv(path)
    print(csvInfo)
    rwFile.writeCsv(toPath, csvInfo)
'''


#【day102】ppt自动化办公
'''
import win32com
import win32com.client
def makePPT(path):
    ppt = win32com.client.Dispatch('PowerPoint.Application')
    ppt.Visble = True
    #增加一个文件
    pptFile = ppt.Presentations.Add()
    #新建幻灯片
    pag1 = pptFile.Slides.Add(1, 1) #参数1：幻灯片的页数，参数2：新建幻灯片的类型（默认样式）
    #幻灯片的第一个文本输入框的输入文字的位置
    t1 = pag1.Shapes[0].TextFrame.TextRange
    #写文字
    t1.Text = '陈艺龙'
    t2 = pag1.Shapes[1].TextFrame.TextRange
    # 写文字
    t2.Text = '彭淑贤'
    #保存
    pptFile.SaveAs(path)
    #关闭幻灯片
    pptFile.Close()
    #关闭软件
    ppt.Quit()

path = r''
makePPT(path)
'''

#【day103】播放音乐
'''
import time
import pygame
#音乐路径
filePath = r'D:\qian_feng_education\first_project\Alan Silvestri - The Avengers.mp3'
#初始化
pygame.mixer.init()
#加载音乐
track = pygame.mixer.music.load(filePath)
#播放音乐
pygame.mixer.music.play()
#程序等待音乐播放
time.sleep(10)
#暂停
pygame.mixer.music.pause()
time.sleep(5)
pygame.mixer.music.play()
time.sleep(10)
#停止音乐
pygame.mixer.music.stop()
'''
'''
#作业：用tkinter实现播放器
'''

#【day104】修改背景图片
#注册表 regedit -> HKEY_CURRENT_USER\Control Panel\Desktop\WallPaper
'''
import win32api #系统api
import win32con
import win32gui
def setWillPaper(path):
    #打开注册表
    reg_key = win32api.RegOpenKeyEx(win32con.HKEY_CURRENT_USER, 'Control Panel\\Desktop', 0, win32con.KEY_SET_VALUE) #把注册表desktop打开，为了设置数据
    #系统内置接口.注册表设置值（修改内容，参数1：打开的注册表，参数2：当前注册表中的表名，参数3，当前表的值，参数4：当前表的类型，参数5：此类型的值，即背景图片的契合度，填充10、适应6、拉伸2、平铺、居中0、跨区）
    win32api.RegSetValueEx(reg_key, 'WallpaperStyle', 0, win32con.REG_SZ, '10')
    #
    #win32api.RegSetValueEx(reg_key)
    #win32api.RegSetValueEx(reg_key, 'WallPaper')
    #系统内置api.系统参数信息（修改壁纸，图片路径，立即生效）
    win32gui.SystemParametersInfo(win32con.SPI_SETDESKWALLPAPER, path, win32con.SPIF_SENDWININICHANGE)

path = r'D:\qian_feng_education\first_project\timg.jpg'
setWillPaper(path)
'''


#【day105】整蛊小程序
'''
import time
import pygame
import win32api
import win32con
import win32gui
import os
import threading #线程模块
def musicGo():
    musicPath = r'D:\qian_feng_education\first_project\music'
    pygame.mixer.init()
    while True:
        for fileName in os.listdir(musicPath):
            absFile = os.path.join(musicPath, fileName)
            track = pygame.mixer.music.load(absFile)
            pygame.mixer.music.play()
            time.sleep(20)
            pygame.mixer.music.stop()
def setWillPaper():
    imgPath = r'D:\qian_feng_education\first_project\img'
    while True:
        for imgName in os.listdir(imgPath):
            img = os.path.join(imgPath, imgName)
            reg_key = win32api.RegOpenKeyEx(win32con.HKEY_CURRENT_USER, 'Control Panel\\Desktop', 0, win32con.KEY_SET_VALUE) #把注册表desktop打开，为了设置数据
            win32api.RegSetValueEx(reg_key, 'WallpaperStyle', 0, win32con.REG_SZ, '10')
            win32gui.SystemParametersInfo(win32con.SPI_SETDESKWALLPAPER, img, win32con.SPIF_SENDWININICHANGE)
            time.sleep(10)
#增加一个线程来进行musicGo()
th = threading.Thread(target=musicGo, name='LoopThread') #target只给函数名，不要调用，否则无法执行多线程
th.start()
setWillPaper()
'''

#【day106】键盘模拟
'''
import win32con
import win32api
import time
#系统内置api.点击键盘触发事件（win键（ASKII）, 0, 点击，0）
win32api.keybd_event(91, 0, 0, 0) #点击
time.sleep(0.1)
#系统内置api.点击键盘触发事件（win键, 0, 抬起, 0）
win32api.keybd_event(91, 0, win32con.KEYEVENTF_KEYUP, 0)
'''

#模拟返回桌面
'''
while True:
    win32api.keybd_event(91, 0, 0, 0)  # 点击
    time.sleep(0.1)
    # 系统内置api.点击键盘触发事件（D键, 0, 抬起, 0）
    win32api.keybd_event(77, 0, 0, 0)  # 点击
    time.sleep(5)
'''

#语音控制游戏
#游戏地址：http://www.4399.com/flash/122374.htm#sim2|122374
'''
import win32com.client
from win32com.client import constants
import win32com.client
import pythoncom
import win32con
import win32api
import time
import os

class SpeechRecognition:
    def __init__(self, wordsToAdd):
        self.speaker = win32com.client.Dispatch('SAPI.SpVoice')
        self.listener = win32com.client.Dispatch('SAPI.SpSharedRecognizer')
        self.context = self.listener.CreateRecoContext()
        self.grammar = self.context.CreateGrammar()
        self.grammar.DictationSetState(0)
        self.wordsRule = self.grammar.Rules.Add('wordsRule', constants.SRATopLevel + constants.SRADynamic, 0)
        self.wordsRule.Clear()
        [self.wordsRule.InitialState.AddWordTransition(None, word) for word in wordsToAdd]
        self.grammar.Rules.Commit()
        self.grammar.CmdSetRuleState('wordsRule', 1)
        self.grammar.Rules.Commit()
        self.eventHandler = ContextEvents(self.context)
        self.say('Started successfully')

    def say(self, phrase):
        self.speaker.Speak(phrase)

class ContextEvents(win32com.client.getevents('SAPI.SpSharedRecoContext')):
    def OnRecognition(self, StreamNumber, StreamPosition, RecognitionType, Result):
        newResult = win32com.client.Dispatch(Result) #生成语音识别对象
        voiceCommand = newResult.PhraseInfo.GetText()  # 生成语音转文字变量
        if voiceCommand == '上':
            print('上')
            win32api.keybd_event(38, 0, 0, 0)  # 点击
            time.sleep(0.1)
            win32api.keybd_event(38, 0, win32con.KEYEVENTF_KEYUP, 0)

        elif voiceCommand == '下':
            print('下')
            win32api.keybd_event(40, 0, 0, 0)  # 点击
            time.sleep(0.1)
            win32api.keybd_event(40, 0, win32con.KEYEVENTF_KEYUP, 0)

        elif voiceCommand == '左':
            print('左')
            win32api.keybd_event(37, 0, 0, 0)  # 点击
            time.sleep(0.1)
            win32api.keybd_event(37, 0, win32con.KEYEVENTF_KEYUP, 0)

        elif voiceCommand == '右':
            print('右')
            win32api.keybd_event(39, 0, 0, 0)  # 点击
            time.sleep(0.1)
            win32api.keybd_event(39, 0, win32con.KEYEVENTF_KEYUP, 0)

if __name__ == '__main__':
    wordsToAdd = ['上', '下', '左', '右']
    speechReco = SpeechRecognition(wordsToAdd)
    while True:
        pythoncom.PumpWaitingMessages()
'''

#【day107】鼠标模拟
'''
import win32con
import win32api
import time

#设置鼠标位置
win32api.SetCursorPos([30, 40])
time.sleep(0.1)

#鼠标事件
win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0) #鼠标左键按下
win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0) #鼠标左键抬起
win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0) #鼠标左键按下
win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0) #鼠标左键抬起

#所有事件见：介绍API函数的中文帮助文件.chm
'''

#【day108】map与reduce
#map()
#原型：map(fn, lsd) fn:函数 lsd：序列
#功能：将传入的函数依次作用在序列中的每一个元素，并把结果作为新的Iterator返回
'''
def cha2int(chr):
    return {'0':0,'1':1,'2':2,'3':3,'4':4,'5':5,'6':6,'7':7,'8':8,'9':9}[chr]

list1 = ['2','1','6','5']
res = map(cha2int, list1)
#map的原理类似：[chr2int('2'), chr2int('1'), chr2int('6'), chr2int('5')]
print(res) #惰性列表
print(list(res)) #显示列表
'''
'''
#需求：将整数型元素的序列，转为字符串型
#[1,2,3,4] -> ['1','2','3','4']
list1 = [1,2,3,4]
res = map(str, list1)
print(list(res))
'''

#map大致的底层实现，内置的map函数适合分布式数据处理
'''
def myMap(func, lis):
    resList = []
    for parase in lis:
        res = func(parase)
        resList.append(res)
'''

#reduce(fn, lsd) 参数1：函数 参数2：列表
#功能：一个函数作用在序列上，这个函数必须接受两个参数，reduce把结果继续和序列的下一个元素累计运算
#reduce(f, [a,b,c,d])
#f(f(f(a,b),c),d)

'''
#求一个序列的和
from functools import reduce
list1 = [1, 2, 3, 4, 5]
def function(x, y):
    return x + y
r = reduce(function, list1)
print(r)
'''

#reduce(fn1, map(fn2, iterator))分布式处理数据作用很大
'''
#将字符串转化为对应字面量数字from functools import reduce
def str2int(str1):
    def fc(x, y):
        return x * 10 + y
    def fs(chr):
        return {'0': 0, '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9}[chr]
    return reduce(fc, map(fs, str1))
str1 = input('请输入数字：')
result = str2int(str1)
print(result)
print(type(result))
'''

#【day109】filter过滤器
#原型：filter(fn, lsd) fn函数 lsd序列
#功能：用于过滤序列
#白话文：把传入的函数依次作用于序列的每个元素，根据返回的是True还是False，决定是否保留该元素
'''
def function(num):
    return num % 2 == 0
list1 = [1,2,3,4,5,6]
list2 = filter(function, list1)
print(list(list2))
'''
#[function(1)=True 保留, function(2)=False 不保留]
#过滤excel数据，指定过滤其中某些元素

#过滤一行
'''
data = [['姓名', '年龄', '爱好'], ['tom', 25, '无'], ['hanmeimei', 26, '金钱']]
def myFilter(data):
    for i in data:
        if i == '无':
            return False
    return True
newData = filter(myFilter, data)
print(list(newData))
'''
#过滤其中指定元素
'''
data = [['姓名', '年龄', '爱好'], ['tom', 25, '无'], ['hanmeimei', 26, '金钱']]
def function(v):
    v = str(v)
    if v == '无':
        return False
    return True
newData = []
for line in data:
    lis = filter(function, line)
    newData.append(list(lis))
print(newData)
'''


#【day110】sorted
#排序：冒泡排序，选择排序，快速排序，插入排序，计数器排序
#普通排序 sorted()默认升序
'''
list1 = [4, 7, 2, 6, 3]
list2 = sorted(list1)
print(list1)
print(list2)
'''

#按绝对值升序排序
'''
list1 = [4, -7, -2, 6, 3]
list2 = sorted(list1, key=abs)#key接受函数来实现自定义排序规则
#list2 = sorted(map(abs, list1)) 大致原理相同
print(list1)
print(list2)
'''

#降序
'''
list1 = [2, 6, 4, 8, 9, 3]
list2 = sorted(list1, reverse=True)
print(list2)
'''

#按照字符串长短升序排序
'''
list1 = ['a435234','b123','c235779','d13452357']
list2 = sorted(list1, key=len)
print(list2)
'''
'''
def myLen(data): #函数可以自己写
    return len(data)
list1 = ['a435234','b123','c235779','d13452357']
list2 = sorted(list1, key=myLen)
print(list2)
'''


#【day111】单元测试
#作用：用来对一个函数、一个类或者一个模块来进行正确性校验工作的
#结果：1、单元测试通过，说明我们测试的函数功能正常
#     2、单元测试不通过，说明函数功能有Bug，要么测试条件输入有误

#测试函数
'''
def mySum(x, y):
    return x + y
def mySub(x, y):
    return x - y
'''
#测试函数最好另外新建一个文件myTest
'''
import unittest
#导入需要测试的函数
from qianfeng import mySum
from qianfeng import mySub

class Test(unittest.TestCase):
    def setUp(self): #正式开发时：可以写入链接数据库的函数
        print('开始测试时，自动调用')
    def teatDown(self): #正式开发时：可以写入断开数据库的函数
        print('结束测试时自动调用')
    def test_mySum(self): #为了测试mySum
        self.assertEqual(mySum(1, 2), 3, '加法有误') #断言相等（被测函数，正确值， 有误则提醒）
    def test_mySub(self):
        self.assertEqual(mySub(2, 1), 1, '减法有误')
if __name__ == '__mian__':
    unittest.main()
'''

#测试类
'''
class Person(object):
    def __init__(self, name, age):
        self.name = name
        self.age = age

    def getAge(self):
        return self.age + 1
'''
#单元类测试，写到另一个文件中,在定义测试函数的时候，函数名必须包含test，否则不进行测试
'''
import unittest
from knowledgePoints import Person
class Test(unittest.TestCase):
    def test_init(self):
        p = Person('hanmeimei', 20)
        self.assertEqual(p.name, 'hanmeimei', '属性赋值有误') #断言不等于（属性， 正确值，如果得到的值不等于正确值则报错）
    def test_getAge(self):
        p = Person('hanmeimei', 20)
        self.assertEqual(p.getAge(), p.age, 'getAge函数有误')

if __name__ == '__main__':
    unittest.main()
'''


#【day112】文档测试
'''
import doctest
'''
#doctest模块可以提取注释中的代码执行
#doctest模块严格按照python交互模式的输入提取
#执行格式：>>> 代码
'''
def mySum(x, y):
'''
'''
    get the sum from x and y
    :param x:firstNum
    :param y:secondNum
    :return:sum
    # 注意 文本测试交互符与代码中间必须有空格
    example:
    >>> print(mySum(1, 2))
    3
'''
'''
    return x + y + 1
print(mySum(1, 2))
'''
#进行文档测试
'''
doctest.testmod() #>>> 代码 ，交互得到的正确值，如果修改函数后，得不到交互的正确值，则报错
'''


#【day113】Telnet远程登录协议
#功能：远程登录同网络中的其他电脑
#条件：1、ping ip地址 能通，2、知道windows登录名和密码
#方法：1、控制面板 -> 程序和功能 -> 启动或关闭windos功能 -> 选中Telnet客户端
#     2、cmd -> telnet ip地址 -> 用户名和密码输入 -> 登录对方cmd成功
#获取信息：1、暴力破解密码
#进入桌面操作：cd user dir cd 用户名 dir cd Desktop mkdir 文件名
#python代码实现远程登录操作
'''
import telnetlib
def telnetDoSomethin(Ip, user, password, command):
    try:
        # 链接服务器
        telnet = telnetlib.Telnet(Ip)
        # 设置调试级别
        telnet.set_debuglevel(2)
        # 读取输入用户名信息
        rt = telnet.read_until('Login username:'.encode('utf-8'))
        # 写入用户名, \r\n：windows的cmd中的回车
        telnet.write((user+'\r\n').encode('utf-8'))
        # 读取输入密码信息
        rt = telnet.read_until('Login password:'.encode('utf-8'))
        telnet.write((password + '\r\n').encode('utf-8'))
        # 读取验证ip信息
        rt = telnet.read_until('Domain name:'.encode('utf-8'))
        telnet.write((Ip + '\r\n').encode('utf-8'))
        # 登录成功，写指令
        rt = telnet.read_until('>'.encode('utf-8'))
        telnet.write((command + '\r\n').encode('utf-8'))
        # 上面命令执行成功，会继续读到>，失败，一般不是>
        rt = telnet.read_until('>'.encode('utf-8'))
        #telnet.close()
        return True
    except:
        return False

if __name__ == '__main__':
    Ip = '10.0.142.197'
    user = 'xumingbin'
    password = ''
    command = 'tasklist'
    print(telnetDoSomethin(Ip, user, password, command))
'''

#【day114】破解密码
#密码：数字的排列组合


#排列：从n个不同元素中取出m（m≤n）个元素，按照一定的顺序排成一列，叫做从n个元素中取出m个元素的一个排列(permutation)。
#4个数中去3个数，总共有多少个排列方式：
#1 2 3 4
#_ _ _： 百位有4种取值方式，十位有3种取值方式，个位有2种取值方式，则总共有4*3*2种排列组合方式
'''
import itertools
from functools import reduce
mylist1 = list(itertools.permutations(['1', '2', '3', '4'], 3)) #迭代工具.排列（序列， 取多少个元素）
print(mylist1)
print(len(mylist1))
'''
'''
4-3 24 4个数字中取3个，排列的方式有24种
4-2 12 4个数字中取2个，排列的方式有12种
4-1 4 4个数字中取1个，排列的方式有4种
找规律，总结通项式
n!/(n-m)!
'''

#组合:从n个不同的元素中，任取m（m≤n）个元素为一组，叫作从n个不同元素中取出m个元素的一个组合。
'''
import itertools
mylist = list(itertools.combinations(['1', '2', '3', '4'], 3)) #迭代工具.组合（序列， 取多少个元素）
print(mylist)
'''
'''
m n s
5-5 1
5-4 5
5-3 10
5-2 10
m！/n!(m-n)!
'''

#排列组合
'''
import itertools
mylist = list(itertools.product('0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ', repeat=4)) #密码位数
'''

#密码破解
'''
import time
import itertools
iterator = (''.join(x) for x in itertools.product('0123456789', repeat=4))
while True:
    try:
        str1 = next(iterator)
        time.sleep(0.5)
    except StopIteration as e:
        break
'''


#【day115】正则表达式
#判断手机号码(非正则)
'''
def checkPhone(str1):
    if len(str1) != 11:
        return False
    elif str1[0] != '1':
        return False
    elif str1[1:3] != '30' and str1[1:3] != '31' and str1[1:3] != '80':
        return False
    for i in range(3, 11):
        if str1[i] < '0' or str1[i] > '9':
            return False
    return True
print(checkPhone('18086829907'))
#print(checkPhone('180868299074'))
#print(checkPhone('180868a9907'))
#print(checkPhone('28086829907'))
#print(checkPhone('19086829907'))
'''

#正则的概念
'''
python自1.5后，增加了一个re模块，提供了正则表达式模式
re模块使python语音拥有了全部的正则表达功能
'''

#【day116】re模块简介
'''
import re
'''

#re.match()
#原型：re.match(pattern, string, flags=0)
#参数：
    #patter:正则表达式的模式
    #string:要匹配的字符串
    #flags:标志位，用于控制正则表达式的匹配方式，值如下：
        #re.I   忽略大小写
        #re.L   做本地化识别
        #re.M   多行匹配，影响^和$
        #re.S   使.匹配包括换行符在内的所有字符
        #re.U   根据Unicode字符集解析字符，影响\w \W \b \B
        #re.X   使我们以更灵活的格式理解正则表达式
#功能：尝试从字符串的起始位置匹配一个模式，如果不是起始位置匹配成功的话，返回None

#判断www.baidu.com中是否有www
'''
print(re.match('www', 'www.baidu.com'))
print(re.match('www', 'www.baidu.com').span())#获得match对象，span表示匹配到的字符的索引值
print(re.match('www', 'ww.baidu.com'))
print(re.match('www', 'baidu.www.com'))
print(re.match('www', 'wwW.baidu.com', flags=re.I))
'''
#扫描整个字符串，返回从起始位置成功的匹配

#re.search()
#原型：search(pattern, string, flags=0)
#参数：同match()的参数
#功能：扫描整个字符串，并返回第一个成功的匹配
'''
print(re.search('www', 'www.baidu.com'))
print(re.search('www', 'www.baidu.com').span())
print(re.search('www', 'ww.baidu.com'))
print(re.search('www', 'baidu.www.com'))
print(re.search('www', 'baiduwwW.www.com', flags=re.I))

#提取匹配成功后的指定字段
html = "itemId":"610155403874"
obj = re.compile(r'"itemId":"(.*?)"')
itemID = obj.search(html)
print(itemID.group(1)) #提取正则表达式的括号部分
'''

#re.findall()
#原型：findall(pattern, string, flags=0)
#参数：同match()的参数
#功能：扫描整个字符串，并以列表的方式返回所有成功的匹配
'''
print(re.findall('www', 'www.baidu.com'))
print(re.findall('www', 'www.baidu.com'))
print(re.findall('www', 'ww.baidu.com'))
print(re.findall('www', 'baidu.www.com'))
print(re.findall('www', 'baiduwwW.www.com', flags=re.I))
'''

#匹配单个字符
# .:除换行符以外的任意字符
'''
print(re.match('.','www.baidu.com'))
print(re.search('.','www.baidu.com'))
print(re.findall('.','www.baidu.com'))
'''

# []:是字符集合，表示匹配中括号中所包含的任意一个字符
#[0123456789] 匹配单个数字
'''
print(re.match('[0123456789]', 'askdjfasjwe813asjlkfj34'))
print(re.search('[0123456789]', 'askdjfasjwe813asjlkfj34'))
print(re.findall('[0123456789]', 'askdjfasjwe813asjlkfj34'))
'''
#[abckef] 匹配其中的任意一个字母
'''
print(re.match('[abckef]', 'askdjfasjwe813asjlkfj34'))
print(re.search('[abckef]', 'skdjfasjwe813asjlkfj34'))
print(re.findall('[abckef]', 'askdjfasjwe813asjlkfj34'))
'''
#[a-z]        匹配任意小写字母
#[A-Z]        匹配任意大写字母
#[0-9]        匹配任意数字
#[0-9a-zA-Z]  匹配任意的数字和大小写字母
#[0-9a-zA-Z_] 匹配任意的数字和大小写字母和下划线
#[^justin]    匹配除了justin以外的所有字符,^称为脱字符，表示不匹配集合中的字符
#[^0-9]       匹配所有的非数字字符
#[\u4E00-\u9FA5]    匹配汉字
#\d           匹配所欲的数字，效果同[0-9]
#\D           匹配非数字字符，效果同[^\d]
#\w           匹配数字字母和下划线，效果同[0-9a-zA-Z_]
#\W           匹配非数字字母和下划线，效果同[^0-9a-zA-Z_]
#\s           匹配任意的空白符（空格，换行，回车，换页，制表），效果同[ \f\n\r\t]
#\S           匹配任意的非空白符（空格，换行，回车，换页，制表），效果同[^ \f\n\r\t]
'''
print(re.match('.', '31234awefawe\nasdf', flags=re.S))
print(re.search('.', '31234awefawe\nasdf', flags=re.S))
print(re.findall('.', '31234awefawe\nasdf', flags=re.S))
'''

#锚字符(便捷字符)
#^     行首匹配，和在[]里面的脱字符^不是一个意思
'''
print(re.findall('^31234', '31234awefawe\n31234')) #行首要是1234的数据
print(re.findall('^31234', '31234awefawe\n31234', flags=re.M)) #行首要是1234的数据
'''
#$     行尾匹配
'''
print(re.findall('fawe$', '31234awefawe\n31234fawe')) #行首要是1234的数据
print(re.findall('fawe$', '31234awefawe\n31234fawe', flags=re.M)) #行首要是1234的数据
'''
#\A    匹配字符串开始，它和^的区别是：\A只匹配整个字符串的开头，即使在re.M模式下也不会匹配它行的行首
'''
print(re.findall('\A31234', '31234awefawe\n31234')) #行首要是1234的数据
'''
#\Z    匹配字符串结束，它和$的区别是：\Z只匹配整个字符串的结束，即使在re.M模式下也不会匹配它行的行尾
'''
print(re.findall('fawe\Z', '31234awefawe\n31234fawe')) #行首要是1234的数据
print(re.findall('fawe\Z', '31234awefawe\n31234fawe', flags=re.M)) #行首要是1234的数据
'''
#\b    匹配一个单词的边界，也就是值单词和空格间的位置,word\b可以匹配以word 结尾的单词
'''
print(re.findall(r'word\b', 'word 341word_ word '))
'''
#\B    匹配一个非单词的边界，也就是值单词和空格间的位置
'''
print(re.findall(r'word\B', 'word 1241word_ word '))
'''

#字符数量修饰字符
#(xyz)    匹配小括号内的xyz(作为一个整体去匹配) 说明：下方的x、y、z、n、m均为假设的普通字符，不是正则表达式的字符
'''
print(re.findall(r'(justin)', 'justin123 just123in 123justin justin'))
'''
#x?       匹配0个或者1个x，非贪婪匹配，尽可能少的匹配
'''
print(re.findall(r'j?', 'justin123 just123in 123justin justin')) #j是一个j匹配，u是0个j匹配，因此空值其实是0个j匹配所得
'''
#x*       匹配0个x或任意多个x，贪婪匹配，尽可能多的匹配
'''
print(re.findall(r'a*', 'a,aa,aaa')) #a一个a的匹配，空是0个a的匹配，aa是两个a的匹配
'''
#.*       匹配0个或任意多个任意字符（换行符除外）
'''
print(re.findall(r'.*','112asdasd24  sdfasd234\n asdewrq'))
'''
#x+       匹配至少1个x,x个数>0
'''
print(re.findall(r'a+', 'fearaag'))
'''
#x{n}     匹配确定的n个x（n是非负整数）,{}中不能有空格
'''
print(re.findall(r'a{3}', 'a,aa,aaa,aaaa,aaaaa,aaaaaa,aaabaaa'))
'''
#x{n,}    匹配至少n个x,{}中不能有空格
'''
print(re.findall(r'a{3,}', 'a,aa,aaa,aaaa,aaaaa,aaaaaa,aaabaaa'))
'''
#x{n,m}   匹配至少n个x,至多m个x（n<=m,m非负整数）,{}中不能有空格
'''
print(re.findall(r'a{3,4}', 'a,aa,aaa,aaaa,aaaaa,aaaaaa,aaabaaa'))
'''
#x|y      |表示或，匹配一个x或一个y
'''
print(re.findall(r'a|b', 'a,b,aout,bout,aa'))
print(re.findall(r'((a|b)out)', 'a,b,aout,bout,aa'))
'''

#需求：提取justin.....man数据
'''
print(re.findall(r'^justin.*?man$', 'justin is good man ! justin is very good man'))
print(re.findall(r'justin.*?man', 'justin is good man ! justin is very good man'))
'''
#取消贪婪匹配
# *? +? x? 最小匹配
#通常都是尽可能多的匹配,我们可以使用这种方式解决贪婪匹配
'''
提取注释即注释中的内容/* part1 */ /* part2 */
print(re.findall(r'//*.*/*/', r'/* part1 */ /* part2 */'))
'''
#特殊
#(?:x)    类似(xyz),但不表示一个组

'''
#作业：
Mobile QQ Mail Phone Url 的正则表达式
'''


#【day117】re模块的深入
#re.split()字符串切割
'''
str1 = 'asdfa,,,,,asdfas,sdfga,asdfa'
print(str1.split(','))
print(re.split(r',+', str1)) #正则.切割(至少一个'，'号即',,,,,'也用作切割, 切谁：str1)
'''

#re.finditer()
#原型：finditer(pattern, string, flags=0)
#参数：同match search findall
#功能：与findall类似，扫描整个字符串，返回一个所有成功匹配的迭代器
'''
for i in re.finditer('justin.*?man', 'justin is good man ! justin is very good man'):
    print(i)
'''

#re.sub() re.subn()
#功能：替换
#原型：sub(pattern, repl, string, count=0, flags=0)
#     subn(pattern, repl, string, count=0, flags=0)
#参数：
#   pattern：正则表达式
#   repl：指定的用来替换的字符串
#   string：目标字符串
#   count：最多替换次数
#   flags：标志位
#功能：在目标字符串中，以正则表达式的规则匹配字符串，再把他们替换成指定的字符串,可以指定替换次数，如果不指定，则替换所有匹配字符串。
#区别：re.subn()会返回替换次数
'''
str1 = 'justin,Justin'
str2 = re.sub(r'justin', 'mia', str1, count=2, flags=re.I)
print(str2)
'''

#【day118】分组
#概念：除了简单的判断是否匹配之外，正则表达式还有提取子串的功能，用()括起来就表示一个组，组里面的数据可以使用group(num)方法提取，num表示第几组，当num=0时表示提取匹配到的字符串整体
'''
str1 = '010-88990022'
m = re.match('((\d{3})-(\d{8}))', str1)
'''
#使用序号获取对应组的信息，group(0)一直代表的是原始字符串
'''
print(m.group(0))
print(m.group(1))
print(m.group(2))
print(m.group(3))
'''
#查看匹配的各组的情况
'''
print(m.groups())
'''
#使用默认组名匹配字符串,默认组名\1\2\3...表示第一个组的名字，第二个组的名字...
'''
import re
Str = '<p><div><spon>猪八戒</spon></div></p>'
pat = r'<(\w+)><(\w+)>\w+</\2></\1>'
re_mobile = re.compile(pat)
data = re_mobile.search(Str)
print(data.group(0))
'''
#给分组命名,按分组名提取数据
'''
str1 = '010-88990022'
m = re.match('((?P<first>\d{3})-(?P<secent>\d{8}))', str1)
print(m.group('first'))
'''
#findall()和分组的运用
# import re
# Str = 'ab123cd456efg'
# pat = r'ab(\d+?)cd(\d+?)efg'
# pattern = re.compile(pat)
# lis = pattern.findall(Str)
# print(lis) ---> [('123', '456')]
#说明
#    findall()返回的是一个列表，列表中的元素都是元组，元组中第一个元素就是正则中第一个小括号匹配到的内容，元组中的二个元素就是正则中第二个小括号匹配到的内容
#    因此需要获取什么内容，只需要在表达式中加括号即可

#【day119】编译
#概念:当我们使用正则表达式时，re模块会做两件事
#    1、编译正则表达式，如果正则表达式本身不合法，会报错
#    2、用编译后的正则表达式去匹配对象
#re.compile()
#原型：compile(pattern, flags=0)
#参数：pattern:要编译的正则表达式，flags:标志位
#功能：编译正则表达式，并返回正则表达式对象
'''
pat = r'^1(([3578]\d)|(47))\d{8}$'
re_mobile = re.compile(pat)
data = re_mobile.match('18086829907')
print(data.group(0))
re_mobile.findall()
'''

#【day120】总结
#编译前
'''
re.match(pattern, string, flags)
re.search(pattern, string, flags)
re.findall(pattern, string, flags)
re.finditer(pattern, string, flags)
re.split(pattern, string, flags)
re.sub(pattern, string, flags)
re.subn(pattern, string, flags)
'''
#编译后
'''
object = re.compile(pattern)
object.match(string)
object.search(string)
object.findall(string)
object.finditer(string)
object.split(string)
object.sub(string)
object.subn(string)
'''

#阿里云账号名:龙良雨good

#【day121】爬虫
#什么是爬虫：
#    可以自动化浏览网页中的信息，当然浏览信息的时候需要按照我们制定的规则进行，这些规则称之为网络爬虫算法
#为什么学习爬虫：
#    私人定制一个搜索引擎，并且可以对搜索引擎的数据采集工作原理进行更深层次地理解
#    获取更多的数据源，并且这些数据源可以按我们的目的进行采集，去掉很多无关数据
#    更好地进行SEO(搜索引擎优化)
#网络爬虫的组成
#    控制节点
#        叫做爬虫中央控制器，主要负责根据url地址分配线程，并调用爬虫节点进行具体的爬行
#    爬虫节点
#        按照相关的算法，对网页进行具体的爬行，主要包括下载网页以及对网页的文本处理，爬行后会将对应的爬行结果存储到对应的资源库中
#    资源库构成
#        存储爬虫爬去到的响应数据，一般为数据库
#爬虫设计思路
#    首先确定需要爬取的网页URL地址
#    通过HTTP协议来获取对应的HTML页面
#    提取HTML页面里的有用数据
#        如果是需要的数据就保存起来
#        如果页面里是其他的URL，那么就继续执行第二步
#为什么要用python去做爬虫
#    PHP
#        虽然是世界上最好的语言，但是先天不是干爬虫的命，PHP对多线程、异步支持不足，并发不足。爬虫是工具性程序，对速度和效率要求较高
#    Jave
#        生态圈完善，是python最大对手，但java本身很笨重，代码量大，重构成本比较高，任何修改都会导致大量代码的变动。最要命的是爬虫需要经常修改部分代码
#    C/C++
#        运行效率和性能几乎最强，但学习成本非常高，代码成型较慢，能用C/C++写爬虫，说明能力很强，但是不是最正确的选择
#    Python
#        语法优美、代码简洁、并发效率高、三方模块多，调用其他接口方便。有强大的爬虫Scrapy，以及成熟高效的scrapy-redis分布式策略
#需要技能
#    python语法
#    如何抓取页面
#        HTTP请求处理，urllib处理后的请求可以模拟浏览器发送请求，获取服务器响应的文件
#    解析服务器响应内容
#        re、xpath、BeautifulSoup4、jsonpath练习、pyquery
#        目的是使用某种描述性语法来提取匹配规则的数据
#    如何采集动态HTML、验证码的处理
#        通用的动态页面采集
#            selenium练习+PhantomJS(无界面浏览器),模拟真实浏览器加载js、ajax等动态页面数据
#        Tesseract
#            机器学习库，机器图像识别系统（识别图片中的文本）
#    scrapy框架
#        中国常见的框架Scrapy、Pyspider
#        高定制性高性能(异步网络框架twisted)，所以数据下载速度非常快，提供了数据存储、数据下载、提取规则等组件
#    分布式策略
#        scrapy-redis
#        在Scrapy的基础上添加了一套以Redis数据库为核心的一套组件，让scrapy框架支持分布式的功能，主要在Redis里做请求指纹去重，请求分配、数据临时存储
#爬虫与反爬虫与反反爬虫三角之争
#    最头痛的人
#        爬虫做到最后，最头痛的不是复杂的页面，也不是晦涩的数据，而是网站另一头的反爬虫人员
#    反爬虫技术
#        User-Agent：检测是否是程序在访问服务器
#        代理：检测高频访问ip并屏蔽
#        验证码：检测是否是人工操作
#        动态数据加载：不展示所有数据
#        加密数据：数据抓取后不展示正确数据
#    是否要反爬虫
#        机器成本+人力成本 > 数据价值，就不反了，一般做到封ip就可以了
#        面子的战争
#    爬虫与反爬虫之间的战争，最后一定是爬虫胜利
#网络爬虫的类型
#    通用网络爬虫
#        概念
#            搜索引擎用的爬虫系统
#        用户群体
#            搜索引擎用的爬虫系统
#        目标
#            尽可能把互联网上的所有网页下载下来，放到本地服务器里形成备份。再对这些网页做相关处理(提取关键字、去掉广告等)，最后提供一个用户检测接口
#        抓取流程
#            首先选取一部分已有的URL，把这些URL放到待爬队列
#            从队列里提取出这些URL，首先在本地hots文件里面去提取ip（hots会保存你曾经访问过的网站以及对应ip），hots中如果没有，就会解析DNS（云端服务器中有访问过的网站及对应ip备份）得到主机IP，然后去这个IP对应的服务器下载HTML页面，保存到搜索引擎的本地服务器里，之后把爬过的URL放入已爬取队列
#            分析这些网页内容，找出网页里的URL链接，继续执行第二步，直到爬取条件结束
#        搜索引擎如何获取一个新网站的URL
#            主动向搜索引擎提交网址(百度站长平台) http://zhanzhang.baidu.com/linksubmit/url
#            在其他网站里设置网站的外链接 开一个帖子放上你的新网站网址or友情链接
#            搜索引擎会和DNS服务商合作，可以快速收录新的网站 ip备案、域名注册时DNS就知道你的新网站网址了
#        Robots协议 通用爬虫并不是万物皆可爬的，它需要遵守规则
#            协议会指明通用爬虫可以爬取网页的权限
#            Robots.txt只是一个建议，并不是所有爬虫都遵守，一般只有大型的搜索引擎爬虫才会遵守 www.taobao.com/robots.txt
#        通用爬虫工作流程
#            爬取网页
#            存储数据
#            内容处理
#            提供检索/排名服务
#                搜索引擎排名
#                    PageRank值
#                        根据网站的流量(点击量、浏览值、人气)统计，流量越高，网站越值钱、排名越靠前
#                    竞价排名
#                        谁钱多谁靠前
#        通用网络爬虫的缺点
#            只能提供和文本相关的内容(HTML、Word、PDF)等，但是不能提供多媒体(音乐、图片、视频)和二进制文件(程序、脚本)等
#            提供结果千篇一律，不能针对不同人群提供不同的搜索结果
#            不能理解人类语义上的检索
#    聚焦网络爬虫
#        概念
#            爬虫程序员写的针对某种内容的爬虫
#        特点
#            面向主题爬虫，面向需求爬虫
#                会针对某种特定的内容去爬取信息，而且会保证信息和需求尽可能相关
#    增量式网络爬虫
#    深层网络爬虫
#URL
#    同一资源定位符，是互联网上的资源地址
#    https://   www.aspxfans.com   :8080/    news/    index.asp    ?boardID=5&ID=5&ID=2345&page=1    #name
#    协议       域名                 端口号     目录       文件名         参数：浏览器给服务器发送的数据       锚：不太用，初始界面的位置定位
#安装Charies抓包工具
#    下载地址：https://www.charlesproxy.com/
#    安装方式：下一步，知道完成
#    安装完成后打开charies软件
#    设置需要爬取网站的https
#        参考：https://blog.csdn.net/weixin_42336579/article/details/80621410
#        在pc端安装证书
#           生成证书
#               点击Help
#                   点击SSL Proxying
#                       点击Install Charies Root Certificate
#                       点击安装证书
#                       点击下一步
#                       点击将所有的证书都放入下列存储
#                       点击浏览
#                       点击受信任的根证书颁发机构
#                       点击确定
#                       点击下一步
#                       点击完成
#                       点击确定
#                       点击确定
#           导出证书
#               点击Help
#                   点击SSL Proxying
#                       点击Save Charles Root Certificate
#                           选择桌面
#                               文件夹名：C:\Users\surface\Desktop\ssl
#                               文件类型：Binary certificate(.cer)
#           谷歌浏览器安装证书
#               地址栏输入chrome://settings/
#                   搜索管理证书
#                   点击管理证书
#                       点击导入
#                           下一步
#                               点击浏览
#                               选择刚才导出到桌面的证书
#                               点击打开
#                               点击完成
#           设置代理监听的网站
#               点击Proxy
#                   点击SSL Proxying Settings
#                       点击add
#                           host：*
#                           port：*
#                           点击ok
#        点击add 将需要爬取的https的域名部分添加入host port：443
#        ok ok
#header
#    Get                客户端访问服务器的请求方式
#    Host	            请求的主机 movie.douban.com
#    Connection	        长链接 keep-alive(百度账号登录后，重新打开或打开百度新闻，账号是保持登录状态)
#    Accept	            请求对象 application/json, text/plain, */*
#    User-Agent	        浏览器信息 Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.108 Safari/537.36
#        KHTML              内核 Chrome/74.0.3729.108 谷歌浏览器及版本
#    Referer	        发起请求的网址 https://movie.douban.com/tag/
#    Accept-Encoding	返回内容的压缩方式 gzip, deflate, br
#    Accept-Language	编码 zh-CN,zh;q=0.9
#    Cookie             登录存储的信息


#【day122】urllib爬虫
'''
import urllib.request
'''
#向指定的url地址发起请求，并返回服务器返回响应的数据(文件对象)
'''
response = urllib.request.urlopen('http://www.baidu.com')
'''
#读取方式一：内容全部读取，会把读取到的数据赋值给字符串变量
'''
data = response.read()
#将爬取到的信息存入文件
path = r'D:\qian_feng_education\first_project\baidu.html'
with open(path, 'wb') as f:
    f.write(data)
'''
#读取方式二：内容行读取
'''
data = response.readline()
'''
#读取方式三：内容全部读取，会把读取到的数据赋值给一个列表变量，元素为html的一行
'''
data = response.readlines()
print(type(data))
print(len(data))
print(type(data[0]))
print(type(data[0].decode('utf-8')))
'''

#response的方法

#.info()
#功能：返回当前环境的有关信息
'''
print(response.info())
'''
#.getcode()
#功能：返回状态码
'''
print(response.getcode())
if response.getcode() == 200 or response.getcode() == 304:
    #去处理信息
    pass
'''
#.geturl()
#功能：返回当前正在爬取的url地址
'''
print(response.geturl())
'''

#urllib.request.unquote(url)
#功能：汉字解码,将发送请求url中的汉字编码转为汉字
'''
url = r'https://www.baidu.com/s?wd=%E9%BE%99%E8%89%AF%E9%9B%A8&rsv_spt=1&rsv_iqid=0xa3059481000da4ee&issp=1&f=8&rsv_bp=1&rsv_idx=2&ie=utf-8&rqlang=cn&tn=baiduhome_pg&rsv_enter=1&oq=%25E7%258A%25B6%25E6%2580%2581%25E7%25A0%2581&rsv_t=94f8EoWpEPsdsU7QkvIySdfYMngr%2Bwp4HBpL7r2Xi3muYB%2B4wDLYWZ4eFi0FizC1LHcm&inputT=1556&rsv_pq=abf3a4b90011516f&rsv_sug3=24&rsv_sug1=16&rsv_sug7=100&rsv_sug2=0&rsv_sug4=1556'
newUrl = urllib.request.unquote(url)
print(newUrl)
'''

#urllib.request.quote(url)
#功能：汉字编码,将发送请求url中的汉字转为汉字编码
'''
NewUrl = urllib.request.quote(newUrl)
print(NewUrl)
'''

'''
#test
import urllib.request
url = 'http://www.baidu.com'
newUrl1 = urllib.request.quote(url)#编码
newUrl2 = urllib.request.unquote(newUrl1)#解码
print(newUrl1)
'''

#urllib.request.urlretrieve()
#功能：直接将url取回到文件
'''
import urllib.request
url = 'http://www.baidu.com'
path = r'D:\qian_feng_education\first_project\test.html'
urllib.request.urlretrieve(url, path)
'''
#问题一：产生缓存
#解决：清除缓存
'''
urllib.request.urlcleanup()#清除缓存
'''
#问题二：容易被反爬虫屏蔽
#解决：模拟浏览器
'''
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.108 Safari/537.36'
}
'''
#设置请求体
#原型：urllib.request.Request(url, data=None, headers={},origin_req_host=None, unverifiable=False,method=None)
#参数：data：需要传参数的时候使用
#     headers：请求头
#功能：
'''
req = urllib.request.Request(url, headers=headers)
'''
#发起请求
'''
response = urllib.request.urlopen(req)
'''
#阅读信息
'''
result = response.read().decode('utf-8')
print(result)
'''


#【day123】伪造User-Agent请求头
#模拟真实浏览器访问网站
'''
import urllib.request
import random
AgentList = [
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 OPR/26.0.1656.60 Opera/8.0 (Windows NT 5.1; U; en)',
    'Mozilla/5.0 (Windows NT 5.1; U; en; rv:1.8.1) Gecko/20061208 Firefox/2.0.0 Opera 9.50',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; en) Opera 9.50',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:34.0) Gecko/20100101 Firefox/34.0',
    'Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.57.2 (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Chrome/10.0.648.133 Safari/534.16',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER',
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; LBBROWSER)',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E; LBBROWSER)"',
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; QQBrowser/7.0.3698.400)',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)',
    'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 SE 2.X MetaSr 1.0',
    'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; SE 2.X MetaSr 1.0)',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Maxthon/4.4.3.4000 Chrome/30.0.1599.101 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/38.0.2125.122 UBrowser/4.0.3214.0 Safari/537.36',
    'Mozilla/5.0 (iPhone; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5',
    'Mozilla/5.0 (iPod; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5',
    'Mozilla/5.0 (iPad; U; CPU OS 4_2_1 like Mac OS X; zh-cn) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8C148 Safari/6533.18.5',
    'Mozilla/5.0 (iPad; U; CPU OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5',
    'Mozilla/5.0 (Linux; U; Android 2.2.1; zh-cn; HTC_Wildfire_A3333 Build/FRG83D) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1',
    'Mozilla/5.0 (Linux; U; Android 2.3.7; en-us; Nexus One Build/FRF91) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1',
    'MQQBrowser/26 Mozilla/5.0 (Linux; U; Android 2.3.7; zh-cn; MB200 Build/GRJ22; CyanogenMod-7) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1',
    'Opera/9.80 (Android 2.3.4; Linux; Opera Mobi/build-1107180945; U; en-GB) Presto/2.8.149 Version/11.10',
    'Mozilla/5.0 (Linux; U; Android 3.0; en-us; Xoom Build/HRI39) AppleWebKit/534.13 (KHTML, like Gecko) Version/4.0 Safari/534.13',
    'Mozilla/5.0 (BlackBerry; U; BlackBerry 9800; en) AppleWebKit/534.1+ (KHTML, like Gecko) Version/6.0.0.337 Mobile Safari/534.1+',
    'Mozilla/5.0 (hp-tablet; Linux; hpwOS/3.0.0; U; en-US) AppleWebKit/534.6 (KHTML, like Gecko) wOSBrowser/233.70 Safari/534.6 TouchPad/1.0',
    'Mozilla/5.0 (SymbianOS/9.4; Series60/5.0 NokiaN97-1/20.0.019; Profile/MIDP-2.1 Configuration/CLDC-1.1) AppleWebKit/525 (KHTML, like Gecko) BrowserNG/7.1.18124',
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows Phone OS 7.5; Trident/5.0; IEMobile/9.0; HTC; Titan)',
    'UCWEB7.0.2.37/28/999',
    'NOKIA5700/ UCWEB7.0.2.37/28/999',
    'Openwave/ UCWEB7.0.2.37/28/999',
    'Mozilla/4.0 (compatible; MSIE 6.0; ) Opera/UCWEB7.0.2.37/28/999'
]
agentStr = random.choice(AgentList)
url = 'http://www.baidu.com'
res = urllib.request.Request(url)
#向请求体里添加User-Agent
res.add_header('User-Agent', agentStr)
result = urllib.request.urlopen(res)
data = result.read().decode('utf-8')
print(data)
'''


#【day124】完整请求头
'''
headers = {
    'User-Agent': 'Openwave/ UCWEB7.0.2.37/28/999', #模拟浏览器
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3', #允许访问的格式
    'X-Requested-With': 'XMLHttpRequest', #请求对象类型
    'Content-Type': 'text/html;charset=utf-8' #链接类型
}
'''


#【day125】设置超时
#如果网页长时间未响应，系统应判断超时，无法爬取
'''
import urllib.request
for i in range(100):
    try:
        response = urllib.request.urlopen('http://www.baidu.com', timeout=0.01)
    except urllib.error.URLError as e:
        print('请求超时，爬取下一个ulr')
'''


#【day126】http请求
#使用场景：进行客户端与服务端之间的消息传递时使用
#http请求
#   GET     通过URL网址传递信息，可以直接在URL网址上添加传递信息（百度搜索龙良雨，龙良雨包含在URL中传递给百度服务器，这就叫GET请求）
#   POST    可以向服务器提交数据，是一种比较流行的，比较安全的数据传递方式（URL中？号后面的参数，就是post，因为它是将数据打包传送给服务器，包截获不了，因此安全。常用语修改服务器中的变量值，比如修改秘密、修改账号等）
#   PUT     请求服务器存储一个资源，通常要指定存储的位置
#   DELETE  请求服务器删除一个资源
#   HEAD    请求获取对应的HTTP报头信息（传送信息前会有一个数据包，数据包就是报头，报头中存储了端口号、主机ip等）
#   OPTIONS 可以获取当前URL所支持的请求类型


#GET请求
#特点：把数据拼接到请求路径的后面传递给服务器
#优点：速度快
#缺点：承载的数据量小，不安全


#【day127】json数据解析
#概念：它是一种保存数据的格式
#作用：可以保存本地的json文件，也可以将json串进行传输，通常将json称为轻量级的传输方式
#json的组成: {} 字典
#           [] 列表
#           ,  分割两个部分
#           :  键值对
#jsonStr =
'''{
  "breakfast_menu": {
    "food": [
      {"calories": "650",
      "description": "two of our famous Belgian Waffles with plenty of real maple syrup",
      "name": "Belgian Waffles",
      "price": "$5.95"
      },{
        "calories":"900",
        "description": "light Belgian Waffles covered with strawberries and whipped cream",
        "name": "Strawberry Belgian Waffles",
        "price": "$7.95"
      },{
        "calories":"900",
        "description": "light Belgian Waffles covered with an assortment of fresh berries and whipped cream",
        "name": "Berry-Berry Belgian Waffles",
        "price": "$8.95"
      },{
        "calories":"600",
        "description": "thick slices made from our homemade sourdough bread",
        "name": "French Toast",
        "price": "$4.5"
      },{
        "calories":"950",
        "description": "two eggs, bacon or sausage, toast, and our ever-popular hash browns",
        "name": "Homestyle Breakfast",
        "price": "$6.95"
      }
    ]
  }
}'''

#json.loads(jsonStr)
#功能：json格式的字符串转为python的数据类型字典
'''
import json
jsonData = json.loads(jsonStr)
#print(jsonData)
for i in range(5):
    print(jsonData['breakfast_menu']['food'][i]['price'])
'''

#json.dumps(pythonDict)
#功能：将python数据类型字典转为json格式
'''
jsonDict = {
  "breakfast_menu": {
    "food": [
      {"calories": "650",
      "description": "two of our famous Belgian Waffles with plenty of real maple syrup",
      "name": "Belgian Waffles",
      "price": "$5.95"
      },{
        "calories":"900",
        "description": "light Belgian Waffles covered with strawberries and whipped cream",
        "name": "Strawberry Belgian Waffles",
        "price": "$7.95"
      },{
        "calories":"900",
        "description": "light Belgian Waffles covered with an assortment of fresh berries and whipped cream",
        "name": "Berry-Berry Belgian Waffles",
        "price": "$8.95"
      },{
        "calories":"600",
        "description": "thick slices made from our homemade sourdough bread",
        "name": "French Toast",
        "price": "$4.5"
      },{
        "calories":"950",
        "description": "two eggs, bacon or sausage, toast, and our ever-popular hash browns",
        "name": "Homestyle Breakfast",
        "price": "$6.95"
      }
    ]
  }
}
import json
jsonStr = json.dumps(jsonDict)
print(jsonStr)
print(type(jsonStr))
'''

#json.load()
#功能：读取本地的json文件，返回jsonDict
'''
import json
path = r'D:\qian_feng_education\first_project\json\caidanJson.json'
with open(path, 'rb') as f:
    jsonStr = json.load(f)
print(jsonStr)
print(type(jsonStr))
'''

#json.dump(jsonDict, filePath)
#功能：json写入本地
'''
jsonDict = {
  "breakfast_menu": {
    "food": [
      {"calories": "650",
      "description": "two of our famous Belgian Waffles with plenty of real maple syrup",
      "name": "Belgian Waffles",
      "price": "$5.95"
      },{
        "calories":"900",
        "description": "light Belgian Waffles covered with strawberries and whipped cream",
        "name": "Strawberry Belgian Waffles",
        "price": "$7.95"
      },{
        "calories":"900",
        "description": "light Belgian Waffles covered with an assortment of fresh berries and whipped cream",
        "name": "Berry-Berry Belgian Waffles",
        "price": "$8.95"
      },{
        "calories":"600",
        "description": "thick slices made from our homemade sourdough bread",
        "name": "French Toast",
        "price": "$4.5"
      },{
        "calories":"950",
        "description": "two eggs, bacon or sausage, toast, and our ever-popular hash browns",
        "name": "Homestyle Breakfast",
        "price": "$6.95"
      }
    ]
  }
}
import json
path = r'D:\qian_feng_education\first_project\json\test.json'
with open(path, 'w') as f:
    json.dump(jsonDict, f)
'''


#【day128】post请求
'''
from myClass.headers import Headers
headers = Headers()
headersDict = headers.allHeaders()

#特点：把参数进行打包，单独传输
#优点：承载数据量大，安全（当对服务器数据进行修改时，建议使用post）
#缺点：数据量大，速度慢
import urllib.request
import urllib.parse #对参数进行打包的库
#网站登录
url = 'http://user.51sole.com/Default.aspx'
#将要发送的数据合成一个字典
#字典的键去网址里找，一般为input标签的name属性的值
data = {
    "txtUserName": "chenyilong",
    "txtPwd": "135cylpsx"
}
#对要发送的数据进行打包
postData = urllib.parse.urlencode(data).encode('utf-8')
#请求体
req = urllib.request.Request(url, data=postData, headers=headersDict)
#发起请求
response = urllib.request.urlopen(req)
#阅读数据
html = response.read().decode('utf-8')
#写入文件
path = r'D:\qian_feng_education\first_project\test1.html'
with open(path, 'w', encoding='utf-8') as f:
    f.write(html)
'''


#【day129】抓取网页动态Ajax请求的数据
#案例网站：豆瓣
#爬取内容：json数据
'''
import urllib.request
import ssl #解析https的包
from myClass.headers import Headers
import json
import time
def ajaxCrawler(url):
    headers = Headers()
    headersDict = headers.allHeaders()
    req = urllib.request.Request(url, headers=headersDict)
    #使用ssl创建一个不需要验证的环境 #解决https问题
    context = ssl._create_unverified_context()
    response = urllib.request.urlopen(req, context=context)
    jsonStr = response.read().decode('utf-8')
    return json.loads(jsonStr)

#抓取一次动态页面

#url = 'https://movie.douban.com/j/chart/top_list?type=17&interval_id=100%3A90&action=&start=20&limit=20'
#jsonDict = ajaxCrawler(url)
#print(jsonDict)

#抓取所有动态页面
#url分析
#   页面有20个数据，再加载20个数据   start=20&limit=20
for i in range(1, 3):
    url = 'https://movie.douban.com/j/chart/top_list?type=17&interval_id=100%3A90&action=&start={}&limit=20'.format((i*20))
    jsonDict = ajaxCrawler(url)
    print(jsonDict)
    time.sleep(3)
'''

#案例网站：糗事百科爬虫
#爬取内容：文本
'''
from myClass.headers import Headers
import urllib.request
import time
import re
import pickle


def jockCrawler(url):
    headers = Headers()
    headersDict = headers.allHeaders()
    res = urllib.request.Request(url, headers=headersDict)
    response = urllib.request.urlopen(res)
    time.sleep(10)
    HTML = response.read().decode('utf-8')
    pat = r'<div class="author clearfix">(.*?)<span class="stats-vote"><i class="number">'
    re_joke = re.compile(pat, re.S) #如果编码urf-8和正则表达式都没问题，任然出现提取零数据，则让.匹配换行
    divList = re_joke.findall(HTML)
    jokeDict = {}
    num = 0
    for div in divList:
        time.sleep(5)
        #用户名
        patH2 = r'<h2>(.*?)</h2>'
        reH2 = re.compile(patH2, re.S)
        reName = reH2.findall(div)[0]
        reName = reName.replace('\n', '')

        #段子
        patSpan = r'<div class="content">\n<span>(.*?)</span>' #html显示了换行就有换行符
        reSpan = re.compile(patSpan, re.S)
        reContent = reSpan.findall(div)[0]
        reContent = reContent.replace('\n', '')
        #存字典
        num += 1
        print('成功爬取{}条'.format(num))
        jokeDict[reName] = reContent
    return jokeDict

List = []
for page in range(1, 14):
    url = r'https://www.qiushibaike.com/text/page/{}/'.format(page)
    jokeDict = jockCrawler(url)
    List.append(jokeDict)
    print('成功爬取第{}页'.format(page))

path = r'D:\qian_feng_education\first_project\jock.txt'
with open(path, 'wb') as f:
    pickle.dump(List, f)
'''

#案例网站：一号店女装
#爬取内容：图片
'''
from myClass.headers import Headers
import urllib.request
import re
import os

def imageCrawler(url, toPath):
    headers = Headers()
    getHeaders = headers.allHeaders()
    res = urllib.request.Request(url, headers=getHeaders)
    response = urllib.request.urlopen(res)
    HTML = response.read().decode('utf-8')
    #html用wb写入文件时不需要.decode()，这样才不乱码
    #HTMLpath = r'D:\qian_feng_education\first_project\myClass\一号店女装\first.html'
    #with open(HTMLpath, 'wb') as f:
        #f.write(HTML)
    pat = r'<div style="position: relative">\n<img (src)|(original)="//(.*?)/>'
    re_imgList = re.compile(pat, re.S)
    imgList = re_imgList.findall(HTML)
    fileNum = 0
    for i in imgList:
        if i[2] != '':
            imageUrl = i[2].rstrip(' ').rstrip('\"')
            absPath = os.path.join(toPath, '{}.jpg'.format(fileNum))
            fileNum += 1
            #把图片存入
            urllib.request.urlretrieve('http://' + imageUrl, filename=absPath)

url = r'https://search.yhd.com/c9719-0-0/mbname-b/a-s1-v0-p1-price-d0-f0b-m1-rt0-pid-mid0-color-size-k%E5%A5%B3%E8%A3%85/'
toPath = r'D:\qian_feng_education\first_project\myClass\一号店女装\image'
imageCrawler(url, toPath) 下载图片 保存图片
#因为一号店的图片不是一次性全部加载，而是通过ajax获取的图片再加载，因此以上方法无法获取全部图片
'''

#案例网站：豆瓣小组——留下你的qq号，看有多少人加你（百度：留下你的qq号）
#爬取内容：QQ号
'''
import urllib.request
import ssl
from myClass.myCrawler import MyCrawler
from myClass.regularExpression import RegularExpression
from myClass.rwfile import RwFile
from collections import deque
import time

def writeFileBytes2Bytes(htmlBytes, toPath):
    with open(toPath, 'wb') as f:
        f.write(htmlBytes)

def writeFileBytes2String(htmlBytes, toPath):
    with open(toPath, 'w') as f:
        f.write(str(htmlBytes))

#得到一个HtmlBytes
def getHtmlBytes(url):
    get_headers = MyCrawler()
    headers = get_headers.getHeaders()
    res = urllib.request.Request(url, headers=headers)
    context = ssl._create_unverified_context()  # 创建一个不验证的环境：解决https安全拦截
    time.sleep(10)
    response = urllib.request.urlopen(res, context=context)
    return response.read()

def qqCrawler(url, toPath):
    htmlBytes = getHtmlBytes(url)
    #toPath1 = r'D:\qian_feng_education\first_project\myClass\豆瓣qq号\file1.html'
    #toPath2 = r'D:\qian_feng_education\first_project\myClass\豆瓣qq号\file2.txt'
    #writeFileBytes2Bytes(htmlBytes, toPath1)
    #writeFileBytes2String(htmlBytes, toPath2)
    #爬取QQ号
    htmlStr = str(htmlBytes)
    pat = RegularExpression()
    re_QQ = pat.getQQ(htmlStr)
    qqList = list(set(re_QQ))
    rwFile = RwFile()
    rwFile.listToTxt(toPath, qqList)
    print('成功爬取一页qq号')
    #爬取网址
    re_Url = pat.getUrl(htmlStr)
    urlList = list(set(re_Url))
    return urlList

def center(url, toPath):
    queue = deque()
    queue.append(url)
    while len(queue) != 0:
        targetUrl = queue.popleft()
        time.sleep(5)
        urlList = qqCrawler(targetUrl, toPath)
        for urlTuple in urlList:
            needUrl = urlTuple[0]
            queue.append(needUrl)

if __name__ == '__main__':
    url = 'https://www.douban.com/group/topic/17359302/?start=0'
    toPath = r'D:\qian_feng_education\first_project\data\QQ号\qqList.txt'
    center(url, toPath)
'''

#网络编程
#   网络概述
#       自从互联网诞生以来，现在基本上所有的程序都是网络程序，很少有单机版的程序了
#       计算机网络就是把各个计算机链接到一起，让网络中的计算机可以互相通信。网络编程就是如何在程序中实现两台计算机的通信
#       用python进行网络编程，就是在python程序本身这个进程内，链接别的服务器进程的通信端口进行通信（要通信需要知道另一台计算机的ip和端口）
#   IPv4包头协议
#       发送通信信息，包含：(4位版本号、4位首部长度、8位服务类型、16位总长度（字节数）、16位标识、3位标识、13位片偏移、8位生存时间、8位协议、16位首部检验和、32位原IP地址、32位目的IP地址)以上信息总共20字节、选项（如果有——TCP协议）12bOr32b、最后才是数据
#           版本：IP头中前4位标识了IP的操作版本，比如版本4或版本6
#           Internet头长度：头中下面4位包括头长度，以32单位表示
#           服务类型
#           总长度（Total Length）
#           标识（Identifier）：每个IP报文被赋予一个唯一的16位标识，用于标识数据报的分段
#           分段标志（Fragmentation Flag）：下一个域包括3个1位标志，标识报文是否允许被分配和是否使用了这些域
#           分段偏移（Fragmentation Offest）：8位的域指出分段报文相对于整个报文开始处的偏移。这个值以64位为单位递增
#           生存时间（TTL）：IP报文不允许在广域网中永久漫游。它必须限制在一定的TTL内
#           协议：8位域指示IP头之后的协议，如VINES、TCP、UDP等
#           效验和（checksum）：效验和是16位的错误检测域。目的机的网络中的每个网关要重新计算报文头的效验和，就如同源机器所做的一样
#           源IP地址：源计算机IP地址
#           目的IP地址：目的计算机的IP地址
#           填充：为了保证IP头长度是32位的整数倍，要填充额外的0
#   网络七层模型
#       物理层
#       数据链路层
#       网络层
#       传输层
#       会话层
#       表示层
#       应用层
#   TCP包头结构
#       TCP数据在IP数据报中的封装
#           IPv4数据包：20字节IP首部、20字节TCP首部、TCP数据
#       TCP包首部
#           包括：16位端口号、16位目的端口号、32位序号、32位确认序号、4位数据偏移、6位保留、标志（URG\ACK\PSM\RST\SYN\FIN）、16位窗口大小、16位检验和、16位紧急指针、选项、数据
#       结构
#           TCP源端口
#               16位的源端口域包含初始化通信的端口号
#               源端口和源IP地址的作用是标识报文的返回地址
#           TCP目的端口
#               16位的目的端口域定义传输的目的
#               这个端口指明报文接收计算机上的应用程序地址接口
#           TCP序列号：32位的序列号由接收端计算机使用，重组分段的报文成最初形状
#           TCP应答号（确认序列号）：TCP使用32位的应答（ACK）域标识下一个希望收到的报文的第一个字节
#           数据偏移
#               这个4位域包括TCP头大小
#               以32位数据结构或称为'双宇'为单位
#           保留
#               6位恒置0的域
#               为将来定义新的用途保留
#           标志
#               6位标志域，每1位标志可以打开一个控制功能
#               这个6位标志是：紧急标志、有意义的应答标志、推(返回丢包信息)、重置链接标志、同步序列号标志、完成发送数据标志
#           窗口大小：目的机使用16位的域告诉源主机，它想以收到的每个TCP数据段大小
#           效验和：TCP头也包括16位的错误检查域——效验和域
#           紧急：紧急指针是一个可选的16位指针，指向段内的最后一个字节位置，这个域只在URG标志设置了时才有效
#           选项：至少一个字节的可变长域标识哪个选项
#           数据
#           填充
#   TCP/IP简介
#       计算机网路的出现比互联网要早很多
#       计算机为了联网，就必须规定通信协议，早期的计算机网络，都是由各个厂商自己规定一套协议，IBM、Apple和Microsoft都有各自的网络协议，互不兼容，这就好比一群人有的说英语，有的说中文，有的说德语，说同一种语言的人可以交流，不同的语言之间就不行了
#       为了把全世界的所有不同类型的计算机都连接起来，就必须规定一套全球通用的协议，为了实现互联网这个目标，互联网协议簇（Internet Protocol Suite）就是通用协议标准。Internet是由inter和net两个单词组合起来的，原意就是连接‘网络’的网络，有了Internet,任何私有网络，只要支持这个协议，就可以联入互联网
#       因为互联网协议包含了上百种协议标准，但是最重要的两个协是TCP和IP协议，所以，大家把互联网的协议简称为TCP/IP协议
#       通信的时候，双方必须知道对方的标识，好比发邮件必须知道对方的邮件地址。互联网上每个计算机的唯一标识就是IP地址，类似123.123.123.123。如果一台计算机同时接入到两个或更多的网络，比如路由器，它就会有两个或多个IP地址，所以，IP地址对应的实际上是计算机的网络接口，通常是网卡
#       IP协议负责把数据从一台计算机通过网络发送到另一台计算机。数据被分割成一小块一小块，然后通过IP包发送出去。IP包的特点是按块发送，途径多个路由，但不保证能到达，也不保证顺序达到
#           计算机A、计算机B，路由器abc，A与B通信是通过路由器发送信息，连接方式非常多，ab、ac、bc。每条路径的长度也不同，因此分块发送的IP包常常会不按顺序到达或游荡其中不到达，时间长了就自动消失。这就叫做丢包
#       IP地址实际上就是一个32位整数（称为IPv4），以字符串表示的IP地址，如192.168.0.1实际上是把32位整数按8位分组后的数字表示，目的是便于阅读
#       IPv6地址实际上是一个128整数，它是目前使用的IPv4的升级版，以字符串表示类似于2001:Odb8:85a3:0042:1000:8a2e:0370:7334
#       TCP协议则是建立在IP协议之上的。TCP协议负责在两台计算机之间建立可靠连接，保证数据包顺序达到。TCP协议会通过握手建立连接，然后，对每个IP包编号，确保对方按顺序收到，如果包丢了，就自动重发
#       建立一个TCP连接
#           握手连接：确认走哪条路
#               客户端发送：SYN标志，ISN以及目的端口。
#               服务器返回:SYN标志、ISN及确认信息（ACK）。
#               客户端发送：确认信息（ACK）
#       断开一个TCP链接
#           服务器发送：FIN标识，停止服务器端到客户端的数据传输
#           客户端返回：确认信息（ACK）、FIN标识，停止客户端到服务器端的数据传输
#           服务器发送：确认信息（ACK）
'''
#TCP连接建立
IP包TCP包 客-服
IP包TCP包 服-客
IP包TCP包 客-服

客-服数据发送
IP包TCP包序号1 [justin]
IP包TCP包序号2 [is]
IP包TCP包序号3 [good]
IP包TCP包序号4 [man]

服-客数据发送
IP包TCP包序号1 [yes]
IP包TCP包序号2 [you]
IP包TCP包序号3 [are]
IP包TCP包序号4 [good]
IP包TCP包序号5 [man]

#TCP断开
IP包TCP包FIN 服-客
IP包TCP包ACK 客-服
IP包TCP包FIN 客-服
IP包TCP包ACK 服-客
'''

#【day130】网络编程
#客户端：创建TCP连接时，主动发起连接的叫做客户端
#服务端：接收客户端的链接

#实现通信的流程：
#服务端开启
#   socket()  建立一个socket对象
#   bind()    绑定端口号
#   listen()  设置监听个数（让几个客户端连接）
#   accept()  等待客户端连接
#客户端
#   socket()
#   connect() 连接服务器
#   send()    发送数据
#服务端
#   read()    读取数据并处理
#   send()   写数据做应答
#   close()   结束连接
#客户端
#   read()    读取数据
#   close()   结束连接

#客户端编程
'''
import socket
#1、建立socket对象，即TCP/IP包头
    #参数1：指定协议 AF_INET表示IPv4协议 AF_INET6表示IPv6
    #参数2：指定使用面向流的TCP协议
sk = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
#2、建立连接
#   参数1：元组（服务器IP地址，端口号）
sk.connect(('www.sina.com.cn', 80))
#3、发送请求
sk.send(b'GET / HTTP/1.1\r\nHost: www.sina.com.cn\r\nConnection: close\r\n\r\n')
#1、2、3步骤相当于在写检查.Network. www.sina.com.cn .Headers.General
#4、等待接收数据
from myClass.myCrawler import MyCrawler
data = []
while True:
    #每次接收1k的数据
    tempData = sk.recv(1024) #接收的是新浪的主页HTMLBytes
    if tempData:
        data.append(tempData)
    else:
        break
dataBytes = b''.join(data)
dataStr = dataBytes.decode('utf-8')
#5、断开链接
sk.close()
#获取headers和HTML
headers, HTML = dataStr.split('\r\n\r\n', 1)
print(headers)
print(HTML)
'''

#【day131】TCP协议建立客户端与服务端通信
#连个端交互数据必须要两个文件，本次知识点见 练习 ->客户端与服务端
#服务端编程：
'''
import socket
#创建
server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
#绑定IP和端口
server.bind(('192.168.1.10', 8080)) #本机ip：cmd -> ipconfig -> ipv4地址
#监听
server.listen(5)
print('服务器启动成功......')
#等待连接
clientSocket, clientAddress = server.accept()
print('{}--{} 链接成功'.format(str(clientSocket), clientAddress))
#循环等待接收客户端发送来的数据
while True:
    data = clientSocket.recv(1024) #接收数据1k数据 recv能阻塞程序，等待接收数据
    print('客户端说:{}'.format(data.decode('utf-8')))
    data = input('对客户端说：')
    clientSocket.send(data.encode('utf-8'))

'''

#客户端编程：
'''
import socket
client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
client.connect(('192.168.1.10', 8080))

while True:
    data = input('对服务器说：')
    client.send(data.encode('utf-8'))
    info = client.recv(1024).decode('utf-8') #阻塞程序进行，等待接收数据
    print('服务器说：', info)
'''

#【day132】UDP协议
#1、冒充飞Q
#2、客户端与服务端通信——实现群聊功能

#TCP是建立可靠的链接，并且通信双方都可以以流的形式发送数据。UDP则是面向无链接的协议
#使用UDP协议时不需要建立连接，只需要知道对方的IP地址和端口号就可以直接发送数据包
#但是能不能达到就不知道了
#虽然UDP传输数据不可靠，但是它的优点是速度，比TCP传输快。对于要求不高的数据可以使用UDP
#用途：视频广播
#实现流程
#服务器
#   socket()
#   bind()
#   recvfrom()
#客户端
#   socket()
#   bind()
#   sendto()
#   close()
#服务器
#   send()
#   recvfrom
#   close()
#代码见：练习->UDP->客户端与服务端通信


#【day133】多任务原理
#现代操作系统：Windows、Mac OS X、Linux等都支持多任务
#什么叫多任务？
#   操作系统同时可以运行多任务
#   早期电脑都是单核CPU， 单核CPU如何实现多任务原理
#       操作系统轮流让各个任务交替执行，任务1执行2us，切换任务2执行2us，切换任务3执行2us..，表面上看每个任务反复执行，但cpu调度执行速度太快了，导致我们感觉所有任务都在同时执行一样
#   多核心电脑的多任务实现原理
#       真正的并行执行多任务只能在多核CPU上实现，但是由于任务数量远远多于CPU的核心数量，所以操作系统也会自动的把多任务轮流调度到每个核心上执行
#           并发：看上去一起执行，任务数多于CPU核心数
#           并行：真正一起执行，任务数小于等于CPU核心数
#   实现多任务的方式
#       多进程模式       多任务执行
#       多线程模式       多任务执行
#       协程模式         使用人数不多，但有其优点
#       多进程+多线程模式 不建议使用
#进程概念
#   对于操作系统而言,一个任务就是一个进程
#   进程是操作系统中程序执行和资源分配的基本单位。
#   每个进程都有自己的数据段、代码段、堆栈段

#【day134】单任务
'''
#单任务无法执行print(mia),只有当print('justin')循环结束才能执行print('mia')
def run():
    while True:
        print('mia')
        sleep(1.2)

if __name__=='__main__':
    while True:
        print('justin')
        sleep(1)
    run()
'''
#多任务可以让两个print同时执行
#实现多任务的方法：启动多线程

#【day135】启动多进程
#multiprocessing库
#   multi 多 processing 进程
#   跨平台版本的多进程模块
#   提供Process类来代表一个进程对象
'''
from multiprocessing import Process
from time import sleep
import os
'''

#多进程实现多任务
'''
def run():
    while True:
        print('mia')
        sleep(1)

if __name__=='__main__':
    print('父进程已启动...')
    # 创建子进程
    # target指定这个子进程要执行的函数任务
    p = Process(target=run)
    p.start()
    while True:
        print('justin')
        sleep(1)
'''

#多进程传参
'''
def run(str): #子进程需要执行的代码
    while True:
        print(str)
        sleep(1)

if __name__=='__main__':
    print('父进程已启动...')
    # 创建子进程
    # target指定这个子进程要执行的函数任务
    p = Process(target=run, args=('coco',)) #args=() 元素1为函数参数1，元素2位tuple如果只有一个元素，需要在此元素后加,号
    p.start()
    while True:
        print('justin')
        sleep(1)
'''

#os.getpid()
#功能：获取当前进程Id号
#os.getppid()
#功能：在子进程中运行getppid函数，获取当前子进程的父进程id号
'''
def run():
    while True:
        #
        print('子进程ID号：{}'.format(os.getpid()))
        print('该子进程的父进程ID号：{}'.format(os.getppid()))
        sleep(1)

if __name__=='__main__':
    print('父进程已启动...')
    # 创建子进程
    # target指定这个子进程要执行的函数任务
    p = Process(target=run) #args=() 元素1为函数参数1，元素2位tuple如果只有一个元素，需要在此元素后加,号
    p.start()
    while True:
        print('父进程ID号：{}'.format(os.getpid()))
        sleep(1)
'''


#【day136】父子进程的先后顺序
'''
def pro():
    print('子进程启动')
    time.sleep(3)
    print('子进程结束')

if __name__=='__main__':
    print('父进程启动')
    p = Process(target=pro)
    p.start()
    print('父进程结束')
    #子进程不会被父进程的结束而结束
'''

#让父进程等待子进程结束在执行父进程
'''
def pro():
    print('子进程启动')
    time.sleep(3)
    print('子进程结束')

if __name__=='__main__':
    print('父进程启动')
    p = Process(target=pro)
    p.start()
    p.join() #子进程加入到父进程中
    print('父进程结束')
#这样不就又变成单进程了吗？其实不然，主要是今后在写多进程时，会写过个子进程，父进程是不干活的
'''


#【day137】全局变量在多个进程中不能共享
#父子进程不共享，兄弟进程也不共享
'''
from multiprocessing import Process
import time
#在子进程中修改全局变量对父进程中的全局变量没有影响
#因为在创建子进程时，对全局变量做了一个备份，父进程中的与子进程中的num是两个完全不同的变量
num = 100
def pro():
    print('子进程开始')
    time.sleep(1)
    global num #相当于是备份了num = 100
    num += 1
    print('子进程中的全局变量 =', num)
    print('子进程结束')

if __name__ == '__main__':
    print('父进程开始')
    p = Process(target=pro)
    p.start()
    p.join()
    print('全局变量 =', num)
    print('父进程结束')
'''

#【day138】 启动大量子进程
#创建大量子进程时，使用Proceess()创建不方便管理
#Pool 进程池就能很好的管理
#原型：Pool([int])
#参数：1、代表能同时进行的进程数量、默认本机的核心数
'''
from multiprocessing import Pool
import random, time, os
def pro(name):
    print('子进程{}启动：{}'.format(name, os.getpid()))
    start = time.time()
    time.sleep(random.choice([1,2,3]))
    end = time.time()
    print('子进程{}结束：{},耗时：{}'.format(name, os.getpid(), (end - start)))

if __name__=='__main__':
    print('父进程启动')
    processPool = Pool(8) #int，启动进程池后能同时执行的进程个数
    for i in range(9): #range数必须大于核心数才有意义
        #创建进程，放入进程池统一管理
        processPool.apply_async(pro, args=(i,))
    processPool.close()
    #进程池对象调用的join方法，会等待进程池中的所有子进程结束完毕，再执行父进程
    processPool.join() #注意：如果调用的进程池的join，必须在之前要调用该进程池的close，并且在close之后就不能添加新的进程了（在close之前可以随意添加进程到进程池）

    print('父进程结束')
'''

#多进程执行不同的函数,在核心数以内的子进程数是同时执行的，没有先后顺序
'''
from multiprocessing import Pool
def pro1(num):
    print('mia\'s {} years'.format(num))

def pro2(num):
    print('justin\'s {} years'.format(num))

if __name__ == '__main__':
    print('父开始')
    pp = Pool()
    pp.apply_async(pro1, args=(9,))
    pp.apply_async(pro2, args=(30,))
    pp.close()
    pp.join()
    print('父结束')
'''

#作业：用多进程将一个文件夹中的文件全部复制到另一个文件

#遍历复制文件，总耗时0.11s
'''
import os, time
def copyFile(path, toPath):
    with open(path, 'r') as f:
        context = f.read()
    with open(toPath, 'w') as f:
        f.write(context)
path = r'D:\qian_feng_education\first_project\test\file'
toPath = r'D:\qian_feng_education\first_project\test\toFile'
start = time.time()
fileList = os.listdir(path)
for fileName in fileList:
    copyFile(os.path.join(path, fileName), os.path.join(toPath, fileName))
end = time.time()
print('总耗时：{}'.format((end - start)))
'''


#多进程复制文件，总耗时：0.72s 拷贝文件
'''
import time, os
from multiprocessing import Pool
def copyFile(path, toPath):
    with open(path, 'r') as f:
        context = f.read()
    with open(toPath, 'w') as f:
        f.write(context)
if __name__=='__main__':
    path = r'D:\qian_feng_education\first_project\test\file'
    toPath = r'D:\qian_feng_education\first_project\test\toFile'
    start = time.time()
    pp = Pool() #创建进程很耗时，但凡处理的文件量大，跑起来之后，会比遍历快很多
    fileList = os.listdir(path)
    for fileName in fileList:
        absFilePath = os.path.join(path, fileName)
        absToPath = os.path.join(toPath, fileName)
        pp.apply_async(copyFile, args=(absFilePath, absToPath))
    pp.close()
    pp.join()
    end = time.time()
    print('总耗时%s' % (end - start))
'''


#【day139】二次封装Process
#二次封装一个进程类，见 myClass -> myProcess
#优点：1、启动一个子进程，子进程执行的方法不用和父进程写在一起，代码清晰
#     2、今后需要不同的子进程方法，只需要再封装一个对应的子进程模块，将需要的方法封装到模块里即可
'''
from myClass.myProcess.myProcesse import MyProcess
if __name__ == '__main__':
    print('启动父进程')
    #创建子进程
    p = MyProcess('justin') #这个子进程中，执行的函数需要什么参数，就在这里传入即可
    #自动调用p进程对象的testProcesse方法
    p.start()
    p.join()
    print('父进程结束')
'''


#【day139】进程间通信
'''
import os, time
from multiprocessing import Process, Queue
def write(q):
    print('启动写子进程%s' % os.getpid())
    for chr in ['A', 'B', 'C', 'D']:
        q.put(chr)
        time.sleep(1)
    print('结束写子进程%s' % os.getpid())

def read(q):
    print('启动读子进程%s' % os.getpid())
    while True: #一直读数据，因为并不知道写方法会写多少个元素
        value = q.get(True)
        print('value =', value)
    print('结束读子进程%s' % os.getpid())

if __name__=='__main__':
    #父进程创建队列，并传递给子进程（子进程引用同一个队列），实现子进程之间的通信
    q = Queue()
    pw = Process(target=write, args=(q, ))
    pr = Process(target=read, args=(q, ))
    #子进程开始
    pw.start()
    pr.start()
    #等待pw结束再执行父进程
    pw.join()
    #pr进程里是一个死循环，无法等待其结束，只能强行结束。写在pw.join()后面的意义是pw结束了，pr也结束
    pr.terminate() #强制结束进程
    print('父进程结束')
'''

#作业：将500个文件分为5组，建立5个进程同时复制这些文件到另一个文件中
'''
import os
from multiprocessing import Process, Queue
def pro1(q, path, start, end):
    fileList = os.listdir(path)
    for fileName in fileList[start:end]:
        absFileName = os.path.join(path, fileName)
        q.put(absFileName)

def pro2(q, toPath):
    while True:
        absFileName = q.get(True)
        with open(absFileName, 'r') as f:
            content = f.read()
        absToPath = os.path.join(toPath, os.path.split(absFileName)[-1])
        with open(absToPath, 'w') as f:
            f.write(content)

if __name__ == '__main__':
    path = r'D:\qian_feng_education\first_project\test\file'
    toPath = r'D:\qian_feng_education\first_project\test\toFile'
    #创建队列
    q1 = Queue()
    q2 = Queue()
    q3 = Queue()
    q4 = Queue()
    q5 = Queue()
    #创建子进程
    p1 = Process(target=pro1, args=(q1, path, 0, 100))
    p2 = Process(target=pro2, args=(q1, toPath))
    p3 = Process(target=pro1, args=(q2, path, 100, 200))
    p4 = Process(target=pro2, args=(q2, toPath))
    p5 = Process(target=pro1, args=(q3, path, 200, 300))
    p6 = Process(target=pro2, args=(q3, toPath))
    p7 = Process(target=pro1, args=(q4, path, 300, 400))
    p8 = Process(target=pro2, args=(q4, toPath))
    p9 = Process(target=pro1, args=(q5, path, 400, 500))
    p10 = Process(target=pro2, args=(q5, toPath))

    #子进程启动
    p1.start()
    p2.start()
    p3.start()
    p4.start()
    p5.start()
    p6.start()
    p7.start()
    p8.start()
    p9.start()
    p10.start()

    #进程结束
    p1.join()
    p3.join()
    p5.join()
    p7.join()
    p9.join()
    while True:
        path = r'D:\qian_feng_education\first_project\test\toFile'
        numLen = len(os.listdir(path))
        if numLen == 500:
            p2.terminate()
            p4.terminate()
            p6.terminate()
            p8.terminate()
            p10.terminate()
            break
'''

#【day140】线程
#在一个进程内部要同时做多件事，就需要同时运行多个子任务，我们把进程内的子任务叫做线程
#   例：word中即一个进程中，可以实现打字、插入、保存、打印等等功能，这些在一个进程里面实现的功能就是启动了线程实现的
#线程通常叫做轻型的进程，线程是共享内存空间并发执行的多任务，每一个线程都共享一个进程的资源。
#线程是最小的执行单元，而进程里至少有一个线程组成。如何调度进程和线程完全是由操作系统决定的，程序自己不能决定什么时候执行，也不能决定执行多长时间
#模块：1、_thread模块     低级模块——接近底层（直接封装的c语言）
#     2、threading模块   高级模块——对_thread进行了封装


#【day141】启动线程
'''
import threading
import time
a = 10
def thr():
    print('子线程{}启动'.format(threading.current_thread().name))
    #实现线程功能
    time.sleep(2)
    print('打印')
    time.sleep(2)
    print(a,'如果主线程结束，内存空间销毁了，a的值就不能被打印出来了')
    print('子线程{}结束'.format(threading.current_thread().name))

if __name__ == '__main__':
    #任何进程默认会启动一个线程，默认启动的线程称为主线程，主线程可以启动新的子线程
    #current_thread():返回当前线程的实例。它有name属性
    print('主线程{}启动'.format(threading.current_thread().name))
    #创建子线程
    t = threading.Thread(target=thr, name='thrThread')
    t.start()
    print('主线程{}结束'.format(threading.current_thread().name))
    #主线程结束了，内存空间会在一定时间内销毁，销毁后，所有变量将不能被调用，会影响整个程序的线程的执行
'''
#主线程必须等待子线程结束才结束
'''
import threading
import time
def thr(a):
    print('子线程{}启动'.format(threading.current_thread().name))
    time.sleep(2)
    print('打印')
    time.sleep(2)
    print(a)
    print('子线程{}结束'.format(threading.current_thread().name))

if __name__ == '__main__':
    print('主线程{}启动'.format(threading.current_thread().name)) #主进程的名字MainThread
    a = 100
    t = threading.Thread(target=thr, args=(a,)) #给子线程传参也用args,传的是元组，如果元组只有一个元素记得加，
    t.start()
    #等待子线程结束
    t.join()
    print('主线程{}结束'.format(threading.current_thread().name))
'''

#【day142】线程间共享数据
#多线程和多进程最大的不同在于
#   多进程中，同一个变量，各自有一份拷贝存在每个进程中，互不影响。而多线程中，所有变量都由所有线程共享
#   所以，任何一个变量都可以被任意一个线程修改，
#   因此，线程之间共享数据最大的危险在于
#       多个线程同时修改一个变量，容易把内容改乱了
'''
import threading
num = 100
def run(n):
    global num
    for i in range(1000000):
        num = num + n
        num = num - n
if __name__ == '__main__':
    t1 = threading.Thread(target=run, args=(9, ))
    t2 = threading.Thread(target=run, args=(6, ))
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    print('num =', num)
'''

#【day143】多线程锁
#两个线程同时工作，一个取20元,一个存20元，可能导致数据错乱
#   100 - 20 = 80 ， 80的值还没有赋值给余额变量时，第二个线程存20，此时余额变量为100+20=120元，但是第一个线程的变量此时才赋值，余额变成80，结果你的卡里少了20元
#线程锁解决数据混乱
#   解决思路：线程1加锁，执行完线程1解锁，当线程2遇到的是锁上的锁，程序会阻塞等待解锁后，线程2再执行函数
#锁确保了函数代码只能由一个线程从头到尾的完整执行
'''
import threading
#创建锁对象
lock = threading.Lock()
num = 100
def run(n):
    global num
    for i in range(1000):
        lock.acquire() #加锁
        try:
            num = num + n
            num = num - n
        finally: #任何线程不管能不能正确执行，都必须执行到解锁，否则其他进程都无法再进入函数，这就叫作死锁
            lock.release() #解锁
if __name__ == '__main__':
    t1 = threading.Thread(target=run, args=(9, ))
    t2 = threading.Thread(target=run, args=(6, ))
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    print('num =', num)
'''
#加了锁函数会阻止多线程的并发执行，包含锁的某段代码实际上只能以单线程模式执行，所以效率大打折扣
#由于可以存在多个锁，而且不同线程可以持有不同的锁，并试图获取其他的锁，可能造成死锁，导致多个线程挂起，只能靠操作系统强制终止，
#为了降低死锁几率，可以采用上锁解锁的简单写法

#【day143】开锁和解锁的简单写法
'''
import threading
lock = threading.Lock()
num = 100
def run(n):
    global num
    for i in range(1000):
        with lock: #自动上锁 执行代码 自动解锁
            num = num + n
            num = num - n

if __name__ == '__main__':
    t1 = threading.Thread(target=run, args=(9, ))
    t2 = threading.Thread(target=run, args=(6, ))
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    print('num =', num)
'''

#【day144】ThreadLocal
#全局变量局部化
'''
import threading
num = 100
#创建一个全局的ThreadingLocal对象
#每个线程有独立的存储空间
#每个线程对ThreadingLocal对象都可以读写，但互不影响
local = threading.local()

def run(x, n):
    x += n
    x -= n

def func(n):
    #每个线程都有一个local.x,就是线程的局部变量
    local.x = num #给local对象增加一个x属性，将全局变量赋值给局部变量
    for i in range(1000000):
        run(local.x, n)
    print('线程名：{}，线程值：{}'.format(threading.current_thread().name, local.x))

if __name__ == '__main__':
    t1 = threading.Thread(target=func, args=(9, ))
    t2 = threading.Thread(target=func, args=(6, ))
    t1.start()
    t2.start()
    t1.join()
    t2.join()
'''
#作用:为每个线程绑定一个数据库链接或http请求或用户身份信息等，这样一个线程的所有调用到的处理函数都可以非常方便的访问这些资源


'''
作业：用多线程修改客户端与服务段通信代码
#见：练习->TCP——多线程
#服务器编程 
import socket
import threading

server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server.bind(('192.168.1.2', 8080)) #本机ip：cmd -> ipconfig -> ipv4地址
server.listen(5)
print('服务器开启成功，等待客户端的链接')

def run(clientSocket, clientAddress):
    data = clientSocket.recv(1024)
    print('客户端{}说:{}'.format(clientAddress, data.decode('utf-8')))
    clientSocket.send('网页主页HTML'.encode('utf-8'))


while True:
    clientSocket, clientAddress = server.accept()
    print('{}--{} 链接成功'.format(str(clientSocket), clientAddress))
    t1 = threading.Thread(target=run, args=(clientSocket, clientAddress))
    t1.start()
    
#客户端编程：
import socket
client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
client.connect(('192.168.1.2', 8080))

while True:
    data = input('输入网址，发送get请求：')
    client.send(data.encode('utf-8'))
    info = client.recv(1024).decode('utf-8') #阻塞程序进行，等待接收数据
    print('服务器说：', info)
'''

'''
作业：写一个聊天室
见 finishedObject -> chatRoom
'''


#【day144】信号量控制线程数量
#一次只能允许有几个线程开始执行
'''
import threading, time
sem = threading.Semaphore(2)
def run():
    with sem:
        for i in range(4):
            print('当前线程名{}，遍历数{}'.format(threading.current_thread().name, i))
            time.sleep(1)
if __name__=='__main__':
    for i in range(4):
        threading.Thread(target=run).start()
'''

#凑一定线程数量才能一起执行
'''
import threading, time
bar = threading.Barrier(4)
def run():
    print('当前线程名{}，start'.format((threading.current_thread().name)))
    time.sleep(1)
    bar.wait() #等待线程数量，凑够了4个才能往下执行，即程序阻塞，够4执行
    print('当前线程名{}，end'.format((threading.current_thread().name))) #凑够了4个线程，执行这4个线程的print

if __name__=='__main__':
    for i in range(5):
        threading.Thread(target=run).start()
'''

#定时线程（延时线程）
'''
import threading
def run():
    print('justin')

if __name__=='__main__':
    print('父线程开始')
    t = threading.Timer(5, run) #延时执行
    t.start()
    t.join()
    print('父线程结束')
'''

#【day144】线程通信
'''
import threading
import time
def func():
    #创建事件对象
    event = threading.Event()
    def run():
        for i in range(5):
            #阻塞等待事件的触发
            event.wait() #阻塞只阻塞一次，但凡是接收到set的信息，就不在阻塞
            event.clear() #重置wait，使其恢复阻塞功能。
            print('justin{}'.format(i))
    t = threading.Thread(target=run).start()
    return event
e = func() #只要能把event对象传出来，就能在任何地放，使用set触发wait
#触发事件
for i in range(5):
    time.sleep(2)
    e.set() #set()能给wait()传递通行信息
'''

#【day144】生产者与消费者
#功能：线程之间的数据传递
#流程:
    #生产者:产生数据
    #将数据加入队列
    #消费者从队列中取出数据
    #消费者:处理数据
'''
import threading
from queue import Queue
import random
import time
def product(i, q):
    while True:
        data = random.randrange(100)
        print('生产者{},产生了数据{},加入队列'.format(i, data))
        q.put(data)
        time.sleep(3)
    #标记任务完成
    q.task_done()
def customer(i, q):
    while True:
        data = q.get()
        if data is None:
            break
        print('消费者{},取出数据{}'.format(i, data))
        time.sleep(3)
    q.task_done()
if __name__ == '__main__':
    #启动生产者
    q = Queue(5) #5代表队列的长度，即q只能存放5个元素，不写代表任意长度
    for i in range(4):
        threading.Thread(target=product, args=(i, q)).start()
    for i in range(4):
        threading.Thread(target=customer, args=(i, q)).start()
    #启动消费者
'''


#【day145】线程的调度
#复杂调度可以网上搜素，一般是做算法的人才能做
'''
import threading
import time
#线程条件变量
cond = threading.Condition()
def run1():
    with cond:
        for i in range(0, 10, 2):
            print(threading.current_thread().name, i)
            #time.sleep(3)
            cond.notify() #通知线程2放行
            cond.wait()  #线程1等待
def run2():
    with cond:
        for i in range(1, 10, 2):
            print(threading.current_thread().name, i)
            #time.sleep(3)
            cond.notify() #通知线程1放行
            cond.wait() #线程2等待
if __name__=='__main__':
    threading.Thread(target=run1).start()
    threading.Thread(target=run2).start()
    #不一定能使数值按顺序打印（0,1,2,3,4,5,6,7,8,9）
'''

#【day】 定时器
'''
import threading
#定义函数
def fun_timer():
    print('hello timer')   #打印输出
    global timer  #定义变量
    timer = threading.Timer(60,fun_timer)   #60秒调用一次函数
    #定时器构造函数主要有2个参数，第一个参数为时间，第二个参数为函数名
    timer.start()    #启用定时器
timer = threading.Timer(1,fun_timer)  #首次启动
timer.start()
'''

#【day146】进程vs线程
#多任务的实现原理
#   通常我们会设计Master_Worker模式，Master负责分配任务，Worker负责执行任务，因此，多任务环境下，通常是一个Master，多个Worker
#多进程
#   主进程就是Master，其他进程就是Worker
#   优点
#       稳定性高：一个子进程崩溃了，不会影响主进程和其他字进程，当然主进程挂了所有进程就全挂了，但是Master进程只负责分配任务，挂掉的概率低
#   缺点
#       创建进程的代价大
#           在Unix/Linux系统下，用fork调用还行，在windows下创建进程开销巨大，因为依赖三方库
#       操作系统能同时运行的进程数也是有限的
#           在内存和CPU的限制下，如果有几千个进程同时运行，操作系统连调度都会成为问题
#多线程
#   主线程就是Master，其他线程就是Worker
#   优点
#       多线程模式通常比多进程块一点，但是也快不到哪去
#       在windows下，多线程的效率比多进程要高
#   缺点
#       任何一个线程挂掉都可能直接造成整个进程崩溃
#           所有线程共享进程的内存，在windows上，如果一个线程执行的代码出了问题，你经常可以看到这样的提示:"该程序执行了非法操作，即将关闭"，其实往往是某个线程出了问题，但是操作系统会强制结束整个进程
#计算密集型 vs IO密集型
#   计算密集型
#       要进行大量的计算，消耗CPU资源，比如计算圆周率、对视频进行高清解码等，全靠CPU的运算能力。这种计算密集型任务虽然也可以用多任务完成，但是任务越多，花在任务切换的时间就越多，CPU执行任务的效率就越低，所以，要高效地利用CPU，计算密集型任务同时进行的数量应当等于CPU的核心数
#   IO密集型
#       涉及到网络、磁盘IO的任务都是IO密集型任务，这类任务的特点是CPU消耗很少，任务的大部分时间都在等待IO操作完成(因为IO的速度远远低于CPU和内存的速度)。对于IO密集型任务，任务越多，CPU效率越高，但也有一个限度。常见的大部分任务都是IO密集型任务，比如Web应用


#【day147】协程
#子程序/子函数（正常的子函数调用过程）：
#   在所有语言中都是层级调用，比如A函数调用B函数，在B执行的过程中，我们又可以调用C，C执行完毕返回，B执行完毕返回，最后A执行完毕，以下代码是通过栈实现的，一个线程就是执行一个子程序，子程序调用总是一个入口，一次返回，而且调用的入口是明确的
'''
def A():
    print('A_start')
    B()
    print('A_end')
def B():
    print('B_start')
    C()
    print('B_end')
def C():
    print('C_start')
    print('C_end')
A()
'''
#协程概念：
#   看上去也是子程序，但执行过程中，在子程序的内部可以中断，然后转而去执行别的子程序，不是函数调用
#协程与线程相比
#   优点：执行效率极高，因为只有一个线程，也不存在变量的冲突，在协程中共享资源不加锁，它只需要判断状态
'''
#1 2 4 5 6 3
#执行成这个结果单A中没有B的调用
#看起来A、B执行过程有点像线程，但协程的特点在于是一个线程执行
def A():
    print('1')
    print('2')
    print('3')
def B():
    print('4')
    print('5')
    print('6')
'''
#协程原理：
#  python对协程的支持是通过generator实现的
#yield函数生成器
#功能：1、返回值
#       类似return，可以对实例化对象返回值，但不结束函数，
#     2、yield以上部分代码全部生成一个生成器函数
#       用生成器函数实例化对象的send('a')方法传入值来调用生成器函数，此时的返回值可以用变量接收
'''
def run():
    print(1)
    yield 10
    print(2)
    yield 20 #多个yield，将一个函数分成多段函数，
    print(3)
    yield 30
#协程的最简单风格，控制函数的阶段执行，节约线程或进程的切换
#yield的返回值是一个生产器,包裹yield以上的代码
m = run() #创建生成器对象，有yield只返回生成器，不执行其以上的代码
print(next(m))
print(next(m))
print(next(m))
print(next(m)) #StopIteration报错
'''


#数据传输
#创建：四个生成器
#   生成器1：data = '', data
#   生成器2：x = 待send传值, print(1,x,data), data
#   生成器3：y = 待send传值, print(1,y,data), data
#   生成器4：z = 待send传值, print(1,z,data), data
'''
def run():
    #创建空变量，存储的作用，data始终为空
    data = ''
    x = yield data #x = a(send传来的值，用来定义x变量) , yiled将data = '', data生成第一个生成器
    print(1, x, data)
    y = yield data #y = b
    print(2, y, data)
    z = yield data #z = c
    print(3, z, data)
    yield data

m = run()
#启动m
print(m.send(None)) #1、data = None; 2、打印data = 'yield的返回值' = m.send('Nane')
print(m.send('a'))  #1、执行x = a和print(1, x, data); 2、打印data = 'yield的返回值' = m.send('a')
print(m.send('b'))  #1、执行y = b和print(1, y, data); 2、打印data = 'yield的返回值' = m.send('b')
print(m.send('c'))  #1、执行z = c和print(1, z, data); 2、打印data = 'yield的返回值' = m.send('c')
print('***')
'''

#【day147】协程的生产者与消费者
'''
#用queue实现
import random
import time
import queue

def production(q):
    data = random.choice([1, 2])
    q.put(data)
    print('进队%d' % data)
    time.sleep(2)

def consumer(q):
    data = q.get()
    print('出队%d' % data)
    time.sleep(2)

def run():
    data = ''
    q = yield data
    production(q)
    q = yield data
    consumer(q)
    yield data

if __name__ == '__main__':
    q = queue.Queue()
    m = run()
    m.send(None)
    m.send(q)
    print('***')
    m.send(q)
'''
'''
#不用queue实现
def production(c):
    c.send(None)
    for i in range(1, 5):
        print('生产者产生数据%d' % i)
        r = c.send(i)
        print('消费者消费了数据%s' % r)
    c.close()

def consumer():
    data = ''
    while True:
        n = yield data #
        if not n:
            return
        print('消费者消费了%s' % n)
        data = 200
c = consumer()
production(c)
'''

'''
Python 中的列表和元组有什么区别？
Python 的主要功能是什么？
python 是编程语言还是脚本语言？
Python 是一种解释性语言吗？
什么是 pep 8？
python 是如何进行内存管理的？
什么是 Python 的命名空间？
python 语言中 PYTHONPATH 是什么？
什么是 python 模块？在 Python 中命名一些常用的内置模块？
Python 中的局部变量和全局变量
python 是否区分大小写？
Python 中的类型转换
Python 中有哪些内置类型？
python 中是否需要缩进？
Python 中 array和 list 有什么区别？
Python 中的函数是什么？
init 的含义是什么？ 18.什么是 lambda 函数？
Python 中的 self 是什么？
break、continue、pass 的用法？
[:: - 1} 的作用是什么？
如何在 Python 中随机化列表中的项目？
什么是 python 迭代器？
如何在 Python 中生成随机数？
range＆xrange 有什么区别？
你如何在 python 中写注释？
什么是 pickling 和 unpickling？
python 中的生成器是什么？
你如何把字符串的第一个字母大写？
如何将字符串转换为全小写？
如何在 python 中注释多行？
Python 中的 Docstrings 是什么？
操作符 is、not 和 in 的目的是什么？
Python 中 help（）和 dir（）函数的用法是什么？
每当 Python 退出时，为什么不是所有的内存都被解除分配？
Python 中的字典（dictionary）是什么？
如何在 python 中使用三元运算符？
* args，** kwargs 是什么？我们为什么要用呢？
len（）的作用是什么？
Python 中“re”模块的 split（），sub（），subn（）这三个方法的作用是什么？
什么是负索引，为什么使用它们？
什么是 Python 包？
如何在 Python 中删除文件？
什么是 python 的内置类型例如：使用负索引取出列表的最后一个数？
NumPy 阵列在（嵌套）Python 列表中提供了哪些优势？
如何将值添加到 python array？
如何删除 python array的值？
Python 有 OOps 概念吗？
深拷贝和浅拷贝有什么区别？
如何在 Python 中实现多线程？
python 源代码必须经过编译才能进行？
什么是 Python 库？举几个例子。
Python 字符串 split() 方法？
如何在 python 中导入模块？ OOPS 面试问题
解释一下 Python 中的继承。
如何在 Python 中创建类？
什么是猴子补丁？
python 是否支持多重继承？
Python 中的多态是什么？
在 Python 中怎样定义封装？
你如何在 Python 中进行数据抽象？
python 是否使用了访问说明符？
Python中 pass 是什么？
object（）有什么作用？ 基本 Python 编码
用 Python 编写程序来执行冒泡排序算法。
用 Python 编写程序来生成 Star 三角形。
编写一个程序，用 Python 生成 Fibonacci 系列。
用 Python 编写程序来检查数字是否为素数。
用 Python 编写程序来检查序列是否是回文序列。
写一个单行，用于计算文件中大写字母的数量。即使文件太大而无法放入内存，你的代码也应该可以正常工作。
在 Python 中为数值数据集编写排序算法。
查看下面的代码，记下 A0，A1，… An 的最终值。 Python 库面试问题
解释 Flask 是什么及其好处？
Django 比 Flask 好吗？
Django，Pyramid 和 Flask 之间的差异。
讨论 Django 架构。
解释如何在 Django 中设置数据库。
举例说明如何在 Django 中编写 VIEW？
提及 Django 模板的组成部分。
在 Django 框架中解释会话的使用？
Django 中 Model 的继承方式。 Web Scraping - Python 面试问题
如何使用已知的 URL 地址本地保存图像？
如何获取任何网址或网页的 Google 缓存时限？
从 IMDb 前 250 电影页面中删除数据。只有电影名称，年份和评级字段。 数据分析 - Python 面试问题
什么是 Python 中的 map 函数？
为什么要使用 NumPy？
如何在 NumPy array 中获得最大值的索引？
你如何用 Python / NumPy 计算百分位数？
NumPy 和 SciPy 有什么区别？
如何使用 NumPy / SciPy 制作 3D 绘图/可视化？
检查给定数字n是否为2或0的幂
计算将A转换为B所需的位数
在重复元素array中查找两个非重复元素
找到具有相同设置位数的下一个较大和下一个较小的数字
给定 n 个项目的重量和值，将这些物品放入容量为 W 的背包中背包中的最大总价值。
给定一根长度为 n 英寸的杆和一系列价格，其中包含所有尺寸小于 n 的尺寸的价格。确定通过切割杆和销售件可获得的最大值。
给定两个字符串 str1 和 str2 以及可以在 str1 上执行的操作。查找所需的最小编辑数（操作）将’str1’转换为’str2’
给定 0 和 1 的二维矩阵，找到最大的广场，其中包含全部 1。
找到两者中存在的最长子序列的长度。子序列是以相同的相对顺序出现的序列，但不一定是连续的。
找到给定序列的最长子序列的长度，以便对子序列的所有元素进行排序，按顺序递增。
给定成本矩阵成本[] []和成本[] []中的位置（m，n），
将一个集合划分为两个子集，使得子集和的差异最小
给定一组非负整数和一个值和，确定是否存在给定集合的子集，其总和等于给定总和。
HackerRank 问题算法 DP
给定距离 dist，计算用 1,2 和 3 步覆盖距离的总方式
在字符板中查找所有可能的单词
广度优先搜索遍历
深度优先搜索遍历
在有向图中检测周期
检测无向图中的循环
Dijkstra的最短路径算法
在给定的边缘加权有向图中找出每对顶点之间的最短距离
图形实现
Kruskal的最小生成树算法
拓扑排序
以下哪个语句创建字典？（多选题）
其中哪一个是分区？
标识符的最大可能长度是多少？
为什么不鼓励以下划线开头的局部变量名？
以下哪项是无效声明？
以下是什么Output？
假设 list1 是[2,33,222,14,25]，什么是 list1 [-1]？
要打开文件 c：scores.txt 进行写作，我们使用
这段代码的 Output 是什么？
try-except-else 的 else 部分什么时候执行？
两个字符串是否是变位词
单链表逆置
前序中序求后序
求两棵树是否相同
求最大树深
前中后序遍历
二叉树节点
层次遍历
深度遍历
找零问题
快排
合并两个有序列表
农场里有鸡和兔子，总共有 35个脑袋和 94条腿，计算一下兔子和鸡分别有多少只？
写一个程序打印出[1、2、3]所有的排列
请编写一个程序，它接受来自控制台的字符串并以相反的顺序打印。
'''


#【day】logging
#概念：提供日志保存的三方库,即可以用来print重要信息，也可以将这些信息保存在本地持久化
#导入：import logging
#方法
#   warning(msg, *args, **kwargs)
#       功能：输出日志，格式为时间、保存位置（root）、信息。在根记录器上记录严重性为“WARNING”的消息。 如果记录器有-没有处理程序，调用basicConfig()来添加一个带有预定义的控制台处理程序
#       示例：logging.warring(item)
#       说明：item是爬虫爬取字段后定义的字典，如果爬虫没有爬到数据，item则是空字典，此时会记录warring日志到本地 
#       注意：在scrapy中使用
#   debug(msg, *args, **kwargs)
#       功能：输出日志，格式为时间、保存位置（root）、信息。在根记录器上记录严重性为“debug”的消息。 如果记录器有-没有处理程序，调用basicConfig()来添加一个带有预定义的控制台处理程序
#       示例：logging.debug(item)
#       说明：item是爬虫爬取字段后定义的字典，如果爬虫没有爬到数据，item则是空字典，此时会记录debug日志到本地 
#       注意：在scrapy中使用
#   info(msg, *args, **kwargs)
#       功能：输出日志，格式为时间、保存位置（root）、信息。在根记录器上记录严重性为“info”的消息。 如果记录器有-没有处理程序，调用basicConfig()来添加一个带有预定义的控制台处理程序
#       示例：logging.info(item)
#       说明：item是爬虫爬取字段后定义的字典，如果爬虫没有爬到数据，item则是空字典，此时会记录info日志到本地 
#       注意：在scrapy中使用
#   getLogger(__name__)
#       功能
#           类，让实例化的对象具有warning方法、debug方法、info方法
#           实例化对象调用这些方法的功能同上面三种方法
#           却别在于输出格式变成了，时间、文件名路径、item数据信息，这样就方便我们观察阅读排错
#       参数：__name__当前文件名变量
#       示例
#           logger = logging.getLogger(__name__)
#           logger.warning(item)
#           logger.debug(item)
#           logger.info(item)
#   basicConfig(lever,format,datefmt,filename)
#       功能：设置日志的输出样式
#       参数
#           lever:日志等级
#           format:日志输出格式
#           datefmt:时间
#           filename:保存位置
#       示例
#           去百度logging.basicConfig，网上有很多日志格式供参考
#           logging.basicConfig(level=logging.INFO, format='levelname:%s'......)
#           logger = logging.getLogger(__name__)
#           if __name__ == '__main__':
#               logger.info('this is a info log') --> 日志文件就保存到了在fileName定义的位置的文件中了
#       说明：在一个地放定义了log日志的格式，可以将实例对象导入到其他文件，在其他文件中再调用这个实例对象的方法，也可输出同样的格式日志
#           示例
#               其他文件
#                   from log_a import logger    #log_a是之前定义日志格式的文件的文件名，即把定义logger格式的文件看做模块，进行导入
#                   logger.info('this is a info log')

